import math
from typing import List, Dict, Any
from sqlalchemy import bindparam
from fastapi import APIRouter, Depends, HTTPException, status, Query
from sqlalchemy.orm import Session
from database import get_db2, get_db4
from schemas import *
from passlib.context import CryptContext
from jose import jwt
from datetime import datetime, timedelta, date
from auth_utils import get_current_user
from sqlalchemy import text
from urllib.parse import quote_plus
from new_outbound_dashboard import get_campaign_in_clause
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from io import BytesIO
from collections import defaultdict
import calendar
from datetime import time

router = APIRouter()




# @router.post("/cdr_report", response_model=List[CDRReportResponse])
# def get_cdr_report(request: CDRReportRequest, db: Session = Depends(get_db), db2: Session = Depends(get_db2)):
#     # Step 1: Get campaign ID from registration_master
#     campaign_query = text("SELECT campaignid FROM registration_master WHERE company_id = :company_id")
#     campaign_result = db.execute(campaign_query, {"company_id": request.company_id}).mappings().fetchone()
#
#     if not campaign_result:
#         raise HTTPException(status_code=404, detail="Company ID not found")
#
#     raw_campaign = campaign_result["campaignid"]
#     campaign_list = [c.strip().strip("'") for c in raw_campaign.split(",")]
#
#     # Step 2: Build report query
#     report_query = f"""
#         SELECT
#             t2.uniqueid,
#             SEC_TO_TIME(t6.p) AS parked_time,
#             t2.campaign_id,
#             IF(queue_seconds <= 20, 1, 0) AS call20,
#             IF(queue_seconds <= 60, 1, 0) AS call60,
#             IF(queue_seconds <= 90, 1, 0) AS call90,
#             t2.user AS agent,
#             vc.full_name,
#             t2.lead_id AS leadid,
#             RIGHT(phone_number, 10) AS phone_number,
#             DATE(call_date) AS call_date,
#             SEC_TO_TIME(queue_seconds) AS queuetime,
#             IF(queue_seconds = 0, FROM_UNIXTIME(t2.start_epoch), FROM_UNIXTIME(t2.start_epoch - queue_seconds)) AS queue_start,
#             FROM_UNIXTIME(t2.start_epoch) AS start_time,
#             FROM_UNIXTIME(t2.end_epoch) AS end_time,
#             SEC_TO_TIME(IF(t3.talk_sec IS NULL, t2.length_in_sec, t3.talk_sec)) AS call_duration,
#             IF(t3.talk_sec IS NULL, t2.length_in_sec, t3.talk_sec) AS call_duration1,
#             FROM_UNIXTIME(t2.end_epoch + TIME_TO_SEC(
#                 IF(t3.dispo_sec IS NULL, SEC_TO_TIME(0),
#                     IF(t3.sub_status = 'LOGIN' OR t3.sub_status = 'Feed' OR t3.talk_sec = t3.dispo_sec OR t3.talk_sec = 0,
#                         SEC_TO_TIME(1),
#                         IF(t3.dispo_sec > 100, SEC_TO_TIME(t3.dispo_sec - (t3.dispo_sec / 100) * 100), SEC_TO_TIME(t3.dispo_sec))
#                     )
#                 )
#             )) AS wrap_end_time,
#             IF(t3.dispo_sec IS NULL, SEC_TO_TIME(0),
#                 IF(t3.sub_status = 'LOGIN' OR t3.sub_status = 'Feed' OR t3.talk_sec = t3.dispo_sec OR t3.talk_sec = 0,
#                     SEC_TO_TIME(1),
#                     IF(t3.dispo_sec > 100, SEC_TO_TIME(t3.dispo_sec - (t3.dispo_sec / 100) * 100), SEC_TO_TIME(t3.dispo_sec))
#                 )
#             ) AS wrap_time,
#             t3.sub_status,
#             t2.status,
#             t2.term_reason,
#             t2.xfercallid
#         FROM asterisk.vicidial_closer_log t2
#         LEFT JOIN vicidial_agent_log t3 ON t2.uniqueid = t3.uniqueid AND t2.user = t3.user
#         LEFT JOIN (
#             SELECT uniqueid, SUM(parked_sec) AS p
#             FROM park_log
#             WHERE STATUS = 'GRABBED' AND DATE(parked_time) BETWEEN :from_date AND :to_date
#             GROUP BY uniqueid
#         ) t6 ON t2.uniqueid = t6.uniqueid
#         LEFT JOIN vicidial_users vc ON t2.user = vc.user
#         WHERE DATE(t2.call_date) BETWEEN :from_date AND :to_date
#         AND t2.campaign_id IN :campaign_ids
#         AND t2.lead_id IS NOT NULL
#     """
#
#     # Execute main report query
#     result = db2.execute(text(report_query), {
#         "from_date": request.from_date,
#         "to_date": request.to_date,
#         "campaign_ids": tuple(campaign_list)
#     }).mappings().fetchall()
#
#     enriched_result = []
#
#     if result:
#         # Fetch scenario tagging from call_master
#         scenario_query = text("""
#             SELECT *
#             FROM call_master cm
#             WHERE DATE(cm.calldate) BETWEEN :from_date AND :to_date
#         """)
#         scenario_data = db.execute(scenario_query, {
#             "from_date": request.from_date,
#             "to_date": request.to_date
#         }).mappings().fetchall()
#
#         scenario_map = {str(row["LeadId"]): row for row in scenario_data if row["LeadId"]}
#
#         # Enrich each row with scenario data if LeadId matches
#         for row in result:
#             lead_id = str(row.get("leadid"))
#             scenario = scenario_map.get(lead_id)
#             enriched_row = dict(row)
#
#             if scenario:
#                 enriched_row.update({
#                     "Category1": scenario.get("Category1"),
#                     "Category2": scenario.get("Category2"),
#                     "Category3": scenario.get("Category3"),
#                     "Category4": scenario.get("Category4"),
#                     "Category5": scenario.get("Category5"),
#                     "source": scenario.get("Source"),
#                     "recording": scenario.get("Recording")
#                 })
#             else:
#                 enriched_row.update({
#                     "Category1": None,
#                     "Category2": None,
#                     "Category3": None,
#                     "Category4": None,
#                     "Category5": None,
#                     "source": None,
#                     "recording": None
#                 })
#
#             enriched_result.append(enriched_row)
#
#     return enriched_result




@router.post("/cdr_report", response_model=List[CDRReportResponse])
def get_cdr_report(request: CDRReportRequest, db: Session = Depends(get_db4), db2: Session = Depends(get_db2)):

    # -----------------------------------------------------------------
    # 1. Campaign ID resolution (matches PHP logic with All + category)
    # -----------------------------------------------------------------
    category_qry = ""
    if request.category and request.category != "All":
        category_qry = f"AND client_category = '{request.category}'"

    if request.company_id == "All":
        db.execute(text("SET SESSION group_concat_max_len = 20000"))
        campaign_sql = f"""
            SELECT GROUP_CONCAT(campaignid) AS campaign_id
            FROM registration_master
            WHERE status='A' AND is_dd_client='1' {category_qry}
        """
        campaign_result = db.execute(text(campaign_sql)).mappings().fetchone()
        if not campaign_result or not campaign_result["campaign_id"]:
            raise HTTPException(status_code=404, detail="No campaigns found")
        raw_campaign = campaign_result["campaign_id"]
    else:
        campaign_sql = f"""
            SELECT campaignid
            FROM registration_master
            WHERE company_id = :company_id {category_qry}
        """
        campaign_result = db.execute(text(campaign_sql), {"company_id": request.company_id}).mappings().fetchone()
        if not campaign_result:
            raise HTTPException(status_code=404, detail="Category Not Match")
        raw_campaign = campaign_result["campaignid"]

    # Convert campaign list
    campaign_list = [c.strip().strip("'") for c in raw_campaign.split(",")]

    # -----------------------------------------------------------------
    # 2. Main CDR Query (same as PHP $qry)
    # -----------------------------------------------------------------
    report_query = f"""
        SELECT 
            t2.uniqueid,
            SEC_TO_TIME(t6.p) AS parked_time,
            t2.campaign_id,
            IF(queue_seconds <= 20, 1, 0) AS call20,
            IF(queue_seconds <= 60, 1, 0) AS call60,
            IF(queue_seconds <= 90, 1, 0) AS call90,
            t2.user AS agent,
            vc.full_name,
            t2.lead_id AS leadid,
            RIGHT(phone_number, 10) AS phone_number,
            phone_number AS full_phone_number,
            DATE(call_date) AS call_date,
            SEC_TO_TIME(queue_seconds) AS queuetime,
            IF(queue_seconds = 0, FROM_UNIXTIME(t2.start_epoch), FROM_UNIXTIME(t2.start_epoch - queue_seconds)) AS queue_start,
            FROM_UNIXTIME(t2.start_epoch) AS start_time,
            FROM_UNIXTIME(t2.end_epoch) AS end_time,
            SEC_TO_TIME(IF(t3.talk_sec IS NULL, t2.length_in_sec, t3.talk_sec)) AS call_duration,
            IF(t3.talk_sec IS NULL, t2.length_in_sec, t3.talk_sec) AS call_duration1,
            FROM_UNIXTIME(t2.end_epoch + TIME_TO_SEC(
                IF(t3.dispo_sec IS NULL, SEC_TO_TIME(0),
                    IF(t3.sub_status = 'LOGIN' OR t3.sub_status = 'Feed' OR t3.talk_sec = t3.dispo_sec OR t3.talk_sec = 0,
                        SEC_TO_TIME(1),
                        IF(t3.dispo_sec > 100, SEC_TO_TIME(t3.dispo_sec - (t3.dispo_sec / 100) * 100), SEC_TO_TIME(t3.dispo_sec))
                    )
                )
            )) AS wrap_end_time,
            IF(t3.dispo_sec IS NULL, SEC_TO_TIME(0),
                IF(t3.sub_status = 'LOGIN' OR t3.sub_status = 'Feed' OR t3.talk_sec = t3.dispo_sec OR t3.talk_sec = 0,
                    SEC_TO_TIME(1),
                    IF(t3.dispo_sec > 100, SEC_TO_TIME(t3.dispo_sec - (t3.dispo_sec / 100) * 100), SEC_TO_TIME(t3.dispo_sec))
                )
            ) AS wrap_time,
            t3.sub_status,
            t2.status,
            t2.term_reason,
            t2.xfercallid
        FROM asterisk.vicidial_closer_log t2
        LEFT JOIN vicidial_agent_log t3 
            ON t2.uniqueid = t3.uniqueid AND t2.user = t3.user
        LEFT JOIN (
            SELECT uniqueid, SUM(parked_sec) AS p 
            FROM park_log 
            WHERE STATUS = 'GRABBED' AND DATE(parked_time) BETWEEN :from_date AND :to_date 
            GROUP BY uniqueid
        ) t6 ON t2.uniqueid = t6.uniqueid
        LEFT JOIN vicidial_users vc ON t2.user = vc.user
        WHERE DATE(t2.call_date) BETWEEN :from_date AND :to_date 
        AND t2.campaign_id IN :campaign_ids 
        AND t2.lead_id IS NOT NULL
    """

    cdr_rows = db2.execute(
        text(report_query),
        {"from_date": request.from_date, "to_date": request.to_date, "campaign_ids": tuple(campaign_list)}
    ).mappings().fetchall()

    # -----------------------------------------------------------------
    # 3. Scenario/Category data (two queries, same as PHP)
    # -----------------------------------------------------------------
    scenario_query1 = text("""
        SELECT  LeadId,
    Category1,
    Category2,
    Category3,
    Category4,
    Category5,
    CallType,
    Field9 AS SourceField FROM call_master cm
        WHERE DATE(cm.calldate) BETWEEN :from_date AND :to_date
    """)
    scenario_data1 = db.execute(scenario_query1, {
        "from_date": request.from_date,
        "to_date": request.to_date
    }).mappings().fetchall()

    # Missed tagging query (hardcoded)
    scenario_query2 = text("""
        SELECT * FROM call_master cm
        WHERE DATE(cm.calldate) BETWEEN '2025-03-15' AND '2025-03-30'
    """)
    scenario_data2 = db.execute(scenario_query2).mappings().fetchall()

    # Merge into map by LeadId
    scenario_map = {}
    for row in scenario_data1 + scenario_data2:
        lead_id = str(row.get("LeadId"))
        if lead_id and lead_id not in scenario_map:
            scenario_map[lead_id] = row

    # -----------------------------------------------------------------
    # 4. Enrich rows with Category1-5
    # -----------------------------------------------------------------
    enriched_result = []
    for row in cdr_rows:
        enriched_row = dict(row)
        lead_id = str(row.get("leadid"))

        scenario = scenario_map.get(lead_id)
        if scenario:
            if request.company_id == 605:
                source_value = scenario.get("SourceField")
            else:
                source_value = "Other_client"

            enriched_row.update({
                "Category1": scenario.get("Category1"),
                "Category2": scenario.get("Category2"),
                "Category3": scenario.get("Category3"),
                "Category4": scenario.get("Category4"),
                "Category5": scenario.get("Category5"),
                "Source": source_value,
                "CallType": scenario.get("CallType"),
            })
        else:
            if request.company_id == 605:
                source_value = None
            else:
                source_value = "Other_client"
            enriched_row.update({
                "Category1": None,
                "Category2": None,
                "Category3": None,
                "Category4": None,
                "Category5": None,
                "Source": source_value,
                "CallType": None
            })

        # ✅ Generate recording link using leadid and agent
        leadid = enriched_row.get("leadid")
        agent = enriched_row.get("agent")

        # if leadid and agent:
        #     recording_link = (
        #         f"https://dialdesk.co.in/download-recording/download.php"
        #         f"?mode=DD&filename={leadid}&agent={agent}"
        #     )
        #     enriched_row["Recording"] = recording_link
        # else:
        #     enriched_row["Recording"] = None

        if leadid and agent:
            recording_link = (
                f"https://crmapi.dialdesk.in/auth/recordings/dd-html"
                f"?filename={leadid}&agent={agent}&dater={enriched_row.get('call_date')}"
            )
            enriched_row["Recording"] = recording_link
        else:
            enriched_row["Recording"] = None

        enriched_result.append(enriched_row)

    return enriched_result




# @router.post("/ob_cdr_report")
# def get_ob_cdr_report(request: OBCDRReportRequest, db: Session = Depends(get_db), db2: Session = Depends(get_db2)):
#     # Step 1: Get campaign ID from registration_master
#     campaign_query = text("SELECT campaignid FROM registration_master WHERE company_id = :company_id")
#     campaign_result = db.execute(campaign_query, {"company_id": request.company_id}).mappings().fetchone()
#
#     if not campaign_result:
#         raise HTTPException(status_code=404, detail="Company ID not found")
#
#     raw_campaign = campaign_result["campaignid"]
#     campaign_list = [c.strip().strip("'") for c in raw_campaign.split(",")]
#
#     if not campaign_list:
#         raise HTTPException(status_code=404, detail="No campaigns found for this company")
#
#     # Step 2: Prepare condition for campaign_id filtering
#     campaign_condition = "t2.campaign_id IN :campaign_ids"
#
#     # Step 3: Build query using the provided qry with parameter placeholders
#     report_query = text(f"""
#         SELECT
#             SEC_TO_TIME(t6.`p`) AS ParkedTime,
#             t2.user AS Agent,
#             t2.lead_id AS LeadId,
#             RIGHT(phone_number,10) AS PhoneNumber,
#             DATE(call_date) AS CallDate,
#             SEC_TO_TIME(queue_seconds) AS Queuetime,
#             IF(queue_seconds='0', FROM_UNIXTIME(t2.start_epoch), FROM_UNIXTIME(t2.start_epoch - queue_seconds)) AS QueueStart,
#             FROM_UNIXTIME(t2.start_epoch) AS StartTime,
#             FROM_UNIXTIME(t2.end_epoch) AS Endtime,
#             SEC_TO_TIME(IF(t3.`talk_sec` IS NULL, t2.length_in_sec, t3.`talk_sec`)) AS CallDuration,
#             IF(t3.`talk_sec` IS NULL, t2.length_in_sec, t3.`talk_sec`) AS CallDuration1,
#             FROM_UNIXTIME(t2.end_epoch + TIME_TO_SEC(
#                 IF(t3.dispo_sec IS NULL, SEC_TO_TIME(0),
#                     IF(t3.sub_status='LOGIN' OR t3.sub_status='Feed' OR t3.talk_sec=t3.dispo_sec OR t3.talk_sec=0,
#                         SEC_TO_TIME(1),
#                         IF(t3.dispo_sec>100, SEC_TO_TIME(t3.dispo_sec-(t3.dispo_sec/100)*100), SEC_TO_TIME(t3.dispo_sec))
#                     )
#                 )
#             )) AS WrapEndTime,
#             IF(t3.dispo_sec IS NULL, SEC_TO_TIME(0),
#                 IF(t3.sub_status='LOGIN' OR t3.sub_status='Feed' OR t3.talk_sec=t3.dispo_sec OR t3.talk_sec=0,
#                     SEC_TO_TIME(1),
#                     IF(t3.dispo_sec>100, SEC_TO_TIME(t3.dispo_sec-(t3.dispo_sec/100)*100), SEC_TO_TIME(t3.dispo_sec))
#                 )
#             ) AS WrapTime
#         FROM asterisk.vicidial_closer_log t2
#         LEFT JOIN vicidial_agent_log t3 ON t2.uniqueid=t3.uniqueid AND t2.user=t3.user
#         LEFT JOIN (
#             SELECT uniqueid, SUM(parked_sec) p FROM park_log
#             WHERE STATUS='GRABBED' AND DATE(parked_time) BETWEEN :from_date AND :to_date
#             GROUP BY uniqueid
#         ) t6 ON t2.uniqueid=t6.uniqueid
#         WHERE DATE(t2.call_date) BETWEEN :from_date AND :to_date
#         AND DATE(t2.call_date) BETWEEN DATE_SUB(CURDATE(), INTERVAL 3 MONTH) AND CURDATE()
#         AND {campaign_condition}
#         AND t2.lead_id IS NOT NULL
#     """)
#
#     # Step 4: Execute query and fetch results
#     result = db2.execute(report_query, {
#         "from_date": request.from_date,
#         "to_date": request.to_date,
#         "campaign_ids": tuple(campaign_list)
#     }).mappings().fetchall()
#
#     # Step 5: Return raw result without response_model
#     return [dict(row) for row in result]




@router.post("/ob_cdr_report")
def get_ob_cdr_report(
    request: OBCDRReportRequest,
    db: Session = Depends(get_db4),   # main DB
    db2: Session = Depends(get_db2)  # vicidial
):
    # Step 1: Get campaign IDs
    campaign_query = text("SELECT campaignid FROM registration_master WHERE company_id = :company_id")
    campaign_result = db.execute(campaign_query, {"company_id": request.company_id}).mappings().fetchone()
    if not campaign_result:
        raise HTTPException(status_code=404, detail="Company ID not found")

    raw_campaign = campaign_result["campaignid"]
    campaign_list = [c.strip().strip("'") for c in raw_campaign.split(",") if c.strip()]

    if not campaign_list:
        raise HTTPException(status_code=404, detail="No campaigns found for this company")

    # Step 2: Main Vicidial query (expanded)
    report_query = text("""
        SELECT 
            t2.uniqueid,
            t2.user AS Agent,
            vu.full_name,
            t2.lead_id AS LeadId,
            RIGHT(t2.phone_number,10) AS PhoneNumber,
            DATE(t2.call_date) AS CallDate,
            FROM_UNIXTIME(t2.start_epoch) AS StartTime,
            FROM_UNIXTIME(t2.end_epoch) AS Endtime,
            t2.length_in_sec AS CallDuration,
            SEC_TO_TIME(t2.length_in_sec) AS CallDurationFmt,
            IF(t2.user='VDAD','Not Connected','Connected') AS `Call Type`,
            IF(t2.list_id='998','Manual','Auto') AS DialMode,
            t3.dispo_sec AS WrapTime,
            t3.`wait_sec` AS WaitSec,
            (t3.talk_sec + t3.dispo_sec) AS ACHT,
            t3.talk_sec AS TalkSec,
            t2.status as `System Disposition`,
            t2.campaign_id as `Client Name`,
            t3.dead_sec as `Dead Time`
        FROM asterisk.vicidial_log t2
        LEFT JOIN vicidial_agent_log t3 ON t2.uniqueid=t3.uniqueid
        LEFT JOIN vicidial_users vu ON t2.user=vu.user
        WHERE DATE(t2.call_date) BETWEEN :from_date AND :to_date
        AND t2.campaign_id IN :campaign_ids
        AND t2.lead_id IS NOT NULL
    """)

    vicidial_rows = db2.execute(report_query, {
        "from_date": request.from_date,
        "to_date": request.to_date,
        "campaign_ids": tuple(campaign_list)
    }).mappings().fetchall()

    # Step 3: Sub Scenario categories
    category_query = text("""
        SELECT cm.LiveUniqueId, cm.LiveLeadId,
               cm.Category1, cm.Category2, cm.Category3, cm.Category4, cm.Category5
        FROM call_master_out cm
        LEFT JOIN ob_campaign_data ocd ON cm.DataId = ocd.id
        WHERE DATE(cm.calldate) BETWEEN :from_date AND :to_date
    """)
    category_rows = db.execute(category_query, {
        "from_date": request.from_date,
        "to_date": request.to_date
    }).mappings().fetchall()

    # Build category map (uniqueid → categories)
    category_map = {}
    for row in category_rows:
        if row["LiveUniqueId"]:
            category_map[row["LiveUniqueId"]] = row
        elif row.get("LiveLeadId"):
            category_map[row["LiveLeadId"]] = row

    # Step 4: Merge recording + categories into Vicidial data
    final_data = []
    for row in vicidial_rows:
        row_dict = dict(row)

        # Step: Override Call Type if call not connected
        if not row_dict.get("Endtime") or not row_dict.get("TalkSec"):
            row_dict["Call Type"] = "Not Connected"
        # else leave Call Type as it was from DB (Connected / Not Connected)

        # Recording URL
        row_dict["Recording"] = (
            # f"https://dialdesk.co.in/download-recording/download.php"
            # f"?mode=DD&filename={row['LeadId']}&agent={row['Agent']}"
            f"https://crmapi.dialdesk.in/auth/recordings/dd-html"
            f"?filename={row_dict['LeadId']}"
            f"&agent={row_dict['Agent']}"
            f"&dater={row_dict['CallDate']}"
        )

        # Attach Sub Scenarios (if exists)
        categories = category_map.get(row["uniqueid"], {})
        row_dict["SubScenario1"] = categories.get("Category1")
        row_dict["SubScenario2"] = categories.get("Category2")
        row_dict["SubScenario3"] = categories.get("Category3")
        row_dict["SubScenario4"] = categories.get("Category4")
        row_dict["SubScenario5"] = categories.get("Category5")

        final_data.append(row_dict)

    return final_data




# @router.post("/ob_shared_cdr_report")
# def get_ob_shared_cdr_report(request: OBCDRReportRequest, db: Session = Depends(get_db), db2: Session = Depends(get_db2)):
#     # Step 1: Get campaign ID from registration_master
#     campaign_query = text("SELECT campaignid FROM registration_master WHERE company_id = :company_id")
#     campaign_result = db.execute(campaign_query, {"company_id": request.company_id}).mappings().fetchone()
#
#     if not campaign_result:
#         raise HTTPException(status_code=404, detail="Company ID not found")
#
#     raw_campaign = campaign_result["campaignid"]
#     campaign_list = [c.strip().strip("'") for c in raw_campaign.split(",")]
#
#     if not campaign_list:
#         raise HTTPException(status_code=404, detail="No campaigns found for this company")
#
#     # Step 2: Prepare campaign filter condition
#     campaign_condition = "t2.campaign_id IN :campaign_ids"
#
#     # Step 3: Build and parameterize query
#     report_query = text(f"""
#         SELECT
#             DATE(t2.call_date) AS CallDate,
#             FROM_UNIXTIME(t2.start_epoch) AS StartTime,
#             FROM_UNIXTIME(t2.end_epoch) AS Endtime,
#             LEFT(t2.phone_number, 10) AS PhoneNumber,
#             t2.user AS Agent,
#             vu.full_name AS FullName,
#             IF(t2.user = 'VDAD', 'Not Connected', 'Connected') AS CallType,
#             t2.status AS Status,
#             IF(t2.list_id = '998', 'Manual', 'Auto') AS DialMode,
#             t2.campaign_id AS CampaignID,
#             t2.lead_id AS LeadID,
#             t2.length_in_sec AS LengthInSec,
#             SEC_TO_TIME(t2.length_in_sec) AS LengthInMin,
#             t2.term_reason AS TermReason,
#             t2.length_in_sec AS CallDuration,
#             t2.status AS CallStatus,
#             t3.pause_sec AS PauseSec,
#             t3.wait_sec AS WaitSec,
#             t3.talk_sec AS TalkSec,
#             t3.dispo_sec AS DispoSec
#         FROM asterisk.vicidial_log t2
#         LEFT JOIN vicidial_agent_log t3 ON t2.uniqueid = t3.uniqueid
#         LEFT JOIN vicidial_users vu ON t2.user = vu.user
#         WHERE DATE(t2.call_date) BETWEEN :from_date AND :to_date
#         AND {campaign_condition}
#         AND t2.lead_id IS NOT NULL
#     """)
#
#     # Step 4: Execute and fetch
#     result = db2.execute(report_query, {
#         "from_date": request.from_date,
#         "to_date": request.to_date,
#         "campaign_ids": tuple(campaign_list)
#     }).mappings().fetchall()
#
#     # Step 5: Return raw list of dicts
#     return [dict(row) for row in result]



@router.post("/ob_shared_cdr_report")
def get_ob_shared_cdr_report(
    request: OBCDRReportRequest,
    db: Session = Depends(get_db4),   # main DB
    db2: Session = Depends(get_db2)   # vicidial DB
):
    from_dt = request.from_date
    to_dt = request.to_date

    # Step 1: Get GroupId (or fallback to campaignid)
    client_query = text("""
        SELECT campaignid, GroupId
        FROM registration_master
        WHERE company_id = :company_id
    """)
    client_row = db.execute(client_query, {"company_id": request.company_id}).mappings().first()

    # Step 1️⃣: Use GroupId if available, otherwise campaignid
    campaign_ids_str = client_row["GroupId"] or client_row["campaignid"]

    # Step 2️⃣: Convert coming data "'Fortum','Fortum_O'" → ['Fortum', 'Fortum_O']
    # Removes quotes and splits by comma safely
    campaign_ids = [c.strip().strip("'") for c in campaign_ids_str.split(",") if c.strip()]

    # Step 3️⃣: CDR Query (now uses IN :campaign_ids)
    cdr_query = text("""
        SELECT 
            DATE(t2.call_date) AS call_date,
            FROM_UNIXTIME(t2.start_epoch) AS start_time,
            FROM_UNIXTIME(t2.end_epoch) AS end_time,
            LEFT(t2.phone_number,10) AS phone_number,
            t2.user AS agent_id,
            t4.full_name AS agent_name,
            IF(mcl.id IS NOT NULL, 'Connected', 'Not Connected') AS call_type,
            t2.status AS call_status,
            IF(t2.list_id='998','Mannual','Auto') AS dial_mode,
            t2.uniqueid,
            t2.lead_id,
            t3.talk_sec,
            t3.wait_sec,
            t3.pause_sec,
            t3.dispo_sec,
            t2.term_reason
        FROM vicidial_log t2
        LEFT JOIN vicidial_agent_log t3 ON t2.uniqueid = t3.uniqueid
        LEFT JOIN vicidial_users t4 ON t2.user = t4.user
        LEFT JOIN asterisk.manual_call_log mcl ON RIGHT(mcl.phone_number,10) = RIGHT(t2.phone_number,10) AND mcl.uniqueid = t2.uniqueid
        WHERE t2.campaign_id IN ('dialdesk','Cryst002','Ajmal000','Superher')
          AND DATE(t2.call_date) BETWEEN :from_dt AND :to_dt
          AND t2.list_id in ('998','2001') 
          AND t2.lead_id IS NOT NULL
        ORDER BY 
          start_time ASC
    """)

    cdr_rows = db2.execute(
        cdr_query,
        {
            "campaign_ids": tuple(campaign_ids),
            "from_dt": from_dt,
            "to_dt": to_dt
        }
    ).mappings().all()

    # Step 3: Abandoned call master lookup (for ClientName)
    aband_query = text("""
        SELECT PhoneNo, DATE(Callbackdate) AS CallDate, CompanyName
        FROM aband_call_master
        WHERE ClientId = :client_id
          AND DATE(Callbackdate) BETWEEN :from_dt AND :to_dt
    """)
    aband_rows = db.execute(aband_query, {
        "client_id": request.company_id,
        "from_dt": from_dt,
        "to_dt": to_dt
    }).mappings().all()
    aband_map = {
        (str(row["PhoneNo"])[-10:], str(row["CallDate"])): row["CompanyName"]
        for row in aband_rows
    }

    # Step 4: Call master lookup (for SubScenarios)
    call_master_query = text("""
        SELECT LeadId, Category1, Category2, Category3, Category4, Category5
        FROM call_master
        WHERE ClientId = :client_id
          AND DATE(CallDate) BETWEEN :from_dt AND :to_dt
    """)
    call_master_rows = db.execute(call_master_query, {
        "client_id": request.company_id,
        "from_dt": from_dt,
        "to_dt": to_dt
    }).mappings().all()
    call_master_map = {row["LeadId"]: row for row in call_master_rows}

    # Step 5: Format output
    response_data = []

    for row in cdr_rows:
        phone = str(row["phone_number"])[-10:]
        call_date = str(row["call_date"])

        client_name = aband_map.get((phone, call_date))
        if not client_name:
            continue

        lead_id = row["lead_id"]
        cm = call_master_map.get(lead_id)

        # # Determine CallType with PHP-style logic
        # if not row.get("end_time") or not row.get("talk_sec"):
        #     call_type = "Not Connected"
        # else:
        #     call_type = row.get("call_type", "Connected")  # fallback to DB value


        response_data.append({
            "CallDate": row["call_date"],
            "StartTime": row["start_time"],
            "Endtime": row["end_time"],
            "CustomerNumber": phone,
            "AgentID": row["agent_id"],
            "AgentName": row["agent_name"],
            "CallType": row["call_type"],
            "SystemDisposition": row["call_status"],
            "DialingMode": row["dial_mode"],
            "ClientName": client_name,  # ✅ guaranteed present
            "LeadID": lead_id,
            "ACHT": (row["talk_sec"] or 0) + (row["dispo_sec"] or 0),
            "TalkTime": row["talk_sec"],
            "WaitTime": row["wait_sec"],
            "PauseTime": row["pause_sec"],
            "DispoTime": row["dispo_sec"],
            "DisconnectedBy": row["term_reason"],
            "Scenario": cm["Category1"] if cm else None,
            "SubScenario1": cm["Category2"] if cm else None,
            "SubScenario2": cm["Category3"] if cm else None,
            "SubScenario3": cm["Category4"] if cm else None,
            "SubScenario4": cm["Category5"] if cm else None,
            "Recording": (
                # "https://dialdesk.co.in/download-recording/download.php"
                # f"?mode=DD&filename={lead_id}&agent={row['agent_id']}"
                f"https://crmapi.dialdesk.in/auth/recordings/dd-html"
                f"?filename={lead_id}&agent={row['agent_id']}&dater={row['call_date']}"
            )
        })

    return {"status": "success", "data": response_data}





# @router.post("/ivr_report")
# def get_ivr_report(
#     request: OBCDRReportRequest,  # reuse your schema expecting from_date, to_date, company_id
#     db: Session = Depends(get_db2)
# ):
#     """
#     Returns IVR report for the requested company_id and date range.
#     """
#
#     # Step 1: Build the query with clear field mapping
#     report_query = text("""
#         SELECT
#             il.*,
#             DATE_FORMAT(start_time, '%d-%b-%y') AS Dater,
#             DATE_FORMAT(start_time, '%d-%b-%y %H:%i:%s') AS StartDate,
#             DATE_FORMAT(end_time, '%d-%b-%y %H:%i:%s') AS EndDate
#         FROM ivr_log il
#         WHERE client_id = :client_id
#         AND DATE(start_time) BETWEEN :from_date AND :to_date
#     """)
#
#     # Step 2: Execute and fetch
#     result = db.execute(report_query, {
#         "client_id": request.company_id,
#         "from_date": request.from_date,
#         "to_date": request.to_date
#     }).mappings().fetchall()
#
#     # Step 3: Return raw list of dicts
#     return [dict(row) for row in result]





@router.post("/ivr_report")
def get_ivr_report(
    request: OBCDRReportRequest,
    db4: Session = Depends(get_db4)
) -> List[Dict[str, Any]]:
    """
    Returns IVR report formatted like the PHP ivr_log report.
    """

    query = text("""
        SELECT
            il.call_type,
            il.from_source,
            il.duration,
            il.outcome,
            il.opt,
            DATE_FORMAT(il.start_time, '%d-%b-%y') AS date,
            DATE_FORMAT(il.start_time, '%d-%b-%y %H:%i:%s') AS start_time,
            DATE_FORMAT(il.end_time, '%d-%b-%y %H:%i:%s') AS end_time
        FROM ivr_log il
        WHERE il.client_id = :client_id
        AND DATE(il.start_time) BETWEEN :from_date AND :to_date
        ORDER BY il.start_time ASC
    """)

    rows = db4.execute(query, {
        "client_id": request.company_id,
        "from_date": request.from_date,
        "to_date": request.to_date
    }).mappings().all()

    result = []
    for row in rows:
        r = dict(row)

        # FROM = last 10 digits of from_source
        from_number = str(r.get("from_source", "")).strip()
        r["from"] = from_number[-10:] if len(from_number) >= 10 else from_number
        del r["from_source"]

        # Default outcome
        if not r.get("outcome"):
            r["outcome"] = "No Input"

        # Rename keys to match report table
        formatted = {
            "Date": r["date"],
            "Call Type": r["call_type"],
            "From": r["from"],
            "Start Time": r["start_time"],
            "End Time": r["end_time"],
            "Duration(Sec.)": r["duration"],
            "Outcome": r["outcome"],
            "Option Chosen": r.get("opt", "")
        }
        result.append(formatted)

    return result



# @router.post("/ivr_funnel_report")
# def get_ivr_funnel_report(
#     request: IVRFunnelReportRequest,
#     db: Session = Depends(get_db2)
# ):
#     try:
#         # Core Query
#         report_query = text("""
#             SELECT
#                 uniqueid,
#                 user,
#                 status,
#                 xfercallid,
#                 DATE_FORMAT(call_date, '%Y-%m-%d') as call_date
#             FROM vicidial_closer_log vcl
#             WHERE DATE(call_date) BETWEEN :from_date AND :to_date
#         """)
#
#         results = db.execute(report_query, {
#             "from_date": request.from_date,
#             "to_date": request.to_date,
#         }).mappings().fetchall()
#
#         return [dict(row) for row in results]
#
#     except Exception as e:
#         raise HTTPException(status_code=500, detail=str(e))



@router.post("/ivr_funnel_report")
def get_ivr_funnel_report(
    request: IVRFunnelReportRequest,
    db2: Session = Depends(get_db4),
    db: Session = Depends(get_db2)
):
    try:
        # Step 1 – fetch closer log
        qry_closer_log = text("""
            SELECT uniqueid, user, status, xfercallid
            FROM vicidial_closer_log vcl
            WHERE DATE(call_date) BETWEEN :from_date AND :to_date
        """)
        cdr_results = db.execute(qry_closer_log, {
            "from_date": request.from_date,
            "to_date": request.to_date,
        }).mappings().fetchall()
        cdr_map = {row["uniqueid"]: row for row in cdr_results}

        # Step 2 – fetch IVR log
        qry_ivr = text("""
            SELECT ivl.uniqueid, ivl.outcome,
                   DATE_FORMAT(ivl.start_time, '%b-%y') AS month
            FROM ivr_log ivl
            WHERE ivl.client_id = :client_id
              AND DATE(ivl.start_time) BETWEEN :from_date AND :to_date
        """)
        ivr_results = db2.execute(qry_ivr, {
            "client_id": request.company_id,
            "from_date": request.from_date,
            "to_date": request.to_date,
        }).mappings().fetchall()

        # Step 3 – process monthly counters
        record_list = {}

        def init_month():
            return {
                "Total IVR calls": 0,
                "Total IVR Closed Calls": 0,
                "Transferred from IVR to queue": 0,
                "Customer dropped calls before queue": 0,
                "Customer dropped calls in queue": 0,
                "Customers abandoned in queue": 0,
                "Customers transferred/connected to agent": 0,
                "Customers transferred from agent to CSAT IVR": 0,
            }
        
        for row in ivr_results:
            month = row["month"]
            outcome = row["outcome"]
            uid = row["uniqueid"]
            cdr = cdr_map.get(uid)

            if month not in record_list:
                record_list[month] = init_month()

            rec = record_list[month]
            rec["Total IVR calls"] += 1
            

            # Closed vs Forward
            if outcome != "TransferredtoAgent":
                rec["Total IVR Closed Calls"] += 1
            else:
                rec["Transferred from IVR to queue"] += 1

            # Dropped before queue
            if outcome == "TransferredtoAgent" and not cdr:
                rec["Customer dropped calls before queue"] += 1

            # Queue outcomes (same logic as PHP)
            if cdr:
                status = cdr.get("status")
                user = str(cdr.get("user", "")).lower()

                if status == "TIMEOT":
                    rec["Customers abandoned in queue"] += 1
                elif user == "vdcl":
                    rec["Customer dropped calls in queue"] += 1
                elif user and user != "vdcl":
                    rec["Customers transferred/connected to agent"] += 1
        # -----------------------------
        # Step 4 – Separate CSAT Query (IMPORTANT FIX)
        # -----------------------------
        qry_csat = text("""
            SELECT DATE_FORMAT(call_date, '%b-%y') AS month,
                   COUNT(*) AS total
            FROM vicidial_closer_log
            WHERE DATE(call_date) BETWEEN :from_date AND :to_date
              AND xfercallid != 0
            GROUP BY month
        """)

        csat_results = db.execute(qry_csat, {
            "from_date": request.from_date,
            "to_date": request.to_date,
        }).mappings().fetchall()

        for row in csat_results:
            month = row["month"]

            if month not in record_list:
                record_list[month] = init_month()

            record_list[month]["Customers transferred from agent to CSAT IVR"] = row["total"]

        # -----------------------------
        # Step 5 – Return Sorted Result
        # -----------------------------
        return [
            {
                "Month": month,
                "Total IVR calls": rec["Total IVR calls"],
                "Total IVR Closed Calls": rec["Total IVR Closed Calls"],
                "Transferred from IVR to queue": rec["Transferred from IVR to queue"],
                "Customer dropped calls before queue": rec["Customer dropped calls before queue"],
                "Customer dropped calls in queue": rec["Customer dropped calls in queue"],
                "Customers abandoned in queue": rec["Customers abandoned in queue"],
                "Customers transferred/connected to agent": rec["Customers transferred/connected to agent"],
                "Customers transferred from agent to CSAT IVR": rec["Customers transferred from agent to CSAT IVR"],
            }
            for month, rec in sorted(record_list.items())
        ]

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))







@router.get("/after-hours-calls")
def get_after_hours_calls(
    client_id: int = Query(..., description="Client ID"),
    start_date: str = Query(..., description="YYYY-MM-DD"),
    db: Session = Depends(get_db2),     # call_log DB
    db4: Session = Depends(get_db4)     # did_master DB
):

    # ------------------------------------------------
    # 1. Fetch ALL DID numbers for the client
    # ------------------------------------------------
    did_query = text("""
        SELECT did_number
        FROM did_master
        WHERE client_id = :client_id
    """)

    did_rows = db4.execute(
        did_query,
        {"client_id": client_id}
    ).mappings().all()

    if not did_rows:
        raise HTTPException(
            status_code=404,
            detail="No DIDs found for given client_id"
        )

    did_numbers = [row["did_number"] for row in did_rows]

    # ------------------------------------------------
    # 2. After-hours time windows (same as PHP)
    # ------------------------------------------------
    data1 = f"{start_date} 20:01:00"
    data2 = f"{start_date} 23:59:59"
    data3 = f"{start_date} 00:01:00"
    data4 = f"{start_date} 07:59:59"

    # ------------------------------------------------
    # 3. After-hours calls for ALL DIDs
    # ------------------------------------------------
    query = text("""
        SELECT caller_code, start_time
        FROM call_log
        WHERE number_dialed IN :did_numbers
        AND start_time BETWEEN :data1 AND :data2

        UNION

        SELECT caller_code, start_time
        FROM call_log
        WHERE number_dialed IN :did_numbers
        AND start_time BETWEEN :data3 AND :data4
        ORDER BY start_time
    """)

    rows = db.execute(
        query,
        {
            "did_numbers": tuple(did_numbers),  # IMPORTANT
            "data1": data1,
            "data2": data2,
            "data3": data3,
            "data4": data4
        }
    ).mappings().all()

    # ------------------------------------------------
    # 4. Response
    # ------------------------------------------------
    return {
        "client_id": client_id,
        "did_numbers": did_numbers,
        "date": start_date,
        "total_calls": len(rows),
        "data": [
            {
                "start_time": row["start_time"],
                "caller_code": row["caller_code"]
            }
            for row in rows
        ]
    }


# @router.post("/rl_internal_report")
# def rl_internal_report(
#     from_date: date,
#     to_date: date,
#     report_type: str = "company",
#     company_id: Optional[int] = None,
#     db: Session = Depends(get_db4),
#     db2: Session = Depends(get_db2)
# ):

#     # ---------------- VALIDATE REPORT TYPE ----------------
#     if report_type not in ["company", "entry"]:
#         raise HTTPException(400, "report_type must be 'company' or 'entry'")

#     # ---------------- GET CDR DATA ----------------
#     cdr_rows = db2.execute(text("""
#         SELECT
#             DATE(t2.call_date) AS call_date,
#             RIGHT(t2.phone_number,10) AS phone_number,
#             IF(mcl.id IS NOT NULL,'Connected','Not Connected') AS call_type
#         FROM vicidial_log t2
#         LEFT JOIN asterisk.manual_call_log mcl
#             ON RIGHT(mcl.phone_number,10)=RIGHT(t2.phone_number,10)
#             AND mcl.uniqueid=t2.uniqueid
#         WHERE DATE(t2.call_date) BETWEEN :from_dt AND :to_dt
#           AND t2.campaign_id IN ('dialdesk','Cryst002','Ajmal000','Superher')
#           AND t2.list_id IN ('998','2001')
#           AND t2.lead_id IS NOT NULL
#     """), {
#         "from_dt": from_date,
#         "to_dt": to_date
#     }).mappings().all()

#     # ---------------- GET ABANDON DATA ----------------
#     # if company_id passed → filter client
#     if company_id:
#         aband_query = text("""
#             SELECT
#                 RIGHT(PhoneNo,10) AS PhoneNo,
#                 DATE(EntryDate) AS EntryDate,
#                 DATE(CallDate) AS CallDate,
#                 CompanyName,
#                 Callbackdate
#             FROM aband_call_master
#             WHERE DATE(CallDate) BETWEEN :from_dt AND :to_dt
#               AND ClientId = :company_id
#         """)
#         aband_params = {
#             "from_dt": from_date,
#             "to_dt": to_date,
#             "company_id": company_id
#         }
#     else:
#         # all clients (current behavior)
#         aband_query = text("""
#             SELECT
#                 RIGHT(PhoneNo,10) AS PhoneNo,
#                 DATE(EntryDate) AS EntryDate,
#                 DATE(CallDate) AS CallDate,
#                 CompanyName,
#                 Callbackdate
#             FROM aband_call_master
#             WHERE DATE(CallDate) BETWEEN :from_dt AND :to_dt
#         """)
#         aband_params = {
#             "from_dt": from_date,
#             "to_dt": to_date
#         }

#     aband_rows = db.execute(aband_query, aband_params).mappings().all()

#     # ---------------- BUILD LOOKUPS ----------------
#     stats_map = {}
#     abandon_lookup = {}

#     for row in aband_rows:
#         phone = row["PhoneNo"]
#         entry_date = str(row["EntryDate"])
#         call_date = str(row["CallDate"])
#         company = row["CompanyName"]
#         callback = row["Callbackdate"]

#         # grouping key
#         group_key = company if report_type == "company" else entry_date

#         abandon_lookup[(phone, call_date)] = group_key

#         if group_key not in stats_map:
#             stats_map[group_key] = {
#                 "TotalAbandon": 0,
#                 "Callback": 0,
#                 "FailedAttempt": 0,
#                 "Connected": 0,
#                 "NotConnected": 0,
#                 "UniquePhones": set()
#             }

#         stats_map[group_key]["TotalAbandon"] += 1

#         if callback:
#             stats_map[group_key]["Callback"] += 1
#         else:
#             stats_map[group_key]["FailedAttempt"] += 1

#     # ---------------- APPLY CDR LOGIC ----------------
#     for row in cdr_rows:
#         phone = row["phone_number"]
#         call_date = str(row["call_date"])
#         call_type = row["call_type"]

#         key = (phone, call_date)

#         if key not in abandon_lookup:
#             continue

#         group_key = abandon_lookup[key]
#         stats = stats_map[group_key]

#         stats["UniquePhones"].add(phone)

#         if call_type == "Connected":
#             stats["Connected"] += 1
#         else:
#             stats["NotConnected"] += 1

#     # ---------------- FORMAT RESPONSE (OLD FRONTEND FORMAT) ----------------
#     result = []

#     for group, stats in stats_map.items():
#         row = {
#             ("CompanyName" if report_type == "company" else "EntryDate"): group,
#             "Total_Abandon": stats["TotalAbandon"],
#             "Abandon_Unique": len(stats["UniquePhones"]),
#             "Total_Callback": stats["Callback"],
#             "Connected": stats["Connected"],
#             "Not_Connected": stats["NotConnected"],
#             "Failed_Attempt": stats["FailedAttempt"]
#         }
#         result.append(row)

#     return result






@router.post("/rl_internal_report")
def rl_internal_report(
    from_date: date,
    to_date: date,
    report_type: str = "company",
    company_id: Optional[int] = None,
    db: Session = Depends(get_db4),
    db2: Session = Depends(get_db2)
):

    if report_type not in ["company", "entry"]:
        raise HTTPException(400, "report_type must be 'company' or 'entry'")

    # ---------------- GET ABANDON MASTER ----------------
    if company_id:
        aband_query = text("""
            SELECT
                RIGHT(PhoneNo,10) AS PhoneNo,
                DATE(EntryDate) AS EntryDate,
                DATE(CallDate) AS CallDate,
                CompanyName,
                Callbackdate,
                call_status
            FROM aband_call_master
            WHERE DATE(CallDate) BETWEEN :from_dt AND :to_dt
              AND ClientId = :company_id
            ORDER BY CompanyName ASC
        """)
        aband_params = {
            "from_dt": from_date,
            "to_dt": to_date,
            "company_id": company_id
        }
    else:
        aband_query = text("""
            SELECT
                RIGHT(PhoneNo,10) AS PhoneNo,
                DATE(EntryDate) AS EntryDate,
                DATE(CallDate) AS CallDate,
                CompanyName,
                Callbackdate,
                call_status
            FROM aband_call_master
            WHERE DATE(CallDate) BETWEEN :from_dt AND :to_dt
            ORDER BY CompanyName ASC
        """)
        aband_params = {
            "from_dt": from_date,
            "to_dt": to_date
        }

    aband_rows = db.execute(aband_query, aband_params).mappings().all()

    # # ---------------- GET TOTAL ABANDON FROM CLOSER LOG ----------------
    # closer_rows = db2.execute(text("""
    #     SELECT 
    #         DATE(t2.call_date) AS call_date,
    #         COUNT(t2.phone_number) AS total_abandon
    #     FROM asterisk.vicidial_closer_log t2
    #     LEFT JOIN vicidial_agent_log t3 
    #         ON t2.uniqueid = t3.uniqueid AND t2.user = t3.user
    #     LEFT JOIN (
    #         SELECT uniqueid, SUM(parked_sec) AS p 
    #         FROM park_log 
    #         WHERE STATUS = 'GRABBED'
    #           AND DATE(parked_time) BETWEEN :from_dt AND :to_dt
    #         GROUP BY uniqueid
    #     ) t6 ON t2.uniqueid = t6.uniqueid
    #     LEFT JOIN vicidial_users vc ON t2.user = vc.user
    #     WHERE DATE(t2.call_date) BETWEEN :from_dt AND :to_dt
    #       AND t2.campaign_id IN ('Karaulidiagnostics0001','Karau0001')
    #       AND t2.lead_id IS NOT NULL
    #       AND t2.user = 'VDCL'
    #     GROUP BY DATE(t2.call_date)
    # """), {
    #     "from_dt": from_date,
    #     "to_dt": to_date
    # }).mappings().all()

    # # Map: date → total_abandon
    # closer_map = {str(r["call_date"]): r["total_abandon"] for r in closer_rows}

    # ---------------- INIT ----------------
    stats_map = {}
    abandon_lookup = {}

    # ---------------- PROCESS ABAND MASTER ----------------
    for row in aband_rows:
        phone = row["PhoneNo"]
        entry_date = str(row["EntryDate"])
        call_date = str(row["CallDate"])
        company = row["CompanyName"]
        callback = row["Callbackdate"]

        group_key = company if report_type == "company" else entry_date

        abandon_lookup[(phone, call_date)] = group_key

        if group_key not in stats_map:
            stats_map[group_key] = {
                "TotalAbandon": 0,
                "Callback": 0,
                "FailedAttempt": 0,
                "Connected": 0,
                "NotConnected": 0,
                "AbandonUnique": 0,
                "Done": 0
            }

        # ✅ SIMPLE COUNT
        stats_map[group_key]["AbandonUnique"] += 1

        call_status = row["call_status"]

        # count DONE separately
        if call_status == "DONE":
            stats_map[group_key]["Done"] += 1


        if callback:
            stats_map[group_key]["Callback"] += 1
        else:
            # ✅ only count failed when NOT DONE
            if not call_status or call_status != "DONE":
                stats_map[group_key]["FailedAttempt"] += 1

    # # ---------------- APPLY TOTAL ABANDON FROM SQL ----------------
    # for (phone, call_date), group_key in abandon_lookup.items():
    #     if call_date in closer_map:
    #         stats_map[group_key]["TotalAbandon"] = closer_map[call_date]

    # ---------------- GET CDR DATA ----------------
    cdr_rows = db2.execute(text("""
        SELECT
            DATE(t2.call_date) AS call_date,
            RIGHT(t2.phone_number,10) AS phone_number,
            IF(mcl.id IS NOT NULL,'Connected','Not Connected') AS call_type
        FROM vicidial_log t2
        LEFT JOIN asterisk.manual_call_log mcl
            ON RIGHT(mcl.phone_number,10)=RIGHT(t2.phone_number,10)
            AND mcl.uniqueid=t2.uniqueid
        WHERE DATE(t2.call_date) BETWEEN :from_dt AND :to_dt
          AND t2.campaign_id IN ('dialdesk','Cryst002','Ajmal000','Superher')
          AND t2.list_id IN ('998','2001')
          AND t2.lead_id IS NOT NULL
    """), {
        "from_dt": from_date,
        "to_dt": to_date
    }).mappings().all()

    # ---------------- APPLY CDR ----------------
    for row in cdr_rows:
        phone = row["phone_number"]
        call_date = str(row["call_date"])
        call_type = row["call_type"]

        key = (phone, call_date)

        if key not in abandon_lookup:
            continue

        group_key = abandon_lookup[key]
        stats = stats_map[group_key]

        if call_type == "Connected":
            stats["Connected"] += 1
        else:
            stats["NotConnected"] += 1

    # ---------------- FINAL RESPONSE ----------------
    result = []

    for group, stats in stats_map.items():
        row = {
            ("CompanyName" if report_type == "company" else "EntryDate"): group,
            "Total_Abandon": stats["AbandonUnique"],
            "Abandon_Unique": stats["AbandonUnique"],
            "Total_Callback": stats["Callback"],
            "Connected": stats["Connected"],
            "Not_Connected": stats["NotConnected"],
            "Failed_Attempt": stats["FailedAttempt"],            
            # "Total_Done": stats["Done"], 
        }
        result.append(row)

    return result



@router.post("/company_consumption_month")
def company_consumption_month(
    request: dict,
    db: Session = Depends(get_db4)
):
    company_id = request.get("company_id")
    year = request.get("year")
    month = request.get("month")
    type_filter = request.get("type")

    if not year or not month:
        raise HTTPException(status_code=400, detail="year and month required")

    # ✅ DATE RANGE
    start_date = f"{year}-{str(month).zfill(2)}-01"
    last_day = calendar.monthrange(int(year), int(month))[1]
    end_date = f"{year}-{str(month).zfill(2)}-{last_day}"

    # --------------------------------------------------
    # 1️⃣ COMPANY QUERY (WITH TYPE FILTER)
    # --------------------------------------------------
    base_query = """
        SELECT company_id, campaignid, company_name, is_shared
        FROM registration_master
        WHERE status = 'A'
    """

    params = {}

    # ✅ COMPANY FILTER
    if company_id and company_id != "ALL":
        base_query += " AND company_id = :company_id"
        params["company_id"] = company_id

    # ✅ TYPE FILTER
    if type_filter in ["0", "1"]:
        base_query += " AND is_shared = :is_shared"
        params["is_shared"] = int(type_filter)

    # ✅ SORTING
    base_query += " ORDER BY company_name ASC"

    companies = db.execute(text(base_query), params).mappings().fetchall()

    final_result = []

    # --------------------------------------------------
    # 2️⃣ LOOP
    # --------------------------------------------------
    for comp in companies:
        comp_id = comp["company_id"]
        company_name = comp["company_name"]

        # ✅ TYPE LABEL
        company_type = "Shared" if str(comp.get("is_shared")) == "1" else "Dedicated"

        # PLAN
        bal_row = db.execute(text("""
            SELECT PlanId FROM balance_master
            WHERE clientId = :cid LIMIT 1
        """), {"cid": comp_id}).mappings().fetchone()

        if not bal_row:
            continue

        plan_row = db.execute(text("""
            SELECT rate_per_pulse_day_shift,rate_per_pulse_night_shift,pulse_night_shift,pulse_day_shift,
            pulse_outbound_call_shift,
            rate_per_pulse_outbound_call_shift
            FROM plan_master WHERE Id = :pid
        """), {"pid": bal_row["PlanId"]}).mappings().fetchone()

        day_rate = float(plan_row.get("rate_per_pulse_day_shift") or 0)
        night_rate = float(plan_row.get("rate_per_pulse_night_shift") or 0)
        ab_rate = float(plan_row.get("rate_per_pulse_outbound_call_shift") or 0)

        pulse_night_shift = float(plan_row.get("pulse_night_shift") or 0)
        pulse_day_shift = float(plan_row.get("pulse_day_shift") or 0)
        pulse_ab_shift = float(plan_row.get("pulse_outbound_call_shift") or 0)

        # --------------------------------------------------
        # 3️⃣ CONSUME DATA
        # --------------------------------------------------
        consume = db.execute(text("""
            SELECT
                COALESCE(SUM(ib_total + ibn_total + ob_total), 0) AS total_consume,
                COALESCE(SUM(ib_pulse), 0) AS ib_talktime,
                COALESCE(SUM(ibn_pulse), 0) AS ibn_talktime,
                COALESCE(SUM(ob_pulse), 0) AS ob_talktime

            FROM billing_consume_daily_new
            WHERE client_id = :cid
            AND DATE(cm_date) BETWEEN :start_date AND :end_date
        """), {
            "cid": comp_id,
            "start_date": start_date,
            "end_date": end_date
        }).fetchone()

        final_result.append({
            "Company_Name": company_name,
            "Company_Type": company_type,
            "Total_Consume": round(float(consume.total_consume or 0), 2),
            "IB_Talk_Minutes": float(consume.ib_talktime or 0),
            "IBN_Talk_Minutes": float(consume.ibn_talktime or 0),
            "OB_Talk_Minutes": float(consume.ob_talktime or 0),

            "day_Rate": f"{day_rate} Rs/ {pulse_day_shift} Sec.",
            "night_Rate": f"{night_rate} Rs/ {pulse_night_shift} Sec.",
            "ab_Rate": f"{ab_rate} Rs/ {pulse_ab_shift} Sec."
        })

    # --------------------------------------------------
    # FINAL RESPONSE
    # --------------------------------------------------
    return {
        "year": year,
        "month": month,
        "data": final_result
    }







@router.post("/outbound/Report")
def outbound_Report(
    company_id: int,
    start_date: date,
    end_date: date,
    db=Depends(get_db4),     # call_master_out DB
    db2=Depends(get_db2)     # vicidial DB
):
    campaign_in = get_campaign_in_clause(db, company_id)

    start_dt = f"{start_date} 00:00:00"
    end_dt = f"{end_date} 23:59:59"

    # ---------------------------------------
    # 1️⃣ Total Calls (Aggregated)
    # ---------------------------------------
    total_query = text(f"""
        SELECT 
            vl.user,
            vu.full_name,
            COUNT(*) AS total_calls
        FROM vicidial_log vl
        LEFT JOIN vicidial_agent_log val 
            ON vl.uniqueid = val.uniqueid
        LEFT JOIN vicidial_users vu 
            ON vl.user = vu.user
        WHERE vl.call_date BETWEEN :start AND :end
        AND vl.campaign_id IN ({campaign_in})
        AND vl.lead_id IS NOT NULL
        GROUP BY vl.user, vu.full_name
    """)

    total_rows = db2.execute(
        total_query,
        {"start": start_dt, "end": end_dt}
    ).mappings().all()

    # ---------------------------------------
    # 2️⃣ Connected Calls (STRICT)
    # ---------------------------------------
    connected_query = text(f"""
        SELECT v.user, COUNT(*) AS connected
        FROM vicidial_log v
        INNER JOIN vicidial_agent_log va
            ON v.uniqueid = va.uniqueid
        WHERE v.call_date BETWEEN :start AND :end
        AND v.campaign_id IN ({campaign_in})
        AND v.length_in_sec > 0
        AND v.user != 'VDAD'
        GROUP BY v.user
    """)

    connected_rows = db2.execute(
        connected_query,
        {"start": start_dt, "end": end_dt}
    ).mappings().all()

    connected_map = {r["user"]: int(r["connected"]) for r in connected_rows}

    # ---------------------------------------
    # 3️⃣ Get CONNECTED uniqueids ONLY
    # ---------------------------------------
    connected_uid_query = text(f"""
        SELECT v.uniqueid
        FROM vicidial_log v
        INNER JOIN vicidial_agent_log va
            ON v.uniqueid = va.uniqueid
        WHERE v.call_date BETWEEN :start AND :end
        AND v.campaign_id IN ({campaign_in})
        AND v.length_in_sec > 0
        AND v.user != 'VDAD'
    """)

    uid_rows = db2.execute(
        connected_uid_query,
        {"start": start_dt, "end": end_dt}
    ).fetchall()

    connected_uniqueids = [row[0] for row in uid_rows]

    status_summary = {}
    total_connected = len(connected_uniqueids)
    matched_count = 0

    # ---------------------------------------
    # 4️⃣ Fetch Category2 for Connected Calls
    # ---------------------------------------
    if connected_uniqueids:

        chunk_size = 1000

        for i in range(0, len(connected_uniqueids), chunk_size):
            chunk = connected_uniqueids[i:i + chunk_size]

            status_query = text("""
                SELECT 
                    Category2,
                    COUNT(*) AS count
                FROM call_master_out
                WHERE LiveUniqueId IN :uids
                GROUP BY Category2
            """)

            rows = db.execute(
                status_query,
                {"uids": tuple(chunk)}
            ).mappings().all()

            for r in rows:
                key = r["Category2"] if r["Category2"] else "Blank"
                count = int(r["count"])

                status_summary[key] = status_summary.get(key, 0) + count
                matched_count += count

    # ---------------------------------------
    # 5️⃣ Add Missing as Blank
    # ---------------------------------------
    blank_missing = total_connected - matched_count

    if blank_missing > 0:
        status_summary["Blank"] = status_summary.get("Blank", 0) + blank_missing
    # ---------------------------------------
    # 5️⃣ Build Final Response
    # ---------------------------------------
    overall = {"connected": 0, "notConnected": 0, "totalCalls": 0}
    agents = {}
    ob_auto = {"connected": 0, "notConnected": 0, "totalCalls": 0}

    for r in total_rows:
        user = r["user"]
        full_name = r["full_name"] if r["full_name"] else user

        total = int(r["total_calls"])
        connected = connected_map.get(user, 0)
        not_connected = total - connected

        overall["connected"] += connected
        overall["notConnected"] += not_connected
        overall["totalCalls"] += total

        if user == "VDAD":
            ob_auto["connected"] += connected
            ob_auto["notConnected"] += not_connected
            ob_auto["totalCalls"] += total
        else:
            agents[full_name] = {
                "connected": connected,
                "notConnected": not_connected,
                "totalCalls": total
            }

    
    # ---------------------------------------
    # 6️⃣ Demo Booked Data
    # ---------------------------------------
    demo_query = text("""
        SELECT 
            Field1 AS `Name`,
            Field2 AS `Contact Number`,
            Field3 AS `Email Address`,
            Field4,
            Field6 AS `App Installation Done`,
            Field5 AS `Meeting Arrange Date`,            
            Field7 AS `Location`,
            callcreated,
            CallDate AS `Call Date`
        FROM call_master_out
        WHERE ClientId = :client_id
        AND DATE(CallDate) BETWEEN :start AND :end
        AND Category2 = 'Demo Booked'
    """)

    demo_rows = db.execute(
        demo_query,
        {
            "client_id": company_id,
            "start": start_date,
            "end": end_date
        }
    ).mappings().all()

    demo_rows = [dict(r) for r in demo_rows]

    # ---------------------------------------
    # Fetch Agent Names from vicidial_users
    # ---------------------------------------

    # extract agent id from "DialDesk - IDC60654"
    users = list({
        r["callcreated"].split("-")[-1].strip()
        for r in demo_rows if r["callcreated"]
    })

    agent_map = {}

    if users:
        user_query = text("""
            SELECT user, full_name
            FROM vicidial_users
            WHERE user IN :users
        """)

        user_rows = db2.execute(
            user_query,
            {"users": tuple(users)}
        ).mappings().all()

        agent_map = {r["user"]: r["full_name"] for r in user_rows}


    # ---------------------------------------
    # Replace Agent username with full name
    # ---------------------------------------
    for r in demo_rows:

        raw_agent = r["callcreated"]

        if raw_agent and "-" in raw_agent:
            username = raw_agent.split("-")[-1].strip()
        else:
            username = raw_agent

        r["Agent"] = agent_map.get(username, username)
        del r["callcreated"]

    demo_booked_calls = demo_rows

    return {
        "overall": overall,
        "agents": agents,
        "obAutoDial": ob_auto,
        "statusBreakdown_From_SubScenario2": status_summary,
        "demoBookedCalls": demo_booked_calls
    }







def generate_outbound_excel(company_id, start_date, db, db2):

    # call your existing function logic
    data = outbound_Report(company_id, start_date, start_date, db, db2)

    overall = data["overall"]
    agents = data["agents"]
    status_summary = data["statusBreakdown_From_SubScenario2"]
    demo_rows = data["demoBookedCalls"]
    ob_auto = data["obAutoDial"]

    # --------------------------------
    # SECOND REPORT DATA (OB CDR)
    # --------------------------------
    request = OBCDRReportRequest(
        company_id=company_id,
        from_date=start_date,
        to_date=start_date
    )

    ob_cdr_rows = get_ob_cdr_report(request, db, db2)

    # --------------------------------
    # Create Workbook
    # --------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "DIGICOFFER REPORT"

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def apply_border(ws, start_row, end_row, start_col, end_col):
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                ws.cell(row=r, column=c).border = thin_border

    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    row = 1

    # --------------------------------
    # Report Title
    # --------------------------------
    title = f"DIGICOFFER REPORT ({start_date})"

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value=title).font = Font(bold=True)
    row += 2

    # --------------------------------
    # Overall Summary
    # --------------------------------
    ws.cell(row=row, column=1, value="Connected Calls")
    ws.cell(row=row, column=2, value="Not Connected Calls")
    ws.cell(row=row, column=3, value="Total Calls")

    for col in range(1,4):
        c = ws.cell(row=row, column=col)
        c.fill = header_fill
        c.font = header_font

    row += 1

    ws.cell(row=row, column=1, value=overall["connected"])
    ws.cell(row=row, column=2, value=overall["notConnected"])
    ws.cell(row=row, column=3, value=overall["totalCalls"])

    apply_border(ws, row-1, row, 1, 3)

    row += 3

    # --------------------------------
    # Agent Performance
    # --------------------------------
    ws.cell(row=row, column=1, value="Agent Performance").font = Font(bold=True)
    row += 1

    headers = ["Agent Name", "Connected", "Not Connected", "Total Calls"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    row += 1

    for agent, stats in agents.items():
        ws.cell(row=row, column=1, value=agent)
        ws.cell(row=row, column=2, value=stats["connected"])
        ws.cell(row=row, column=3, value=stats["notConnected"])
        ws.cell(row=row, column=4, value=stats["totalCalls"])
        row += 1

    agent_start = row - len(agents) - 1
    agent_end = row - 1
    apply_border(ws, agent_start, agent_end, 1, 4)

    row += 2

    # --------------------------------
    # Drop Calls
    # --------------------------------
    ws.cell(row=row, column=1, value="Drop Calls").font = Font(bold=True)
    row += 1

    ws.cell(row=row, column=1, value="Not Connected")
    ws.cell(row=row, column=2, value=ob_auto["notConnected"])

    apply_border(ws, row, row, 1, 2)

    row += 3

    # --------------------------------
    # Status Breakdown
    # --------------------------------
    ws.cell(row=row, column=1, value="Status Breakdown").font = Font(bold=True)
    row += 1

    ws.cell(row=row, column=1, value="Status").fill = header_fill
    ws.cell(row=row, column=1).font = header_font

    ws.cell(row=row, column=2, value="Count").fill = header_fill
    ws.cell(row=row, column=2).font = header_font

    row += 1

    for status, count in status_summary.items():
        ws.cell(row=row, column=1, value=status)
        ws.cell(row=row, column=2, value=count)
        row += 1

    status_start = row - len(status_summary) - 1
    status_end = row - 1
    apply_border(ws, status_start, status_end, 1, 2)

    row += 3

    # --------------------------------
    # Demo Booked Details
    # --------------------------------
    ws.cell(row=row, column=1, value="Demo Booked Details").font = Font(bold=True)
    row += 1

    demo_headers = [
        "Name",
        "Email Address",
        "Contact Number",
        "App Installation Done",
        "Meeting Arrange Date",
        "Location",
        "Agent Name",
        "Call Date"
    ]

    for col, h in enumerate(demo_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    row += 1

    for r in demo_rows:
        ws.cell(row=row, column=1, value=r.get("Name"))
        ws.cell(row=row, column=2, value=r.get("Email Address"))
        ws.cell(row=row, column=3, value=r.get("Contact Number"))
        ws.cell(row=row, column=4, value=r.get("App Installation Done"))
        ws.cell(row=row, column=5, value=r.get("Meeting Arrange Date"))
        ws.cell(row=row, column=6, value=r.get("Location"))
        ws.cell(row=row, column=7, value=r.get("Agent"))
        ws.cell(row=row, column=8, value=r.get("Call Date"))
        row += 1

    demo_start = row - len(demo_rows) - 1
    demo_end = row - 1
    apply_border(ws, demo_start, demo_end, 1, 8)

    # --------------------------------
    # Column Widths (ADD HERE)
    # --------------------------------
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 32
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 22
    ws.column_dimensions['H'].width = 22

    # --------------------------------
    # SHEET 2 (OB CDR REPORT)
    # --------------------------------
    ws2 = wb.create_sheet(title="OB CDR REPORT")

    if ob_cdr_rows:

        headers = list(ob_cdr_rows[0].keys())

        # header
        for col, h in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font

        # data
        row_idx = 2
        for record in ob_cdr_rows:
            for col_idx, value in enumerate(record.values(), 1):
                ws2.cell(row=row_idx, column=col_idx, value=value)
            row_idx += 1

        apply_border(ws2, 1, row_idx-1, 1, len(headers))

    # --------------------------------
    # Save to Stream
    # --------------------------------
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return stream








@router.get("/closer-log-report")
def closer_log_report(
    start_date: date = Query(...),
    end_date: date = Query(...),
    is_shared: str = Query("all"),   # default = all
    db: Session = Depends(get_db2),
    db4: Session = Depends(get_db4),
) -> Dict:


    campaign_filter_query = """
        SELECT campaignid 
        FROM registration_master
        WHERE status = 'A'
        AND campaignid IS NOT NULL
    """

    params = {}

    if is_shared.lower() == "true":
        campaign_filter_query += " AND is_shared = 1"
    elif is_shared.lower() == "false":
        campaign_filter_query += " AND is_shared = 0"
    elif is_shared.lower() == "all":
        pass
    else:
        raise HTTPException(status_code=400, detail="is_shared must be true, false, or all")

    campaign_filter_query = text(campaign_filter_query)

    rows = db4.execute(campaign_filter_query, params).fetchall()

    # 🔹 Convert to list
    campaign_ids = []

    for row in rows:
        if row[0]:
            campaign_ids.extend([
                cid.strip()
                for cid in row[0].replace("'", "").split(",")
                if cid.strip()
            ])

    if not campaign_ids:
        return {
            "overall": {},
            "campaigns_date_wise": {},
            "agents_date_wise": {}
        }

    # 🔹 Overall
    overall_query = text("""
        SELECT 
            COUNT(*) AS total_calls,
            SUM(length_in_sec) AS total_talk_time,
            AVG(length_in_sec) AS avg_call_duration
        FROM vicidial_closer_log
        WHERE DATE(call_date) BETWEEN :start AND :end
        AND user != 'VDCL'
        AND campaign_id IN :campaign_ids
    """).bindparams(bindparam("campaign_ids", expanding=True))

    overall = db.execute(overall_query, {
        "start": start_date,
        "end": end_date,
        "campaign_ids": tuple(campaign_ids)
    }).mappings().first()

    # 🔹 Campaign-wise (date-wise)
    campaign_query = text("""
        SELECT 
            DATE(call_date) AS call_date,
            campaign_id,
            COUNT(*) AS total_calls,
            SUM(length_in_sec) AS talk_time
        FROM vicidial_closer_log
        WHERE DATE(call_date) BETWEEN :start AND :end
        AND user != 'VDCL'
        AND campaign_id IN :campaign_ids
        GROUP BY DATE(call_date), campaign_id
        ORDER BY call_date, campaign_id ASC
    """).bindparams(bindparam("campaign_ids", expanding=True))

    campaign_rows = db.execute(campaign_query, {
        "start": start_date,
        "end": end_date,
        "campaign_ids": tuple(campaign_ids)
    }).mappings().all()

    # 🔹 Agent → Campaign breakdown (WITH full_name)
    agent_query = text("""
        SELECT 
            DATE(vcl.call_date) AS call_date,
            vcl.user AS user,
            COALESCE(vu.full_name, vcl.user) AS full_name,
            vcl.campaign_id,
            COUNT(*) AS total_calls
        FROM vicidial_closer_log vcl
        LEFT JOIN vicidial_users vu 
            ON vcl.user = vu.user
        WHERE DATE(vcl.call_date) BETWEEN :start AND :end
        AND vcl.user != 'VDCL'
        AND vcl.campaign_id IN :campaign_ids
        GROUP BY DATE(vcl.call_date), vcl.user, vu.full_name, vcl.campaign_id
        ORDER BY call_date, full_name
    """).bindparams(bindparam("campaign_ids", expanding=True))

    agent_rows = db.execute(agent_query, {
        "start": start_date,
        "end": end_date,
        "campaign_ids": tuple(campaign_ids)
    }).mappings().all()

    # -------------------------------
    # 🔹 Transform Campaign Data
    # -------------------------------
    campaign_dict = defaultdict(list)

    for row in campaign_rows:
        campaign_dict[str(row["call_date"])].append({
            "campaign_id": row["campaign_id"],
            "total_calls": row["total_calls"],
            "talk_time": row["talk_time"]
        })

    # -------------------------------
    # 🔹 Transform Agent Data
    # user + full_name → campaigns
    # -------------------------------
    agent_dict = defaultdict(lambda: defaultdict(lambda: {
        "user": "",
        "full_name": "",
        "campaigns": []
    }))

    for row in agent_rows:
        date_key = str(row["call_date"])
        user_key = row["user"]

        agent_entry = agent_dict[date_key][user_key]

        agent_entry["user"] = row["user"]
        agent_entry["full_name"] = row["full_name"]

        agent_entry["campaigns"].append({
            "campaign_id": row["campaign_id"],
            "total_calls": row["total_calls"]
        })

    # format final output
    formatted_agents = {}

    for date_key, users in agent_dict.items():
        formatted_agents[date_key] = list(users.values())

    return {
        "overall": dict(overall) if overall else {},
        "campaigns_date_wise": dict(campaign_dict),
        "agents_date_wise": formatted_agents
    }












@router.get("/closer-log-report-by-campaign")
def closer_log_report_by_campaign(
    start_date: date = Query(...),
    end_date: date = Query(...),
    is_shared: str = Query("all"),
    db: Session = Depends(get_db2),
    db4: Session = Depends(get_db4)
) -> Dict:
    
    base_query = """
        SELECT campaignid 
        FROM registration_master
        WHERE status = 'A'
        AND campaignid IS NOT NULL
    """

    is_shared = is_shared.lower()

    if is_shared == "true":
        base_query += " AND is_shared = 1"
    elif is_shared == "false":
        base_query += " AND is_shared = 0"
    elif is_shared == "all":
        pass
    else:
        raise HTTPException(status_code=400, detail="is_shared must be true, false, or all")

    rows = db4.execute(text(base_query)).fetchall()

    campaign_ids = []
    for row in rows:
        if row[0]:
            campaign_ids.extend([
                cid.strip()
                for cid in row[0].replace("'", "").split(",")
                if cid.strip()
            ])

    # ✅ handle empty
    if not campaign_ids:
        return {
            "overall": {},
            "campaigns_date_wise": {},
            "campaigns_users_date_wise": {}
        }

    # 🔹 Overall (same)
    overall_query = text("""
        SELECT 
            COUNT(*) AS total_calls,
            SUM(length_in_sec) AS total_talk_time,
            AVG(length_in_sec) AS avg_call_duration
        FROM vicidial_closer_log
        WHERE DATE(call_date) BETWEEN :start AND :end
        AND user != 'VDCL'
        AND campaign_id IN :campaign_ids
    """).bindparams(bindparam("campaign_ids", expanding=True))

    overall = db.execute(overall_query, {
        "start": start_date,
        "end": end_date,
        "campaign_ids": tuple(campaign_ids)
    }).mappings().first()

    # 🔹 Campaign-wise (same)
    campaign_query = text("""
        SELECT 
            DATE(call_date) AS call_date,
            campaign_id,
            COUNT(*) AS total_calls,
            SUM(length_in_sec) AS talk_time
        FROM vicidial_closer_log
        WHERE DATE(call_date) BETWEEN :start AND :end
        AND user != 'VDCL'
        AND campaign_id IN :campaign_ids
        GROUP BY DATE(call_date), campaign_id
        ORDER BY call_date, campaign_id ASC
    """).bindparams(bindparam("campaign_ids", expanding=True))

    campaign_rows = db.execute(campaign_query, {
        "start": start_date,
        "end": end_date,
        "campaign_ids": tuple(campaign_ids)
    }).mappings().all()

    # 🔹 Agent query (same)
    agent_query = text("""
        SELECT 
            DATE(vcl.call_date) AS call_date,
            vcl.user AS user,
            COALESCE(vu.full_name, vcl.user) AS full_name,
            vcl.campaign_id,
            COUNT(*) AS total_calls
        FROM vicidial_closer_log vcl
        LEFT JOIN vicidial_users vu 
            ON vcl.user = vu.user
        WHERE DATE(vcl.call_date) BETWEEN :start AND :end
        AND vcl.user != 'VDCL'
        AND vcl.campaign_id IN :campaign_ids
        GROUP BY DATE(vcl.call_date), vcl.user, vu.full_name, vcl.campaign_id
        ORDER BY call_date, campaign_id, full_name
    """).bindparams(bindparam("campaign_ids", expanding=True))

    agent_rows = db.execute(agent_query, {
        "start": start_date,
        "end": end_date,
        "campaign_ids": tuple(campaign_ids)
    }).mappings().all()

    # -------------------------------
    # 🔹 Campaign Date Wise (same)
    # -------------------------------
    campaign_dict = defaultdict(list)

    for row in campaign_rows:
        campaign_dict[str(row["call_date"])].append({
            "campaign_id": row["campaign_id"],
            "total_calls": row["total_calls"],
            "talk_time": row["talk_time"]
        })

    # -------------------------------
    # 🔹 🔥 NEW: Campaign → Users
    # -------------------------------
    campaign_user_dict = defaultdict(lambda: defaultdict(lambda: {
        "campaign_id": "",
        "users": []
    }))

    for row in agent_rows:
        date_key = str(row["call_date"])
        campaign_key = row["campaign_id"]

        campaign_entry = campaign_user_dict[date_key][campaign_key]

        campaign_entry["campaign_id"] = campaign_key

        campaign_entry["users"].append({
            "user": row["user"],
            "full_name": row["full_name"],
            "total_calls": row["total_calls"]
        })

    # format final output
    formatted_campaign_users = {}

    for date_key, campaigns in campaign_user_dict.items():
        formatted_campaign_users[date_key] = list(campaigns.values())

    return {
        "overall": dict(overall) if overall else {},
        "campaigns_date_wise": dict(campaign_dict),
        "campaigns_users_date_wise": formatted_campaign_users
    }












@router.get("/closer-log-report/excel")
def closer_log_report_excel(
    start_date: date = Query(...),
    end_date: date = Query(...),
    is_shared: str = Query("all"),
    db: Session = Depends(get_db2),
    db4: Session = Depends(get_db4),
):

    # -------------------------------
    # 🔹 KEEP YOUR EXISTING LOGIC
    # -------------------------------
    response = closer_log_report(start_date, end_date, is_shared, db, db4)

    agents_data = response["agents_date_wise"]

    # -------------------------------
    # 🔹 Generate Date List
    # -------------------------------
    date_list = []
    current = start_date
    while current <= end_date:
        date_list.append(str(current))
        current += timedelta(days=1)

    # -------------------------------
    # 🔹 Transform → Pivot Structure
    # user → campaign → date → calls
    # -------------------------------
    pivot = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))

    for date_key, users in agents_data.items():
        for user in users:
            full_name = user["full_name"]

            for camp in user["campaigns"]:
                campaign = camp["campaign_id"]
                calls = camp["total_calls"]

                pivot[full_name][campaign][date_key] += calls

    blue_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    white_bold_font = Font(bold=True, color="FFFFFF")

    # -------------------------------
    # 🔹 Create Excel
    # -------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Agent Wise"

    bold = Font(bold=True)

    # Header
    header = ["Row Labels"]
    header += [d for d in date_list]
    header.append("Grand Total")

    ws.append(header)

    # style header
    for col in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = blue_fill
        cell.font = white_bold_font

    # -------------------------------
    # 🔹 Fill Data
    # -------------------------------
    for user, campaigns in pivot.items():

        # USER ROW
        user_row = [user]
        user_totals = []

        for d in date_list:
            total = sum(campaigns[c].get(d, 0) for c in campaigns)
            user_totals.append(total)

        user_row += user_totals
        user_row.append(sum(user_totals))

        ws.append(user_row)

        # bold user row
        for col in range(1, len(user_row) + 1):
            ws.cell(row=ws.max_row, column=col).font = bold

        # CAMPAIGN ROWS
        for campaign, date_map in campaigns.items():
            row = ["   " + campaign]

            totals = []
            for d in date_list:
                val = date_map.get(d, 0)
                totals.append(val)

            row += totals
            row.append(sum(totals))

            ws.append(row)

    # -------------------------------
    # 🔹 Auto Width
    # -------------------------------
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 2


    # -------------------------------
    # 🔹 GRAND TOTAL (BOTTOM ROW)
    # -------------------------------
    grand_row = ["Grand Total"]

    column_totals = []

    for d in date_list:
        total = 0
        for user, campaigns in pivot.items():
            for campaign in campaigns:
                total += campaigns[campaign].get(d, 0)
        column_totals.append(total)

    grand_row += column_totals
    grand_row.append(sum(column_totals))

    ws.append(grand_row)

    # style grand total row
    for col in range(1, len(grand_row) + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.fill = blue_fill
        cell.font = white_bold_font






    # -------------------------------
    # 🔹 SECOND SHEET (Campaign → Users)
    # -------------------------------
    campaign_response = closer_log_report_by_campaign(start_date, end_date, is_shared, db, db4)
    campaign_users_data = campaign_response["campaigns_users_date_wise"]

    ws2 = wb.create_sheet(title="Client Wise")

    # Header
    header2 = ["Row Labels"]
    header2 += [d for d in date_list]
    header2.append("Grand Total")

    ws2.append(header2)

    # style header
    for col in range(1, len(header2) + 1):
        cell = ws2.cell(row=1, column=col)
        cell.fill = blue_fill
        cell.font = white_bold_font

    # -------------------------------
    # 🔹 Transform → Pivot
    # campaign → user → date → calls
    # -------------------------------
    pivot2 = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))

    for date_key, campaigns in campaign_users_data.items():
        for campaign in campaigns:
            campaign_id = campaign["campaign_id"]

            for user in campaign["users"]:
                name = user["full_name"]
                calls = user["total_calls"]

                pivot2[campaign_id][name][date_key] += calls

    # -------------------------------
    # 🔹 Fill Data
    # -------------------------------
    for campaign, users in pivot2.items():

        # CAMPAIGN ROW
        camp_row = [campaign]
        camp_totals = []

        for d in date_list:
            total = sum(users[u].get(d, 0) for u in users)
            camp_totals.append(total)

        camp_row += camp_totals
        camp_row.append(sum(camp_totals))

        ws2.append(camp_row)

        # bold campaign row
        for col in range(1, len(camp_row) + 1):
            ws2.cell(row=ws2.max_row, column=col).font = bold

        # USER ROWS
        for user, date_map in users.items():
            row = ["   " + user]

            totals = []
            for d in date_list:
                val = date_map.get(d, 0)
                totals.append(val)

            row += totals
            row.append(sum(totals))

            ws2.append(row)

    # -------------------------------
    # 🔹 GRAND TOTAL (BOTTOM)
    # -------------------------------
    grand_row2 = ["Grand Total"]
    column_totals2 = []

    for d in date_list:
        total = 0
        for campaign, users in pivot2.items():
            for user in users:
                total += users[user].get(d, 0)
        column_totals2.append(total)

    grand_row2 += column_totals2
    grand_row2.append(sum(column_totals2))

    ws2.append(grand_row2)

    # style grand total
    for col in range(1, len(grand_row2) + 1):
        cell = ws2.cell(row=ws2.max_row, column=col)
        cell.fill = blue_fill
        cell.font = white_bold_font

    # -------------------------------
    # 🔹 Auto Width (Sheet 2)
    # -------------------------------
    for col in ws2.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws2.column_dimensions[col_letter].width = max_len + 2


    # -------------------------------
    # 🔹 Return File
    # -------------------------------
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=Agent_{start_date}_to_{end_date}.xlsx"
        }
    )




@router.get("/abandon-callback-report")
def abandon_callback_report(
    client_id: str,
    report_date: date,
    db=Depends(get_db2),
    db4=Depends(get_db4)
):
    # 🔹 Step 1: Fetch abandon data
    if client_id == "ALL":
        aband_data = db4.execute(text("""
        SELECT 
            CompanyName,
            PhoneNo,
            CallDate,
            call_status,
            call_attempt_time
        FROM aband_call_master
        WHERE DATE(CallDate) = :report_date
        """), {
            "report_date": report_date
        }).mappings().all()
    
    else:
        aband_data = db4.execute(text("""
            SELECT 
                CompanyName,
                PhoneNo,
                CallDate,
                call_status,
                call_attempt_time
            FROM aband_call_master
            WHERE ClientId = :client_id
            AND DATE(CallDate) = :report_date
        """), {
            "client_id": client_id,
            "report_date": report_date
        }).mappings().all()

    if not aband_data:
        return {"count": 0, "data": []}

    # 🔹 Step 2: Prepare phone list (last 10 digits)
    phone_map = {}
    phone_list = []

    for row in aband_data:
        clean_phone = str(row["PhoneNo"])[-10:]
        phone_map[clean_phone] = []
        phone_list.append(clean_phone)

    # 🔹 Step 3: Fetch ALL logs in ONE query
    logs = db.execute(text(f"""
        SELECT 
            t2.call_date,
            t2.user,
            RIGHT(t2.phone_number, 10) as phone,
            mcl.id as manual_id
        FROM vicidial_log t2
        LEFT JOIN vicidial_users t4 ON t2.user = t4.user
        LEFT JOIN asterisk.manual_call_log mcl 
            ON RIGHT(mcl.phone_number,10) = RIGHT(t2.phone_number,10) 
            AND mcl.uniqueid = t2.uniqueid
        WHERE t2.campaign_id IN ('dialdesk','Cryst002','Ajmal000','Superher','DLFDE000')
        AND DATE(t2.call_date) = :report_date
        AND t2.list_id IN ('998','2001')
        AND t2.lead_id IS NOT NULL
        AND RIGHT(t2.phone_number, 10) IN :phones
        ORDER BY t2.call_date ASC
    """), {
        "report_date": report_date,
        "phones": tuple(phone_list)
    }).mappings().all()

    # 🔹 Step 4: Group logs by phone
    log_map = defaultdict(list)

    for log in logs:
        if log["user"] == "VDAD":
            continue

        phone = log["phone"]

        if len(log_map[phone]) < 3:
            log_map[phone].append({
                "time": log["call_date"],
                "status": "Connected" if log["manual_id"] is not None else "Not connected"
            })

    now = datetime.now()
    today = now.date()

    result = []

    # 🔹 Step 5: Process data
    for row in aband_data:
        phone_raw = row["PhoneNo"]
        phone = str(phone_raw)[-10:]
        call_time = row["CallDate"]
        call_status = row["call_status"]
        call_attempt_time = row["call_attempt_time"]

        is_done = (call_status == "DONE")


        attempts = log_map.get(phone, [])
        attempt_count = len(attempts)

        # Default
        first_time = second_time = third_time = None
        first_status = second_status = third_status = None

        # 🔥 Connected
        if any(a["status"] == "Connected" for a in attempts):
            final_status = "Connected"

            while len(attempts) < 3:
                attempts.append({"time": None, "status": None})

            first_time, first_status = attempts[0]["time"], attempts[0]["status"]
            second_time, second_status = attempts[1]["time"], attempts[1]["status"]
            third_time, third_status = attempts[2]["time"], attempts[2]["status"]

        elif report_date != today:
            final_status = "No callback attempted" if attempt_count == 0 else "Not connected"

            while len(attempts) < 3:
                attempts.append({"time": None, "status": None})

            first_time, first_status = attempts[0]["time"], attempts[0]["status"]
            second_time, second_status = attempts[1]["time"], attempts[1]["status"]
            third_time, third_status = attempts[2]["time"], attempts[2]["status"]

        else:

            # 🔥 Fresh Abandon
            if attempt_count == 0:
                if not is_done:
                    first_status = "Callback Attempt 1 Pending"
                    second_status = "Callback Attempt 2 Pending"
                    third_status = "Callback Attempt 3 Pending"
                    final_status = "Fresh Abandon"
                else:
                    # 🔥 DONE + no attempts → no pending
                    first_status = second_status = third_status = None
                    final_status = "DONE"

            else:
                final_status = "Not connected"

                while len(attempts) < 3:
                    attempts.append({"time": None, "status": None})

                first_time, first_status = attempts[0]["time"], attempts[0]["status"]
                second_time, second_status = attempts[1]["time"], attempts[1]["status"]
                third_time, third_status = attempts[2]["time"], attempts[2]["status"]

        if is_done:
            final_status = "DONE"

        result.append({
            "client_name": row["CompanyName"],
            "date": call_time.date(),
            "phone_number": phone,
            "call_abandon_time": call_time.strftime("%Y-%m-%d %H:%M:%S"),

            "first_attempt_time": first_time.strftime("%Y-%m-%d %H:%M:%S") if first_time else None,
            "first_status": first_status,

            "second_attempt_time": second_time.strftime("%Y-%m-%d %H:%M:%S") if second_time else None,
            "second_status": second_status,

            "third_attempt_time": third_time.strftime("%Y-%m-%d %H:%M:%S") if third_time else None,
            "third_status": third_status,

            "final_status": final_status,
            "call_attempt_time": call_attempt_time
        })

    return {
        "count": len(result),
        "data": result
    }









@router.get("/abandon-callback-report/excel")
def abandon_callback_report_excel(
    client_id: str,
    report_date: date,
    db=Depends(get_db2),
    db4=Depends(get_db4)
):
    # 🔹 Get data from your existing API function
    response = abandon_callback_report(client_id, report_date, db, db4)
    data = response["data"]

    campaigns = []

    if client_id != "ALL":
        campaign_result = db4.execute(text("""
            SELECT campaignid 
            FROM registration_master 
            WHERE company_id = :client_id
        """), {"client_id": client_id}).mappings().fetchone()

        # campaigns = [row[0] for row in campaign_result]
        raw_campaign = campaign_result["campaignid"]
        campaigns = [c.strip().strip("'") for c in raw_campaign.split(",") if c.strip()]


    if client_id == "ALL":
        total_abandoned_calls = db.execute(text("""
            SELECT COUNT(phone_number) AS total
            FROM asterisk.vicidial_closer_log t2
            WHERE DATE(t2.call_date) = :report_date
            AND t2.lead_id IS NOT NULL
            AND t2.user = 'VDCL'
        """), {"report_date": report_date}).scalar()
        

    else:
        total_abandoned_calls = db.execute(text(f"""
            SELECT COUNT(phone_number) AS total
            FROM asterisk.vicidial_closer_log t2
            WHERE DATE(t2.call_date) = :report_date
            AND t2.lead_id IS NOT NULL
            AND t2.user = 'VDCL'
            AND t2.campaign_id IN :campaigns
        """), {
            "report_date": report_date,
            "campaigns": tuple(campaigns)
        }).scalar()


    after_8pm_count = 0
    failed_before_8pm = 0
    failed_after_8pm = 0
    called_not_connected = 0
    connected_count = 0
    need_to_callback = 0
    inbound_received = 0
    unique_phones = set()

    cutoff_time = time(20, 0, 0)  # 8 PM

    if client_id == "ALL":
        safe_company_name = "ALL"
    else:
        company_name = data[0]["client_name"] if data else "Report"
        safe_company_name = company_name[:10]

    # 🔹 Create Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "RL Report"

    # 🔹 Header
    headers = [
        "Client Name", "Abandoned Call Back", "Phone Number", "Call Abandon time",
        "First Attempt Time", "Call Back Done Successfully",
        "Second Attempt Time", "Call Back Done Successfully",
        "Third Attempt Time", "Call Back Done Successfully",
        "Final Status", "Call Attempt Time"
    ]

    # 🔹 Blue Header Style
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Write header
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # 🔹 Fill Data
    for row_num, row in enumerate(data, 2):
        ws.cell(row=row_num, column=1, value=row["client_name"])
        ws.cell(row=row_num, column=2, value=str(row["date"]))
        ws.cell(row=row_num, column=3, value=row["phone_number"])
        ws.cell(row=row_num, column=4, value=row["call_abandon_time"])

        ws.cell(row=row_num, column=5, value=row["first_attempt_time"])
        ws.cell(row=row_num, column=6, value=row["first_status"])

        ws.cell(row=row_num, column=7, value=row["second_attempt_time"])
        ws.cell(row=row_num, column=8, value=row["second_status"])

        ws.cell(row=row_num, column=9, value=row["third_attempt_time"])
        ws.cell(row=row_num, column=10, value=row["third_status"])

        ws.cell(row=row_num, column=11, value=row["final_status"])
        ws.cell(row=row_num, column=12, value=row["call_attempt_time"])

    # 🔹 Auto adjust column width
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2


    # --------------------------------
    # Column Widths (ADD HERE)
    # --------------------------------
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 32
    ws.column_dimensions['F'].width = 24
    ws.column_dimensions['G'].width = 22
    ws.column_dimensions['H'].width = 24
    ws.column_dimensions['I'].width = 22
    ws.column_dimensions['J'].width = 24
    ws.column_dimensions['K'].width = 22
    ws.column_dimensions['L'].width = 22


    for row in data:
        if not row["call_abandon_time"]:
            continue

        raw_time = row["call_abandon_time"]

        # ✅ Track unique phone numbers
        if row["phone_number"]:
            unique_phones.add(row["phone_number"])

        # ✅ Safe parsing
        if isinstance(raw_time, datetime):
            call_time = raw_time.time()
        else:
            call_time = datetime.strptime(raw_time, "%Y-%m-%d %H:%M:%S").time()

        final_status = row["final_status"]

        # -----------------------------
        # Existing Logic (8 PM split)
        # -----------------------------
        if call_time >= cutoff_time:
            after_8pm_count += 1

        if final_status not in ["Connected", "Not connected", "DONE"]:
            if call_time < cutoff_time:
                failed_before_8pm += 1
            else:
                failed_after_8pm += 1

        # -----------------------------
        # ✅ NEW COUNTS
        # -----------------------------

        # 1. Called but not connected
        if final_status == "Not connected":
            called_not_connected += 1

        # 2. Connected
        elif final_status == "Connected":
            connected_count += 1

        # 3. Need to callback
        elif final_status in ["Fresh Abandon", "No callback attempted"]:
            need_to_callback += 1

        # 4. Inbound Received (DONE)
        elif final_status == "DONE":
            inbound_received += 1

    unique_phone_count = len(unique_phones)


    ws2 = wb.create_sheet(title="Summary")

    summary_headers = ["CATEGORY", "COUNT"]

    for col_num, header in enumerate(summary_headers, 1):
        cell = ws2.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font

    summary_data = [
        ("CALLED AFTER 8 PM", after_8pm_count),
        ("TOTAL ABANDONED CALLS", total_abandoned_calls),
        ("UNIQUE ABANDONED CALLS", unique_phone_count),
        ("INBOUND RECEIVED", inbound_received),
        ("NEED TO BE CALLED BACK", need_to_callback),
        ("CONNECTED SUCCESSFULLY", connected_count),
        ("CALLED BUT NOT CONNECTED", called_not_connected),       
        ("MISSED RL (PRE-8 PM – NO AGENT AVAILABILITY)", failed_before_8pm),
        ("MISSED CALLS AFTER 8 PM", failed_after_8pm),        
    ]

    for row_num, (category, value) in enumerate(summary_data, 2):
        ws2.cell(row=row_num, column=1, value=category)
        ws2.cell(row=row_num, column=2, value=value)

    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 20

    # 🔹 Save to memory
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    # 🔹 Return file
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={safe_company_name}_RL_Report_{report_date}.xlsx"
        }
    )