from http.client import HTTPException
from io import BytesIO

from fastapi import APIRouter, Query, Depends, Body, Form, HTTPException, Response
from sqlalchemy import text, bindparam
from typing import List, Dict, Optional, Any, Union
from pydantic import BaseModel
from starlette.responses import StreamingResponse
from datetime import date, time
from schemas import *
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session
from dotenv import load_dotenv
from database import get_engine4, get_engine3, get_db3, get_db4
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from collections import defaultdict
from openpyxl.utils import get_column_letter
from math import ceil
import math

load_dotenv()



router = APIRouter()










@router.get("/call/csat-report_old/{client_id}", response_model=List[Dict])
def get_csat_report_old(
    client_id: int,
    from_date: str = Query(...),
    to_date: str = Query(...),
):
    query = text("""
        SELECT vl.*, vcl.user, vu.full_name
        FROM csat_data vl
        INNER JOIN vicidial_closer_log vcl ON vl.uniqueid = vcl.uniqueid
        INNER JOIN vicidial_users vu ON vcl.user = vu.user
        WHERE vl.dtmf < 4
          AND vl.client_id = :client_id
          AND DATE(vl.call_date) BETWEEN :from_date AND :to_date
    """)

    try:
        engine = get_engine3()
        with engine.connect() as conn:
            result = conn.execute(query, {
                "client_id": client_id,
                "from_date": from_date,
                "to_date": to_date,
            }).mappings().all()

        return [dict(row) for row in result]

    except SQLAlchemyError as e:
        print("SQLAlchemy Error:", str(e))  # Good for local debugging
        raise HTTPException(status_code=500, detail="Database query failed.")









@router.get("/sla/slot-wise-utilization_old")
def slot_wise_utilization_old(
    startdate: str = Query(...),
    enddate: str = Query(...),
    clientID: str = Query(...),
    sd_type: str = Query("All"),
    db1: Session = Depends(get_db4),   # registration_master
    db2: Session = Depends(get_db3),   # vicidial
):
    data = {}
    datetimeArray = {}
    datearray = []
    timearray = []

    # ------------------ AGENT MASTER ------------------
    ag_sql = text("""
        SELECT user, full_name
        FROM vicidial_users
        WHERE active='Y'
    """)
    ag_rows = db2.execute(ag_sql).fetchall()
    ag_list = {r.user: r.full_name for r in ag_rows}

    # ------------------ DATE SETUP ------------------
    from_dt = datetime.strptime(startdate, "%Y-%m-%d")
    to_dt = datetime.strptime(enddate, "%Y-%m-%d") + timedelta(days=1)

    # ------------------ SD TYPE ------------------
    sd_str = ""
    if sd_type == "Shared":
        sd_str = " AND is_shared='1'"
    elif sd_type == "Dedicated":
        sd_str = " AND is_shared='0'"

    # ------------------ CLIENT / CAMPAIGN ------------------
    if clientID == "All":
        db1.execute(text("SET SESSION group_concat_max_len = 20000"))

        camp_sql = text(f"""
            SELECT GROUP_CONCAT(campaignid) AS campaign_id
            FROM registration_master
            WHERE status='A' AND is_dd_client='1' {sd_str}
        """)
        campaignId = db1.execute(camp_sql).fetchone().campaign_id
        campaign_cond = f"t2.campaign_id IN ({campaignId})" if campaignId else "1=0"

        cli_sql = text(f"""
            SELECT GROUP_CONCAT(company_id) AS company_id
            FROM registration_master
            WHERE status='A' AND is_dd_client='1' {sd_str}
        """)
        client_list = db1.execute(cli_sql).fetchone().company_id
        client_list_cond = f"ClientId IN ({client_list})" if client_list else "1=0"

    else:
        camp_sql = text(f"""
            SELECT campaignid
            FROM registration_master
            WHERE company_id=:clientID {sd_str}
            LIMIT 1
        """)
        campaignId = db1.execute(camp_sql, {"clientID": clientID}).fetchone().campaignid
        campaign_cond = f"t2.campaign_id IN ({campaignId})"
        client_list_cond = f"ClientId IN ({clientID})"

    # ------------------ LOOP SLOT WISE (HOURLY) ------------------
    cur = from_dt
    while cur < to_dt:
        date_label = cur.strftime("%Y-%m-%d")
        time_label = int(cur.strftime("%H"))

        start_time = cur.strftime("%Y-%m-%d %H:%M:%S")
        end_time = (cur + timedelta(hours=1)).strftime("%Y-%m-%d %H:%M:%S")

        datearray.append(date_label)
        timearray.append(time_label)
        datetimeArray.setdefault(date_label, []).append(time_label)

        qry = text(f"""
            SELECT 
                t2.user,
                t2.queue_seconds,
                vu.user_nickname,
                t1.talk_sec,
                t1.wait_sec,
                t1.dispo_sec,
                t1.pause_sec,
                IFNULL(t3.p,0) AS hold
            FROM asterisk.vicidial_closer_log t2
            LEFT JOIN asterisk.vicidial_users vu ON t2.user=vu.user
            LEFT JOIN asterisk.vicidial_agent_log t1 ON t1.uniqueid=t2.uniqueid AND t2.user=t1.user
            LEFT JOIN (
                SELECT uniqueid,SUM(parked_sec) p
                FROM park_log
                WHERE STATUS='GRABBED'
                AND parked_time>='{start_time}'
                AND parked_time<'{end_time}'
                GROUP BY uniqueid
            ) t3 ON t1.uniqueid=t3.uniqueid
            WHERE t2.call_date>='{start_time}'
            AND t2.call_date<'{end_time}'
            AND {campaign_cond}
        """)

        rows = db2.execute(qry).fetchall()

        total = 0
        within_sla = 0
        answered = 0

        agents_set = set()
        shared_set = set()
        dedicated_set = set()
        other_set = set()

        talk = wait = dispo = pause = hold = 0

        for r in rows:
            user = (r.user or "")
            nickname = str(r.user_nickname) if r.user_nickname else ""

            total += 1

            if user != "VDCL":
                answered += 1
                agents_set.add(user)

                if (r.queue_seconds or 0) <= 20:
                    within_sla += 1

                if nickname == "1":
                    shared_set.add(user)
                elif nickname == "0":
                    dedicated_set.add(user)
                else:
                    other_set.add(user)

            talk += r.talk_sec or 0
            wait += r.wait_sec or 0
            dispo += r.dispo_sec or 0
            pause += r.pause_sec or 0
            hold += r.hold or 0

        # data.setdefault(date_label, {})[time_label] = {
        #     "Total": r.Total,
        #     "Answered": r.Answered,
        #     "Manpower": r.Manpower,
        #     "Shared": r.Shared,
        #     "Dedicated": r.Dedicated,
        #     "Other": r.Other,
        #     "Talk": r.Talk,
        #     "wait": r.wait,
        #     "dispo": r.dispo,
        #     "pause": r.pause,
        #     "hold": r.hold,
        #     "Al %": round(r.Al or 0, 2),
        #     "SL %": round((r.WIthinSLA / r.Answered) * 100, 2) if r.Answered else 0,
        #     "Total login": r.Total_login,
        #     "Net login": r.Net_login,
        #     "Utilization %": round(r.Utilization or 0, 2),
        #     "WIthinSLA": r.WIthinSLA,
        #     "Manpower Agents": ",".join([f"{ag_list.get(a)}({a})" for a in (r.agents or "").split(",") if a]),
        #     "Shared Agents": ",".join([f"{ag_list.get(a)}({a})" for a in (r.Shared_ag or "").split(",") if a]),
        #     "Dedicated Agents": ",".join([f"{ag_list.get(a)}({a})" for a in (r.Dedicated_ag or "").split(",") if a]),
        #     "Other Agents": ",".join([f"{ag_list.get(a)}({a})" for a in (r.Other_ag or "").split(",") if a]),
        # }


        data.setdefault(date_label, {})[time_label] = {
            "Total": total,
            "Answered": answered,
            "Manpower": len(agents_set),
            "Shared": len(shared_set),
            "Dedicated": len(dedicated_set),
            "Other": len(other_set),

            "Talk": talk,
            "wait": wait,
            "dispo": dispo,
            "pause": pause,
            "hold": hold,

            "Al %": round((answered / total * 100) if total else 0, 2),
            "SL %": round((within_sla / answered * 100) if answered else 0, 2),

            "Total login": talk + wait + dispo + pause + hold,
            "Net login": talk + wait + dispo + hold,

            "Utilization %": round(
                ((talk + dispo + hold) / (talk + wait + dispo + hold) * 100)
                if (talk + wait + dispo + hold) else 0, 2
            ),

            "WIthinSLA": within_sla,

            "Manpower Agents": ",".join([f"{ag_list.get(a)}({a})" for a in agents_set]),
            "Shared Agents": ",".join([f"{ag_list.get(a)}({a})" for a in shared_set]),
            "Dedicated Agents": ",".join([f"{ag_list.get(a)}({a})" for a in dedicated_set]),
            "Other Agents": ",".join([f"{ag_list.get(a)}({a})" for a in other_set]),
        }

        # --- RL and RL % calculation ---
        rl_sql = text(f"""
            SELECT COUNT(1) as cnt
            FROM aband_call_master
            WHERE {client_list_cond}
            AND call_status='answer'
            AND calldate >= '{start_time}'
            AND calldate < '{end_time}'
        """)
        rl_count = db1.execute(rl_sql).fetchone().cnt or 0
        rl_percent = round(((rl_count + answered) / total) * 100, 2) if total else 0

        data[date_label][time_label].update({
            "RL": rl_count,
            "RL %": rl_percent
        })


        cur += timedelta(hours=1)

    return {
        "data": data,
        "datearray": list(set(datearray)),
        "timearray": list(set(timearray)),
        "datetimeArray": datetimeArray
    }




def generate_rl_sl_excel(startdate, enddate, clientID, sd_type, db1, db2):

    # ---------------------------------
    # Call existing API logic
    # ---------------------------------
    result = slot_wise_utilization_old(
        startdate=startdate,
        enddate=enddate,
        clientID=clientID,
        sd_type=sd_type,
        db1=db1,
        db2=db2
    )

    jsonData = result.get("data", {})

    rows = []

    # ---------------------------------
    # JSON → ROWS
    # ---------------------------------
    for date, hours in jsonData.items():
        for hour, values in hours.items():

            rows.append([
                date,
                hour,
                values.get("Total", 0),
                values.get("Answered", 0),
                values.get("Manpower", 0),
                values.get("Shared", 0),
                values.get("Dedicated", 0),
                values.get("Other", 0),
                values.get("Talk", 0),
                values.get("wait", 0),
                values.get("dispo", 0),
                values.get("hold", 0),
                values.get("Al %", 0),
                values.get("SL %", 0),
                values.get("RL %", 0),
                values.get("RL", 0),
                values.get("Total login", 0),
                values.get("Net login", 0),
                values.get("Utilization %", 0),
                values.get("WIthinSLA", 0),
                values.get("Manpower Agents", ""),
                values.get("Shared Agents", ""),
                values.get("Dedicated Agents", ""),
                values.get("Other Agents", "")
            ])

    # ---------------------------------
    # Create Workbook
    # ---------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "RL SL Report"

    headers = [
        "Date","Hour","Total","Answered","Manpower","Shared","Dedicated","Other",
        "Talk","Wait","Dispo","Hold",
        "Al %","SL %","RL %","RL",
        "Total Login","Net Login","Utilization %",
        "Within SLA",
        "Manpower Agents","Shared Agents","Dedicated Agents","Other Agents"
    ]

    ws.append(headers)

    # ---------------------------------
    # Add Data Rows
    # ---------------------------------
    for r in rows:
        ws.append(r)

    # ---------------------------------
    # GRAND TOTAL
    # ---------------------------------
    if rows:

        percent_cols = [12, 13, 14, 18]  # Al %, SL %, RL %, Utilization %

        grand = ["Grand Total", ""]

        for i in range(2, 20):

            col_values = [float(r[i]) if r[i] else 0 for r in rows]

            if i in percent_cols:
                grand.append(round(sum(col_values) / len(col_values), 2))
            else:
                grand.append(round(sum(col_values), 2))

        # agent columns empty
        grand += ["", "", "", ""]

        ws.append(grand)

    # ---------------------------------
    # Auto Column Width
    # ---------------------------------
    for col in ws.columns:

        max_length = 0
        column = col[0].column_letter

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column].width = max_length + 2

    # ---------------------------------
    # Save to Memory
    # ---------------------------------
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output










@router.get("/sla/export-sla-day-wise_old")
def export_sla_day_wise_old(
    startdate: str = Query(...),
    enddate: str = Query(...),
    clientID: str = Query(...),
    sd_type: Optional[str] = Query("All"),
    db1: Session = Depends(get_db4),
    db2: Session = Depends(get_db3),
):

    data = {}

    # ---------------- Agents master ----------------
    ag_rows = db2.execute(
        text("SELECT user, full_name FROM vicidial_users WHERE active='Y'")
    ).fetchall()

    ag_list = {r.user: r.full_name for r in ag_rows}

    # ---------------- SD filter ----------------
    sd_str = ""
    sd_str2 = ""

    if sd_type == "Shared":
        sd_str = " AND is_shared='1'"
        sd_str2 = " AND vu.user_nickname='1'"
    elif sd_type == "Dedicated":
        sd_str = " AND is_shared='0'"
        sd_str2 = " AND vu.user_nickname='0'"

    # ---------------- Campaign & Client filter ----------------
    if clientID == "All":
        db1.execute(text("SET SESSION group_concat_max_len = 20000"))

        camp = db1.execute(
            text(f"""
            SELECT GROUP_CONCAT(campaignid) campaignid
            FROM registration_master
            WHERE status='A' AND is_dd_client='1' {sd_str}
            """)
        ).scalar()

        cl = db1.execute(
            text(f"""
            SELECT GROUP_CONCAT(company_id) company_id
            FROM registration_master
            WHERE status='A' AND is_dd_client='1' {sd_str}
            """)
        ).scalar()

        campaignId = f"t2.campaign_id IN ({camp})"
        client_list_str = f"ClientId IN ({cl})"

    else:
        camp = db1.execute(
            text(f"""
            SELECT campaignid FROM registration_master
            WHERE company_id='{clientID}' {sd_str}
            """)
        ).scalar()

        campaignId = f"t2.campaign_id IN ({camp})"
        client_list_str = f"ClientId IN ({clientID})"

    # ---------------- Date handling ----------------
    FromDate = datetime.strptime(startdate + " 00:00:00", "%Y-%m-%d %H:%M:%S")
    ToDate = datetime.strptime(enddate + " 23:59:59", "%Y-%m-%d %H:%M:%S")

    # ---------------- Daily Loop ----------------
    while FromDate < ToDate:

        start_time_start = FromDate
        start_time_end = FromDate + timedelta(days=1)
        FromDate = start_time_end

        dateLabel = start_time_start.strftime("%d-%m-%Y")

        qry = f"""
        SELECT 
            t2.user,
            t2.queue_seconds,
            vu.user_nickname,
            t1.talk_sec,
            t1.wait_sec,
            t1.dispo_sec,
            t1.pause_sec,
            IFNULL(t3.p,0) as hold

        FROM vicidial_closer_log t2
        LEFT JOIN vicidial_users vu ON t2.user=vu.user
        LEFT JOIN vicidial_agent_log t1 ON t1.uniqueid=t2.uniqueid AND t2.user=t1.user
        LEFT JOIN (
            SELECT uniqueid,SUM(parked_sec) p
            FROM park_log
            WHERE STATUS='GRABBED'
              AND parked_time>='{start_time_start}'
              AND parked_time<'{start_time_end}'
            GROUP BY uniqueid
        ) t3 ON t1.uniqueid=t3.uniqueid

        WHERE t2.call_date>='{start_time_start}'
          AND t2.call_date<'{start_time_end}'
          AND {campaignId}
        """

        row = db2.execute(text(qry)).fetchall()

        Total = 0
        WithinSLA = 0
        Answered = 0

        agents = set()
        shared_agents = set()
        dedicated_agents = set()
        other_agents = set()

        Talk = wait = dispo = pause = hold = 0

        # ---------------- Agent Name Mapping ----------------
        def map_agents(csv):
            if not csv:
                return ""
            return ",".join(
                [f"{ag_list.get(a,a)}({a})" for a in csv.split(",") if a]
            )
        
        for r in row:
            Total += 1

            user = r.user
            queue = r.queue_seconds or 0
            nickname = r.user_nickname

            talk_sec = r.talk_sec or 0
            wait_sec = r.wait_sec or 0
            dispo_sec = r.dispo_sec or 0
            pause_sec = r.pause_sec or 0
            hold_sec = r.hold or 0

            if user != 'VDCL':
                Answered += 1
                agents.add(user)

                if queue <= 20:
                    WithinSLA += 1

                if nickname == '1':
                    shared_agents.add(user)
                elif nickname == '0':
                    dedicated_agents.add(user)
                else:
                    other_agents.add(user)

            Talk += talk_sec
            wait += wait_sec
            dispo += dispo_sec
            pause += pause_sec
            hold += hold_sec

        
        Manpower = len(agents)
        Shared = len(shared_agents)
        Dedicated = len(dedicated_agents)
        Other = len(other_agents)

        Al = (Answered / Total * 100) if Total else 0

        Total_login = Talk + wait + dispo + pause + hold
        Net_login = Talk + wait + dispo + hold

        Utilization = (
            (Talk + dispo + hold) / Net_login * 100
            if Net_login else 0
        )


        agents_str = ",".join(agents)
        shared_str = ",".join(shared_agents)
        dedicated_str = ",".join(dedicated_agents)
        other_str = ",".join(other_agents)

        # ---------------- RL ----------------
        rl = db1.execute(
            text(f"""
            SELECT COUNT(1) cnt
            FROM aband_call_master
            WHERE {client_list_str}
              AND call_status='answer'
              AND calldate>='{start_time_start}'
              AND calldate<'{start_time_end}'
            """)
        ).scalar()

        # ---------------- Final Data ----------------
        # data[dateLabel] = {
        #     "Total": row.Total,
        #     "Answered": row.Answered,
        #     "Manpower": row.Manpower,
        #     "Shared": row.Shared,
        #     "Dedicated": row.Dedicated,
        #     "Other": row.Other,
        #     "Talk": row.Talk,
        #     "wait": row.wait,
        #     "dispo": row.dispo,
        #     "hold": row.hold,
        #     "Al %": round(row.Al, 2),
        #     "SL %": round((row.WIthinSLA / row.Answered) * 100, 2) if row.Answered else 0,
        #     "RL": rl,
        #     "RL %": round(((rl + row.Answered) / row.Total) * 100, 2) if row.Total else 0,
        #     "Total login": row.Total_login,
        #     "Net login": row.Net_login,
        #     "Utilization %": round(row.Utilization, 2),
        #     "Manpower Agents": map_agents(row.agents),
        #     "Shared Agents": map_agents(row.Shared_ag),
        #     "Dedicated Agents": map_agents(row.Dedicated_ag),
        #     "Other Agents": map_agents(row.Other_ag),
        #     "WIthinSLA": row.WIthinSLA,
        # }


        data[dateLabel] = {
            "Total": Total,
            "Answered": Answered,
            "Manpower": Manpower,
            "Shared": Shared,
            "Dedicated": Dedicated,
            "Other": Other,
            "Talk": Talk,
            "wait": wait,
            "dispo": dispo,
            "hold": hold,
            "Al %": round(Al, 2),
            "SL %": round((WithinSLA / Answered) * 100, 2) if Answered else 0,
            "RL": rl,
            "RL %": round(((rl + Answered) / Total) * 100, 2) if Total else 0,
            "Total login": Total_login,
            "Net login": Net_login,
            "Utilization %": round(Utilization, 2),

            "Manpower Agents": map_agents(agents_str),
            "Shared Agents": map_agents(shared_str),
            "Dedicated Agents": map_agents(dedicated_str),
            "Other Agents": map_agents(other_str),

            "WIthinSLA": WithinSLA,
        }

    return data













@router.get("/abandon-trend_old")
def abandon_trend_old(
    start_date: date,
    end_date: date,
    client_id: str = Query("All"),
    category: str = Query("All"),
    no_of_count: int = Query(100),
    db: Session = Depends(get_db4),   # registration_master
    db2: Session = Depends(get_db3)   # vicidial
) -> Dict[str, Any]:

    # ---------------- 🔹 STEP 1: GET CAMPAIGN IDs ----------------
    if client_id != "All":
        camp = db.execute(
            text("SELECT campaignid FROM registration_master WHERE company_id = :cid"),
            {"cid": client_id}
        ).scalar_one_or_none()

        if not camp:
            raise HTTPException(404, "Company ID not found")

        campaign_list = [c.strip().strip("'") for c in camp.split(",") if c.strip()]

    else:
        if category == "All":
            rows = db.execute(text("""
                SELECT campaignid 
                FROM registration_master 
                WHERE status='A' AND is_dd_client='1'
            """)).fetchall()
        else:
            rows = db.execute(text("""
                SELECT campaignid 
                FROM registration_master 
                WHERE status='A' 
                AND is_dd_client='1'
                AND client_category = :category
            """), {"category": category}).fetchall()

        campaign_list = []
        for row in rows:
            if row[0]:
                campaign_list.extend([c.strip().strip("'") for c in row[0].split(",")])

    # 🔴 Safety check
    campaign_list = list(set(campaign_list))  # remove duplicates

    if not campaign_list:
        return {"data": {}, "message": "No campaigns found"}

    # ---------------- 🔹 STEP 2: CAMPAIGN CONDITION ----------------
    campaign_condition = "AND t2.campaign_id IN :cids"
    base_params = {"cids": tuple(campaign_list)}

    # ---------------- 🔹 STEP 3: DATE LOOP ----------------
    current_date = datetime.combine(start_date, datetime.min.time())
    end_datetime = datetime.combine(end_date, datetime.max.time())

    data = {}
    date_array = []
    campaign_array = set()
    datetime_array = {}

    while current_date < end_datetime:

        start_time = current_date
        next_day = current_date + timedelta(days=1)

        date_label = start_time.strftime("%Y-%m-%d")
        date_array.append(date_label)

        # ---------------- 🔹 STEP 4: MAIN QUERY ----------------
        query = f"""
            SELECT 
                t2.campaign_id,
                COUNT(*) AS total,
                SUM(CASE WHEN t2.user != 'VDCL' THEN 1 ELSE 0 END) AS answered,
                SUM(CASE WHEN t2.user = 'VDCL' THEN 1 ELSE 0 END) AS abandon
            FROM asterisk.vicidial_closer_log t2
            LEFT JOIN asterisk.vicidial_agent_log t1 
                ON t1.uniqueid = t2.uniqueid 
                AND t1.lead_id != '' 
                AND t2.user = t1.user
            WHERE t2.call_date >= :start_time
            AND t2.call_date < :end_time
            {campaign_condition}
            GROUP BY t2.campaign_id
        """

        query_params = {
            "start_time": start_time,
            "end_time": next_day,
            **base_params
        }

        rows = db2.execute(text(query), query_params).fetchall()

        # ---------------- 🔹 STEP 5: PROCESS DATA ----------------
        for row in rows:
            campaign_id = row[0]
            answered = row[2] or 0
            abandon = row[3] or 0
            total = answered + abandon

            if total == 0:
                continue

            # 🔥 SAME PHP LOGIC
            if total <= no_of_count:
                abandon_percent = round((abandon * 100) / total)

                data.setdefault(date_label, {})
                data[date_label][campaign_id] = {
                    "abandon_percent": f"{abandon_percent}%"
                }

                campaign_array.add(campaign_id)
                datetime_array.setdefault(date_label, []).append(campaign_id)

        current_date = next_day

    # ---------------- 🔹 FINAL RESPONSE ----------------
    return {
        "data": data,
        "dates": sorted(set(date_array)),
        "campaigns": sorted(campaign_array),
        "datetime_map": datetime_array
    }






@router.get("/abandon-call_old")
def abandon_call_old(
    start_date: date,
    end_date: date,
    client_id: str = Query("All"),
    db: Session = Depends(get_db4),   # registration_master
    db2: Session = Depends(get_db3)   # vicidial DB
) -> Dict[str, Any]:

    # ---------------- 🔹 STEP 1: GET CAMPAIGNS ----------------
    if client_id != "All":
        camp = db.execute(
            text("SELECT campaignid FROM registration_master WHERE company_id = :cid"),
            {"cid": client_id}
        ).scalar_one_or_none()

        if not camp:
            raise HTTPException(404, "Company not found")

        campaign_list = [c.strip().strip("'") for c in camp.split(",") if c.strip()]

    else:
        rows = db.execute(text("""
            SELECT campaignid 
            FROM registration_master 
            WHERE status='A' AND is_dd_client='1'
        """)).fetchall()

        campaign_list = []
        for row in rows:
            if row[0]:
                campaign_list.extend([c.strip().strip("'") for c in row[0].split(",")])

    campaign_list = list(set(campaign_list))

    if not campaign_list:
        return {"data": {}, "message": "No campaigns found"}

    campaign_condition = "AND t2.campaign_id IN :cids"
    base_params = {"cids": tuple(campaign_list)}

    # ---------------- 🔹 STEP 2: DATE LOOP ----------------
    current_date = datetime.combine(start_date, datetime.min.time())
    end_datetime = datetime.combine(end_date, datetime.max.time())

    data = {}
    datetime_array = {}

    while current_date < end_datetime:

        start_time = current_date
        end_time = datetime.combine(current_date.date(), datetime.max.time())
        next_day = current_date + timedelta(days=1)

        # Labels (same as PHP)
        time_label = start_time.strftime("%d-%b-%Y")   # 01-Mar-2026
        date_label = start_time.strftime("%B-%Y")      # March-2026

        datetime_array.setdefault(date_label, []).append(time_label)

        # ---------------- 🔹 STEP 3: QUERY ----------------
        query = f"""
            SELECT 
                SUM(IF(t2.user='VDCL',1,0)) `Abandon`,
                SUM(IF(((t2.user='VDCL') AND (t2.queue_seconds IS NULL OR t2.queue_seconds<=10)),1,0)) `AbndWithinTen`,
                SUM(IF(((t2.user='VDCL') AND (t2.queue_seconds>10 and t2.queue_seconds<=15)),1,0)) `AbndWithinFif`,
                SUM(IF(((t2.user='VDCL') AND (t2.queue_seconds>15 and t2.queue_seconds<=20)),1,0)) `AbndWithinTwe`,
                SUM(IF(((t2.user='VDCL') AND (t2.queue_seconds>20 and t2.queue_seconds<=30)),1,0)) `AbndWithinThr`,
                SUM(IF(((t2.user='VDCL') AND (t2.queue_seconds IS NULL OR t2.queue_seconds>30)),1,0)) `AbndAftertThr`

            FROM asterisk.vicidial_closer_log t2
            LEFT JOIN asterisk.vicidial_agent_log t1 
                ON t1.uniqueid = t2.uniqueid 
                AND t1.lead_id != '' 
                AND t2.user = t1.user

            LEFT JOIN (
                SELECT uniqueid, SUM(parked_sec) p 
                FROM park_log 
                WHERE status='GRABBED'
                AND parked_time >= :start_time
                AND parked_time < :end_time
                GROUP BY uniqueid
            ) t3 ON t1.uniqueid = t3.uniqueid

            WHERE t2.call_date >= :start_time
            AND t2.call_date < :end_time
            {campaign_condition}
        """

        params = {
            "start_time": start_time,
            "end_time": end_time,
            **base_params
        }

        result = db2.execute(text(query), params).fetchone()

        # ---------------- 🔹 STEP 4: STORE DATA ----------------
        if result:
            data.setdefault("Abandon Call", {}).setdefault(date_label, {})[time_label] = result[0] or 0
            data.setdefault("Abandon in 10 Sec", {}).setdefault(date_label, {})[time_label] = result[1] or 0
            data.setdefault("Abandon in 15 Sec", {}).setdefault(date_label, {})[time_label] = result[2] or 0
            data.setdefault("Abandon in 20 Sec", {}).setdefault(date_label, {})[time_label] = result[3] or 0
            data.setdefault("Abandon in 30 Sec", {}).setdefault(date_label, {})[time_label] = result[4] or 0
            data.setdefault("Abandon After 30 Sec", {}).setdefault(date_label, {})[time_label] = result[5] or 0

        current_date = next_day

    # ---------------- 🔹 FINAL RESPONSE ----------------
    return {
        "data": data,
        "datetime_map": datetime_array
    }











from fastapi import FastAPI, Query, HTTPException
from fastapi.responses import StreamingResponse
from typing import Optional
import mysql.connector
from mysql.connector import Error
import io
from datetime import date, datetime, time
import openpyxl
from openpyxl import Workbook
from contextlib import contextmanager



# ─────────────────────────────────────────────
# DB CONFIG
# ─────────────────────────────────────────────
DIALDESK_DB = dict(host="192.168.10.12", user="root", password="dial@mas123", database="db_dialdesk")
ASTERISK_DB  = dict(host="192.168.10.21", user="root", password="vicidialnow", database="asterisk")
VICIDIAL_DB  = dict(host="192.168.10.21", user="root", password="vicidialnow", database="asterisk")  # adjust if different


@contextmanager
def get_db(config: dict):
    conn = mysql.connector.connect(**config)
    try:
        yield conn
    finally:
        conn.close()


def query(conn, sql: str, params=None):
    cur = conn.cursor()
    cur.execute(sql, params or ())
    rows = cur.fetchall()
    cur.close()
    return rows


def query_one(conn, sql: str, params=None):
    rows = query(conn, sql, params)
    return rows[0] if rows else None


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def sec_convert(sec) -> str:
    """Convert seconds → H:MM:SS string."""
    try:
        sec = int(sec or 0)
    except (ValueError, TypeError):
        sec = 0
    h = sec // 3600
    m = (sec % 3600) // 60
    s = sec % 60
    return f"{h}:{m:02d}:{s:02d}"


def safe_div(a, b, decimals=2):
    try:
        return round(a / b * 100, decimals) if b else 0.0
    except Exception:
        return 0.0




def get_unit_agents(agent_type: Optional[str], process: Optional[str] = None):
    """Return agent metadata + tagging counts using SQLAlchemy."""
    print("🔌 Connecting to DB (SQLAlchemy)...")
    engine = get_engine4()

    with engine.connect() as conn:
        print("➡️ Running query")

        sql = "SELECT * FROM agent_master WHERE status = 'A'"
        params = {}

        if agent_type and agent_type != "All":
            sql += " AND agent_type = :agent_type"
            params["agent_type"] = agent_type

        if process and process != "All":
            sql += " AND processname = :process"
            params["process"] = process

        result = conn.execute(text(sql), params)
        agents = [dict(row._mapping) for row in result]

    # ───────── Process Data ─────────
    agent_type_list = {}
    agents_crm = {}
    ag_list2 = {}

    for ag in agents:
        user = ag["username"]
        agent_type_list[user] = ag["agent_type"]
        agents_crm[user] = ag
        ag_list2[int(ag["id"])] = user

    print(f"✅ Total agents fetched: {len(agents)}")
    return agent_type_list, agents_crm, ag_list2


def get_tag_counts(from_date: str, to_date: str, ag_list2: dict):
    engine = get_engine4()
    with engine.connect() as conn:
        result = conn.execute(text("""
            SELECT date(calldate) dater, AgentId, count(1) tagging
            FROM call_master
            WHERE date(calldate) BETWEEN :from_date AND :to_date
            GROUP BY date(calldate), AgentId
        """), {"from_date": from_date, "to_date": to_date})

        rows = result.fetchall()
    tag_count_list: dict = {}
    for dater, agent_id, tagging in rows:
        agent_id = int(agent_id) if agent_id is not None else None
        username = ag_list2.get(agent_id)
        if username:
            tag_count_list.setdefault(str(dater), {})[username] = tagging
    print(tag_count_list)
    return tag_count_list




def get_agent_log_data(from_dt: str, to_dt: str, unit_users: list, vicidial_db_config: dict):

    engine2 = get_engine3()

    sql = """
        SELECT `user`, DATE(event_time) AS dater,
        IF(wait_sec>10000,0,wait_sec) AS wait_sec,
        talk_sec, dispo_sec,
        IF(pause_sec>10000,0,pause_sec) AS pause_sec,
        lead_id, status, dead_sec, sub_status
        FROM vicidial_agent_log
        WHERE event_time <= :to_dt AND event_time >= :from_dt
    """

    params = {
        "to_dt": to_dt,
        "from_dt": from_dt
    }

    if unit_users:
        sql += " AND `user` IN :users"
        params["users"] = tuple(unit_users)

    with engine2.connect() as conn:
        result = conn.execute(text(sql), params)
        rows = result.fetchall()

    # aggregation dicts
    swait  = {}; stalk  = {}; sdispo = {}; spause = {}; sdead = {}; scust = {}
    scalls = {}; stime  = {}
    launch = {}; bio    = {}; tea    = {}; oper   = {}; qual  = {}; refr  = {}; train = {}
    transfer = {}
    date_list            = {}
    date_wise_user_list  = {}

    for row in rows:
        user, dater, wait, talk, dispo, pause, lead, status, dead, sub_stat = row
        dater = str(dater)
        wait  = min(int(wait  or 0), 65000)
        talk  = min(int(talk  or 0), 65000)
        dispo = min(int(dispo or 0), 65000)
        pause = min(int(pause or 0), 65000)
        dead  = min(int(dead  or 0), 65000)

        date_list[dater] = dater
        date_wise_user_list.setdefault(dater, {})[user] = user

        key = (dater, user)
        swait[key]  = swait.get(key, 0)  + wait
        stalk[key]  = stalk.get(key, 0)  + talk
        sdispo[key] = sdispo.get(key, 0) + dispo
        spause[key] = spause.get(key, 0) + pause
        sdead[key]  = sdead.get(key, 0)  + dead

        if lead and status and status.upper() != "NULL":
            scalls[key] = scalls.get(key, 0) + 1

        sub2 = (sub_stat or "").strip().replace(" ", "")

        if sub2 == "Lunch":
            launch[key] = launch.get(key, 0) + pause
        elif sub2 == "Bio":
            bio[key]    = bio.get(key, 0)    + pause
        elif sub2 in ("Short", "Tea"):
            tea[key]    = tea.get(key, 0)    + pause
        elif sub2.lower() == "oper":
            oper[key]   = oper.get(key, 0)   + pause
        elif sub2.lower() == "qualit":
            qual[key]   = qual.get(key, 0)   + pause
        elif sub2.lower() == "refres":
            refr[key]   = refr.get(key, 0)   + pause
        elif sub2.lower() == "traing":
            train[key]  = train.get(key, 0)  + pause

        if (status or "").lower().strip() == "xfer":
            transfer[key] = transfer.get(key, 0) + 1

    return dict(
        swait=swait, stalk=stalk, sdispo=sdispo, spause=spause, sdead=sdead,
        scalls=scalls, launch=launch, bio=bio, tea=tea, oper=oper, qual=qual,
        refr=refr, train=train, transfer=transfer,
        date_list=date_list, date_wise_user_list=date_wise_user_list,
    )



def get_park_data(from_dt: str, to_dt: str, vicidial_db_config: dict):

    sql = """
        SELECT user,
               DATE(parked_time) AS dater,
               COUNT(*) AS cnt,
               SUM(parked_sec) AS secs
        FROM park_log
        WHERE parked_time <= :to_dt
          AND parked_time >= :from_dt
        GROUP BY user, DATE(parked_time)
    """

    park_date: dict = {}
    park_user: dict = {}

    engine2 = get_engine3()

    with engine2.connect() as conn:
        result = conn.execute(
            text(sql),
            {"to_dt": to_dt, "from_dt": from_dt}
        )
        rows = result.fetchall()

    for user, dater, cnt, secs in rows:
        dater = str(dater)

        park_date.setdefault(dater, {}).setdefault(user, [0, 0])
        park_date[dater][user][0] += cnt
        park_date[dater][user][1] += (secs or 0)

        park_user.setdefault(user, [0, 0])
        park_user[user][0] += cnt
        park_user[user][1] += (secs or 0)

    return park_date, park_user




def get_login_logout(user: str, from_dt: str, to_dt: str, vicidial_db_config: dict, date_filter: Optional[str] = None):

    engine2 = get_engine3()

    with engine2.connect() as conn:

        if date_filter:
            login_sql = """
                SELECT TIME(MIN(event_date))
                FROM vicidial_user_log
                WHERE user = :user
                  AND event = 'Login'
                  AND DATE(event_date) = :date_filter
            """

            logout_sql = """
                SELECT TIME(MAX(event_date))
                FROM vicidial_user_log
                WHERE user = :user
                  AND event = 'LOGOUT'
                  AND DATE(event_date) = :date_filter
            """

            params = {
                "user": user,
                "date_filter": date_filter
            }

        else:
            login_sql = """
                SELECT TIME(MIN(event_date))
                FROM vicidial_user_log
                WHERE user = :user
                  AND event = 'Login'
                  AND event_date BETWEEN :from_dt AND :to_dt
            """

            logout_sql = """
                SELECT TIME(MAX(event_date))
                FROM vicidial_user_log
                WHERE user = :user
                  AND event = 'LOGOUT'
                  AND event_date BETWEEN :from_dt AND :to_dt
            """

            params = {
                "user": user,
                "from_dt": from_dt,
                "to_dt": to_dt
            }

        login_row = conn.execute(text(login_sql), params).fetchone()
        logout_row = conn.execute(text(logout_sql), params).fetchone()

    login_t  = str(login_row[0])  if login_row and login_row[0] else "00:00:00"
    logout_t = str(logout_row[0]) if logout_row and logout_row[0] else "00:00:00"

    return login_t, logout_t


# ─────────────────────────────────────────────
# REPORT BUILDER
# ─────────────────────────────────────────────

def build_report(
    query_date: str,
    end_date: str,
    agent_type: Optional[str] = None,
    process: Optional[str] = None,
    shift: str = "ALL",
    start_timexx: str = "00:00:00",
    end_timexx: str = "23:59:59",
    vicidial_db_config: Optional[dict] = None,
):
    if vicidial_db_config is None:
        vicidial_db_config = VICIDIAL_DB

    # --- time range ---
    shift_map = {
        "TEST":  ("09:45:00", "10:00:00"),
        "AM":    ("03:45:00", "15:14:59"),
        "PM":    ("15:15:00", "23:15:00"),
        "ALL":   ("00:00:00", "23:59:59"),
        "9AM-5PM":     ("09:00:00", "16:59:59"),
        "5PM-MIDNIGHT":("17:00:00", "23:59:59"),
    }
    t_begin, t_end = shift_map.get(shift, ("00:00:00", "23:59:59"))
    from_dt = f"{query_date} {t_begin}"
    to_dt   = f"{end_date} {t_end}"

    # --- agents ---
    print("STEP 1: agents")
    agent_type_list, agents_crm, ag_list2 = get_unit_agents(agent_type, process)
    unit_users = list(agents_crm.keys())          # usernames (lowercase)

    # tag counts
    print("STEP 2: tags")
    tag_count_list = get_tag_counts(query_date, end_date, ag_list2)

    # agent log
    print("STEP 3: agent log")
    log = get_agent_log_data(from_dt, to_dt, unit_users, vicidial_db_config)

    # park
    park_date, park_user = get_park_data(from_dt, to_dt, vicidial_db_config)

    # ── aggregate per-user summary ──
    user_summary: dict = {}
    date_list = sorted(log["date_list"].keys())

    for dater in date_list:
        print(dater)
        for user in log["date_wise_user_list"].get(dater, {}):
            k = (dater, user)
            stalk   = log["stalk"].get(k, 0)
            sdispo  = log["sdispo"].get(k, 0)
            spause  = log["spause"].get(k, 0)
            swait   = log["swait"].get(k, 0)
            launch  = log["launch"].get(k, 0)
            bio     = log["bio"].get(k, 0)
            tea     = log["tea"].get(k, 0)
            oper    = log["oper"].get(k, 0)
            qual    = log["qual"].get(k, 0)
            refr    = log["refr"].get(k, 0)
            train   = log["train"].get(k, 0)
            calls   = log["scalls"].get(k, 0)
            xfer    = log["transfer"].get(k, 0)
            park_secs = (park_date.get(dater, {}).get(user) or [0, 0])[1]

            stime = swait + stalk + sdispo + spause
            productive_login = stime - launch - bio - tea

            if user not in user_summary:
                user_summary[user] = dict(
                    stalk=0, sdispo=0, spause=0, swait=0,
                    launch=0, bio=0, tea=0, oper=0, qual=0, refr=0, train=0,
                    calls=0, xfer=0, stime=0, productive_login=0,
                    park_secs=0, tags=0, mandays=0,
                )

            s = user_summary[user]
            s["stalk"]           += stalk
            s["sdispo"]          += sdispo
            s["spause"]          += spause
            s["swait"]           += swait
            s["launch"]          += launch
            s["bio"]             += bio
            s["tea"]             += tea
            s["oper"]            += oper
            s["qual"]            += qual
            s["refr"]            += refr
            s["train"]           += train
            s["calls"]           += calls
            s["xfer"]            += xfer
            s["stime"]           += stime
            s["productive_login"]+= productive_login
            s["park_secs"]       += park_secs
            s["tags"]            += tag_count_list.get(dater, {}).get(user, 0)

            login_t, _ = get_login_logout(user, from_dt, to_dt, vicidial_db_config, dater)
            if login_t != "00:00:00":
                s["mandays"] += 1

    # ── build rows ──
    summary_rows  = []
    raw_data_rows = []

    summary_header = [
        "Agent Type", "Process", "Agent ID", "Agent Name",
        "Calls", "ACHT", "Talktime", "Park Time", "Transfer Count",
        "Net Login", "Productive Login", "Lunch", "Bio", "Tea/Short",
        "Operation", "Quality", "Refresher", "Training",
        "Quality Score", "Utilization%",
        # "Utilization OP%",
        "First Login", "Last Logout", "Tagging no", "Mandays",
    ]

    raw_header = ["Date"] + summary_header[:-1]  # no Mandays in raw

    summary_rows.append(summary_header)
    raw_data_rows.append(raw_header)

    # per-date raw rows
    for dater in date_list:
        for user in log["date_wise_user_list"].get(dater, {}):

            print("TAG USERS:", list(tag_count_list.get(dater, {}).keys())[:5])
            print("LOG USER:", user)
            k     = (dater, user)
            stalk = log["stalk"].get(k, 0)
            sdispo= log["sdispo"].get(k, 0)
            spause= log["spause"].get(k, 0)
            swait = log["swait"].get(k, 0)
            calls = log["scalls"].get(k, 0)
            launch= log["launch"].get(k, 0)
            bio   = log["bio"].get(k, 0)
            tea   = log["tea"].get(k, 0)
            oper  = log["oper"].get(k, 0)
            qual  = log["qual"].get(k, 0)
            refr  = log["refr"].get(k, 0)
            train = log["train"].get(k, 0)
            xfer  = log["transfer"].get(k, 0)
            park_secs = (park_date.get(dater, {}).get(user) or [0, 0])[1]

            stime = swait + stalk + sdispo + spause
            productive_login = stime - launch - bio - tea
            acht  = round((stalk + sdispo) / calls) if calls else 0
            util  = safe_div(stalk, productive_login) if productive_login else 0
            util2 = safe_div(stalk, stime) if stime else 0
            tags  = tag_count_list.get(dater, {}).get(user, 0)

            login_t, logout_t = get_login_logout(user, from_dt, to_dt, vicidial_db_config, dater)
            agent = agents_crm.get(user, {})

            raw_data_rows.append([
                dater,
                agent.get("agent_type", ""),
                agent.get("processname", ""),
                user,
                agent.get("displayname", ""),
                calls,
                acht,
                sec_convert(stalk),
                sec_convert(park_secs),
                xfer,
                sec_convert(stime),
                sec_convert(productive_login),
                sec_convert(launch),
                sec_convert(bio),
                sec_convert(tea),
                sec_convert(oper),
                sec_convert(qual),
                sec_convert(refr),
                sec_convert(train),
                "0",
                # util2,
                util,
                login_t,
                logout_t,
                tags,
            ])

    # per-user summary rows
    tot = dict(calls=0, stalk=0, sdispo=0, launch=0, bio=0, tea=0, oper=0,
               qual=0, refr=0, train=0, xfer=0, stime=0, productive_login=0,
               park_secs=0, tags=0, mandays=0)

    for user, s in user_summary.items():
        calls = s["calls"]
        stalk = s["stalk"]
        sdispo= s["sdispo"]
        acht  = round((stalk + sdispo) / calls) if calls else 0
        util  = safe_div(stalk, s["productive_login"]) if s["productive_login"] else 0
        util2 = safe_div(stalk, s["stime"]) if s["stime"] else 0

        login_t, logout_t = get_login_logout(user, from_dt, to_dt, vicidial_db_config)
        agent = agents_crm.get(user, {})

        summary_rows.append([
            agent.get("agent_type", ""),
            agent.get("processname", ""),
            user,
            agent.get("displayname", ""),
            calls,
            acht,
            sec_convert(stalk),
            sec_convert(s["park_secs"]),
            s["xfer"],
            sec_convert(s["stime"]),
            sec_convert(s["productive_login"]),
            sec_convert(s["launch"]),
            sec_convert(s["bio"]),
            sec_convert(s["tea"]),
            sec_convert(s["oper"]),
            sec_convert(s["qual"]),
            sec_convert(s["refr"]),
            sec_convert(s["train"]),
            "0",
            # util2,
            util,
            login_t,
            logout_t,
            s["tags"],
            s["mandays"],
        ])

        for key in tot:
            tot[key] += s.get(key, 0)

    # totals row
    summary_rows.append([
        "", "", "", "Total",
        tot["calls"], "00:00:00",
        sec_convert(tot["stalk"]), sec_convert(tot["park_secs"]),
        tot["xfer"], sec_convert(tot["stime"]),
        sec_convert(tot["productive_login"]),
        sec_convert(tot["launch"]), sec_convert(tot["bio"]),
        sec_convert(tot["tea"]),    sec_convert(tot["oper"]),
        sec_convert(tot["qual"]),   sec_convert(tot["refr"]),
        sec_convert(tot["train"]),  "0",
        "00:00:00", "00:00:00", "00:00:00",
        tot["tags"], tot["mandays"],
    ])
    raw_data_rows.append([
        "", "", "", "", "Total",
        tot["calls"], "",
        sec_convert(tot["stalk"]), sec_convert(tot["park_secs"]),
        tot["xfer"], sec_convert(tot["stime"]),
        sec_convert(tot["productive_login"]),
        sec_convert(tot["launch"]), sec_convert(tot["bio"]),
        sec_convert(tot["tea"]),    sec_convert(tot["oper"]),
        sec_convert(tot["qual"]),   sec_convert(tot["refr"]),
        sec_convert(tot["train"]),  "0",
        "00:00:00", "00:00:00","00:00:00", tot["tags"],
    ])

    return summary_rows, raw_data_rows


# ─────────────────────────────────────────────
# API ENDPOINTS
# ─────────────────────────────────────────────

@router.get("/apr-report/json_old", summary="APR Report – JSON")
def apr_report_json_old(
    query_date:  str = Query(...,  example="2024-01-01", description="Start date YYYY-MM-DD"),
    end_date:    str = Query(...,  example="2024-01-31", description="End date   YYYY-MM-DD"),
    agent_type:  Optional[str] = Query(None, description="Filter by agent type; omit for all"),
    process:     Optional[str] = Query(None),
    shift:       str = Query("ALL", description="Shift: ALL | AM | PM | TEST | 9AM-5PM | 5PM-MIDNIGHT"),
):
    """Return both Summary and Raw Data as JSON."""
    try:
        print("started")
        summary_rows, raw_data_rows = build_report(query_date, end_date, agent_type, process, shift)
        print("build_report done")
    except Exception as e:
        import traceback
        print("🔥 ERROR OCCURRED:")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

    # convert to list-of-dicts using first row as header
    def rows_to_dicts(rows):
        if len(rows) < 2:
            return []
        header = rows[0]
        return [dict(zip(header, r)) for r in rows[1:]]

    return {
        "summary":  rows_to_dicts(summary_rows),
        "raw_data": rows_to_dicts(raw_data_rows),
        "generated_at": datetime.now().isoformat(),
        "params": {
            "query_date": query_date,
            "end_date":   end_date,
            "agent_type": agent_type,
            "shift":      shift,
        },
    }


@router.get("/apr-report/xlsx_old", summary="APR Report – Excel download")
def apr_report_xlsx_old(
    query_date:  str = Query(...,  example="2024-01-01"),
    end_date:    str = Query(...,  example="2024-01-31"),
    agent_type:  Optional[str] = Query(None),
    process:     Optional[str] = Query(None),
    shift:       str = Query("ALL"),
):
    """Download the APR report as an Excel (.xlsx) file with two sheets."""
    try:
        summary_rows, raw_data_rows = build_report(query_date, end_date, agent_type, process, shift)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    for row in summary_rows:
        ws_summary.append([str(c) if c is not None else "" for c in row])

    ws_raw = wb.create_sheet(title="Raw Data")
    for row in raw_data_rows:
        ws_raw.append([str(c) if c is not None else "" for c in row])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"APR_Report_{query_date}_{end_date}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )








class CorrectiveReportRequest(BaseModel):
    start_date: str  # format YYYY-MM-DD
    end_date: str    # format YYYY-MM-DD
    client_id: int

# --- Helper function to fetch calls from DB ---
def fetch_calls(client_id: int, start: datetime, end: datetime, db: Session):
    query = text("""
        SELECT *
        FROM call_master
        WHERE ClientId = :client_id
          AND DATE(CallDate) BETWEEN :start AND :end
        ORDER BY Category3 ASC
    """)
    result = db.execute(query, {"client_id": client_id, "start": start, "end": end}).fetchall()
    return [dict(row._mapping) for row in result]





@router.post("/corrective_report_old")
def corrective_report_old(
    request: CorrectiveReportRequest,
    db: Session = Depends(get_db4)
):

    # 1️⃣ Validate Dates
    try:
        start = datetime.strptime(request.start_date, "%Y-%m-%d").date()
        end = datetime.strptime(request.end_date, "%Y-%m-%d").date()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    if end < start:
        raise HTTPException(status_code=400, detail="end_date must be >= start_date")

    # 2️⃣ Fetch Data
    try:
        calls = fetch_calls(request.client_id, start, end, db)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"DB query failed: {str(e)}")

    # 3️⃣ Group Data → Category → Phase
    data_arr: Dict[str, Dict[str, Dict[str, int]]] = {}

    for call in calls:
        category = call.get("Category2") or "Undefined"
        phase = call.get("Category3") or "Undefined"
        status = call.get("CloseLoopCate1")

        if category not in data_arr:
            data_arr[category] = {}

        if phase not in data_arr[category]:
            data_arr[category][phase] = {
                "open": 0,
                "close": 0,
                "total": 0
            }

        if status == "Open":
            data_arr[category][phase]["open"] += 1
        else:
            data_arr[category][phase]["close"] += 1

        data_arr[category][phase]["total"] += 1

    # 4️⃣ Calculate Category Totals + Grand Total
    final_data = {}
    grand_open = 0
    grand_close = 0
    grand_total = 0

    for category, phases in data_arr.items():
        cat_open = 0
        cat_close = 0
        cat_total = 0

        for phase, values in phases.items():
            cat_open += values["open"]
            cat_close += values["close"]
            cat_total += values["total"]

        final_data[category] = {
            "phases": phases,
            "category_total": {
                "open": cat_open,
                "close": cat_close,
                "total": cat_total
            }
        }

        grand_open += cat_open
        grand_close += cat_close
        grand_total += cat_total

    # 5️⃣ Final Response
    return {
        "client_id": request.client_id,
        "start_date": request.start_date,
        "end_date": request.end_date,
        "grand_total": {
            "open": grand_open,
            "close": grand_close,
            "total": grand_total
        },
        "data": final_data
    }












class SLAClientwiseReq(BaseModel):
    from_date: date
    to_date: date
    company_id: str  # single or 'ALL'
    sd_type: Optional[str] = None  # "0" = Dedicated, "1" = Shared, None/All
    filter_type: Optional[str] = None  # "All", "SLA", "AL"


class SLAClientwiseRow(BaseModel):
    campaign_id: str
    Total: int
    TalkTime: str
    dispo_time: str
    WrapTime: str
    Answered: int
    Abandon: int
    TotalAcht: int
    WIthinSLA: int
    WIthinSLATen: int
    AbndWithinThresold: int
    AbndAfterThresold: int

class SLAClientwiseResp(BaseModel):
    rows: List[SLAClientwiseRow]






# def timedelta_to_seconds(td):
#     if td is None:
#         return 0
#     return int(td.total_seconds())



def timedelta_to_seconds(td):
    if isinstance(td, int):  # ✅ already seconds
        return td

    if isinstance(td, str):  # ✅ "HH:MM:SS"
        h, m, s = map(int, td.split(":"))
        return h * 3600 + m * 60 + s

    # ✅ timedelta case
    return int(td.total_seconds())


def seconds_to_hms(seconds):
    seconds = int(seconds)  # convert float to int
    h = seconds // 3600
    m = (seconds % 3600) // 60
    s = seconds % 60
    return f"{h:02d}:{m:02d}:{s:02d}"


def sec_to_time(seconds):
    h = int(seconds // 3600)
    m = int((seconds % 3600) // 60)
    s = int(seconds % 60)
    return f"{h:02}:{m:02}:{s:02}"






@router.post("/sla_clientwise_report_excel_old")
def sla_clientwise_report_excel_old(req: SLAClientwiseReq, db2: Session = Depends(get_db3), db: Session = Depends(get_db4)):

    params_shared = {}
    shared_clause = ""
    if str(req.sd_type) in ("0", "1"):
        shared_clause = "AND is_shared = :is_shared"
        params_shared["is_shared"] = req.sd_type

    # Fetch campaigns
    if req.company_id.upper() == "ALL":
        rows = db.execute(
            text(f"""
                SELECT campaignid, company_id, company_name
                FROM registration_master
                WHERE status='A'
                  AND is_dd_client='1'
                  {shared_clause}
                ORDER BY company_id ASC
            """),
            params_shared
        ).fetchall()
    else:
        rows = db.execute(
            text(f"""
                SELECT campaignid, company_id, company_name
                FROM registration_master
                WHERE company_id = :cid
                  AND status='A'
                  AND is_dd_client='1'
                  {shared_clause}
            """),
            {**params_shared, "cid": req.company_id}
        ).fetchall()

    if not rows:
        raise HTTPException(404, "No campaigns found")

    # Flatten campaigns per company
    client_campaigns = {}
    for r in rows:
        company_id = r.company_id
        company_name = r.company_name
        if r.campaignid:
            campaigns = [c.strip().strip("'") for c in r.campaignid.split(",") if c.strip()]
            # client_campaigns[company_id] = campaigns
        else:
            campaigns = []
            # client_campaigns[company_id] = []

        client_campaigns[company_id] = {
            "company_name": company_name,
            "campaigns": campaigns
        }

    # Fetch plan rates per company
    company_rates = {}
    for company_id in client_campaigns.keys():
        balance = db.execute(
            text("SELECT PlanId FROM balance_master WHERE clientId=:cid LIMIT 1"),
            {"cid": company_id}
        ).fetchone()
        if balance:
            plan = db.execute(
                text("""
                    SELECT InboundCallCharge, rate_per_pulse_day_shift, pulse_day_shift
                    FROM plan_master
                    WHERE id = :plan_id LIMIT 1
                """), {"plan_id": balance.PlanId}
            ).fetchone()
            if plan:
                pulse_day_shift = int(plan.pulse_day_shift or 0)
                rate_per_pulse_day_shift = float(plan.rate_per_pulse_day_shift or 0)
                inbound_charge = float(plan.InboundCallCharge or 0)
                rate_per_sec = rate_per_pulse_day_shift / pulse_day_shift if pulse_day_shift > 0 else 0
                company_rates[company_id] = {
                    "InboundCallCharge": inbound_charge,
                    "rate_per_sec": rate_per_sec
                }
            else:
                company_rates[company_id] = {"InboundCallCharge": 0, "rate_per_sec": 0}
        else:
            company_rates[company_id] = {"InboundCallCharge": 0, "rate_per_sec": 0}


    # Prepare data
    all_data = {}
    total_handle = 0
    total_acht = 0
    total_answered = 0
    total_offered = 0
    # rl_list = set()


    grand_totals = {
        "Offered": 0,
        "Handled": 0,
        "Total Talk Time": "00:00:00",
        "Calls Ans (20 Sec)": 0,
        "Calls Ans (10 Sec)": 0,
        "Total Calls Abandoned": 0,
        "Abnd Within (20)": 0,
        "AHT_total_sec": 0,
        "Amount": 0,
        "RL": 0,
        "Call Rate": 0
    }   
        

    for company_id, info in client_campaigns.items():
        # ---------------- RL & RL% calculation ----------------

        rl_sql = """
            SELECT COUNT(1) AS cnt
            FROM aband_call_master
            WHERE ClientId = :cid
            AND call_status = 'answer'
            AND DATE(calldate) BETWEEN :from_date AND :to_date
        """

        rl_row = db.execute(
            text(rl_sql),
            {
                "cid": company_id,
                "from_date": req.from_date,
                "to_date": req.to_date
            }
        ).fetchone()

        company_rl = int(rl_row.cnt or 0)

        company_name = info["company_name"]
        campaigns = info["campaigns"]

        if company_name not in all_data:
            all_data[company_name] = {
                "Client Name": company_name,
                "Offered": 0,
                "Handled": 0,
                "Calls Ans (20 Sec)": 0,
                "Total Calls Abandoned": 0,
                "Abnd Within (20)": 0,
                "Total Talk Time_sec": 0,
                "AHT_total": 0,
                "RL": company_rl
            }

        rate_info = company_rates[company_id]
        for campaign in campaigns:
            sql = f"""
                SELECT 
                    t2.campaign_id,
                    t2.user,
                    t2.queue_seconds,
                    dispo_sec,
                    talk_sec,
                    pause_sec,
                    wait_sec,
                    t1.length_in_sec
                FROM asterisk.vicidial_closer_log t2
                LEFT JOIN asterisk.vicidial_users vu ON t2.user = vu.user
                LEFT JOIN asterisk.call_log t1 ON t1.uniqueid = t2.uniqueid
                LEFT JOIN asterisk.vicidial_agent_log t3 ON t2.uniqueid = t3.uniqueid AND t2.user = t3.user
                WHERE DATE(t2.call_date) BETWEEN :from_date AND :to_date
                  AND t2.term_reason <> 'AFTERHOURS'
                  AND t2.lead_id IS NOT NULL
                  AND t2.campaign_id = :campaign
            """
            row = db2.execute(text(sql), {"from_date": req.from_date, "to_date": req.to_date, "campaign": campaign}).mappings().all()

            # if not row:
            #     row = {k: 0 for k in ["Total","Answered","Abandon","TotalAcht","WIthinSLA","WIthinSLATen","AbndWithinThresold"]}
            #     row["TalkTime"] = "00:00:00"

            if not row:
                row = []   # ✅ must be list


            total = 0
            answered = 0
            abandon = 0
            total_acht_row = 0

            within_sla = 0
            within_sla_ten = 0
            abnd_within = 0
            abnd_after = 0

            talk_time_sec = 0
            dispo_time_sec = 0
            wrap_time_sec = 0

            for r in row:
                total += 1

                user = r["user"]
                queue_sec = r["queue_seconds"] or 0
                dispo_sec = r["dispo_sec"] or 0
                talk_sec = r["talk_sec"] or 0
                pause_sec = r["pause_sec"] or 0
                wait_sec = r["wait_sec"] or 0
                length_sec = r["length_in_sec"] or 0

                # Time calculations
                talk_time_sec += (talk_sec + pause_sec + wait_sec + dispo_sec)
                dispo_time_sec += dispo_sec
                wrap_time_sec += dispo_sec

                if user != "VDCL":
                    answered += 1
                    total_acht_row += length_sec

                    if queue_sec <= 20:
                        within_sla += 1
                    if queue_sec <= 10:
                        within_sla_ten += 1
                else:
                    abandon += 1

                    if queue_sec <= 20:
                        abnd_within += 1
                    else:
                        abnd_after += 1

            # ✅ KEEP SAME VARIABLE NAMES USED BELOW
            offered = total
            handled = answered
            wi_thin_sla = within_sla
            wi_thin_sla_ten = within_sla_ten

            # abandon = float(row["Abandon"] or 0)
            # total_acht_row = float(row["TotalAcht"] or 0)
            # offered = float(row["Total"] or 0)
            # handled = float(row["Answered"] or 0)
            # wi_thin_sla = float(row["WIthinSLA"] or 0)
            # wi_thin_sla_ten = float(row["WIthinSLATen"] or 0)
            # abnd_within = float(row["AbndWithinThresold"] or 0)

            

            # RL %
            denominator = handled + abandon
            # rl_percent = round(((handled + rl) / denominator) * 100) if denominator > 0 else 0

            # -------- FILTER LOGIC --------
            # skips Offered that have 0
            # if req.filter_type == "without_0" and offered <= 0:
            #     continue

            



            data = all_data[company_name]

            data["Offered"] += offered
            data["Handled"] += handled
            data["Calls Ans (20 Sec)"] += wi_thin_sla
            data["Total Calls Abandoned"] += abandon
            data["Abnd Within (20)"] += abnd_within
            data["Total Talk Time_sec"] += talk_time_sec
            # data["RL"] += rl

            if handled > 0:
                data["AHT_total"] += total_acht_row


            # all_data[campaign] = {
            #     "Client Name": campaign,
            #     "Offered": offered,
            #     "Handled": handled,
            #     "SL% (20 Sec)": f"{round(wi_thin_sla*100/handled) if handled else 0}%",
            #     "AL": f"{round(handled*100/offered) if offered else 0}%",
            #     "Calls Ans (20 Sec)": wi_thin_sla,
            #     # "Calls Ans (10 Sec)": wi_thin_sla_ten,
            #     "Total Calls Abandoned": abandon,
            #     "Abnd Within (20)": abnd_within,
            #     "Total Talk Time": seconds_to_hms(talk_time_sec),
            #     "Average Aband Time": "",
            #     # "SL% (10 Sec)": f"{round(wi_thin_sla_ten*100/handled) if handled else 0}%",               
            #     "AHT (In Sec)": round(total_acht_row / handled) if handled else 0,
            #     # "Call Rate": rate_info["InboundCallCharge"],
            #     # "Amount": round(abandon * float(rate_info["rate_per_sec"]) * round((total_acht_row / handled if handled else 0)), 2),
            #     "RL": rl,
            #     "RL%": f"{rl_percent}%"
            # }

            # Accumulate totals safely as float
            total_handle += total_acht_row
            total_acht += total_acht_row
            total_answered += handled
            total_offered += offered

    if req.filter_type == "without_0":
        all_data = {
            k: v for k, v in all_data.items()
            if v["Offered"] > 0
        }



    for company_name, data in all_data.items():
        handled = data["Handled"]
        offered = data["Offered"]

        data["Total Talk Time"] = seconds_to_hms(data["Total Talk Time_sec"])

        data["SL% (20 Sec)"] = f"{round(data['Calls Ans (20 Sec)'] * 100 / handled) if handled else 0}%"
        data["AL"] = f"{round(handled * 100 / offered) if offered else 0}%"

        data["AHT (In Sec)"] = round(data["AHT_total"] / handled) if handled else 0

        denominator = handled + data["Total Calls Abandoned"]
        data["RL%"] = f"{round((handled + data['RL']) * 100 / denominator) if denominator else 0}%"


    # -------------------- Generate Excel --------------------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SLA Clientwise Report"

    # headers = [
    #     "Client Name", "Offered", "Handled", "Calls Ans (20 Sec)", "Calls Ans (10 Sec)",
    #     "Total Calls Abandoned", "Abnd Within (20)", "Average Aband Time", "Total Talk Time", "SL% (20 Sec)", 
    #     "SL% (10 Sec)", "AL", "AHT (In Sec)", "Call Rate", "Amount", "RL", "RL%"
    # ]

    headers = [
        "Client Name", "Offered", "Handled", "SL% (20 Sec)", "AL", "Calls Ans (20 Sec)",
        "Total Calls Abandoned", "Abnd Within (20)", "Total Talk Time",  
        "AHT (In Sec)", "RL", "RL%"
    ]
    fill = PatternFill(start_color="317EAC", end_color="317EAC", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for col_num, header in enumerate(headers,1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    last_data_row = 1

    sorted_data = dict(sorted(all_data.items(), key=lambda x: x[0]))

    for row_num, (company, data) in enumerate(sorted_data.items(), start=2):
        last_data_row = row_num        

        for col_num, key in enumerate(headers, start=1):
            ws.cell(row=row_num, column=col_num, value=data.get(key, ""))
        
        # ---- Accumulate totals ----
        offered = data["Offered"]
        handled = data["Handled"]

        grand_totals["Offered"] += offered
        grand_totals["Handled"] += handled

        grand_totals["Total Talk Time_sec"] = grand_totals.get("Total Talk Time_sec", 0)
        grand_totals["Total Talk Time_sec"] += timedelta_to_seconds(data["Total Talk Time"])

        grand_totals["Calls Ans (20 Sec)"] += data["Calls Ans (20 Sec)"]
        # grand_totals["Calls Ans (10 Sec)"] += data["Calls Ans (10 Sec)"]
        grand_totals["Total Calls Abandoned"] += data["Total Calls Abandoned"]
        grand_totals["Abnd Within (20)"] += data["Abnd Within (20)"]
        # grand_totals["Average Aband Time"] = 0
        # grand_totals["Call Rate"] += data["Call Rate"]
        # grand_totals["Amount"] += data["Amount"]
        grand_totals["RL"] += data["RL"]

        if handled > 0:
            grand_totals["AHT_total_sec"] += handled * data["AHT (In Sec)"]


    # -------- GRAND TOTAL RL% --------
    denominator = (
        grand_totals["Handled"] +
        grand_totals["Total Calls Abandoned"]
    )

    if denominator > 0:
        grand_totals["RL%"] = round(
            ((grand_totals["Handled"] + grand_totals["RL"]) / denominator) * 100)
    else:
        grand_totals["RL%"] = 0


    # -------- GRAND TOTAL SL% (20 Sec) and SL% (10 Sec) --------
    total_handled = grand_totals["Handled"]

    if total_handle > 0:
        grand_totals["SL% (20 Sec)"] = round(
            ((grand_totals["Calls Ans (20 Sec)"]*100)/total_handled)
            )
        grand_totals["SL% (10 Sec)"] = round(
            ((grand_totals["Calls Ans (10 Sec)"]*100)/total_handled)
            )
    else:
        grand_totals["SL% (20 Sec)"] = 0
        grand_totals["SL% (10 Sec)"] = 0

    # -------- GRAND TOTAL AL --------
    if total_offered > 0:
        grand_totals["AL"] = round(
            ((grand_totals["Handled"]*100)/grand_totals["Offered"]))
    else:
        grand_totals["AL"] = 0


    # Auto-width
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2


    grand_row = last_data_row + 1

    ws.cell(row=grand_row, column=1, value="Grand Total").font = Font(bold=True)

    col_map = {h: i + 1 for i, h in enumerate(headers)}

    ws.cell(row=grand_row, column=col_map["Offered"], value=grand_totals["Offered"])
    ws.cell(row=grand_row, column=col_map["Handled"], value=grand_totals["Handled"])
    ws.cell(row=grand_row, column=col_map["SL% (20 Sec)"], value=f'{grand_totals["SL% (20 Sec)"]}%')
    ws.cell(row=grand_row, column=col_map["AL"], value=f'{grand_totals["AL"]}%')
    ws.cell(row=grand_row, column=col_map["Total Talk Time"], value=seconds_to_hms(grand_totals["Total Talk Time_sec"]))
    ws.cell(row=grand_row, column=col_map["Calls Ans (20 Sec)"], value=grand_totals["Calls Ans (20 Sec)"])
    # ws.cell(row=grand_row, column=col_map["Calls Ans (10 Sec)"], value=grand_totals["Calls Ans (10 Sec)"])
    ws.cell(row=grand_row, column=col_map["Total Calls Abandoned"], value=grand_totals["Total Calls Abandoned"])
    ws.cell(row=grand_row, column=col_map["Abnd Within (20)"], value=grand_totals["Abnd Within (20)"])
    # ws.cell(row=grand_row, column=col_map["Average Aband Time"], value=grand_totals["Average Aband Time"])
    # ws.cell(row=grand_row, column=col_map["Call Rate"], value=grand_totals["Call Rate"])
    # ws.cell(row=grand_row, column=col_map["Amount"], value=round(grand_totals["Amount"], 2))
    ws.cell(row=grand_row, column=col_map["RL"], value=grand_totals["RL"])
    ws.cell(row=grand_row, column=col_map["RL%"], value=f'{grand_totals["RL%"]}%')
    
    # ws.cell(row=grand_row, column=col_map["SL% (10 Sec)"], value=f'{grand_totals["SL% (10 Sec)"]}%')
   


    # Weighted AHT
    if grand_totals["Handled"] > 0:
        ws.cell(
            row=grand_row,
            column=col_map["AHT (In Sec)"],
            value=round(grand_totals["AHT_total_sec"] / grand_totals["Handled"])
        )


    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"sla_clientwise_report_{req.from_date}_{req.to_date}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )












class DashboardReq(BaseModel):
    company_id: Union[int, str]
    from_date: date
    to_date: date
    sd_type: str
    category: str = "All"



class HourlyCampaignRow(BaseModel):
    Total: int
    Answered: int
    Abandon: int
    gdate: Optional[str] = None
    ghour: int
    campaign: Optional[str] = None
    SLA : Optional[int] = None

class HourlyCampaignALRow(BaseModel):
    Total: int
    Answered: int
    Abandon: int
    gdate: Optional[str] = None
    campaign: Optional[str] = None
    AL: float


class HourlyCampaignResp(BaseModel):
    rows: List[HourlyCampaignRow]
    al_rows: List[HourlyCampaignALRow]



# Return Customer daily data client wise
@router.post("/hourly_campaign_report_old", response_model=HourlyCampaignResp)
def get_hourly_campaign_report_old(
    req: DashboardReq,
    db: Session = Depends(get_db3),
    db_main: Session = Depends(get_db4),
) -> Any:

    params = {
        "from_date": req.from_date,
        "to_date": req.to_date,
    }

    # -------------------- Shared / Dedicated Logic --------------------
    shared_clause = ""
    params_shared = {}

    if str(req.sd_type) in ("0", "1"):
        shared_clause = "AND is_shared = :is_shared"
        params_shared["is_shared"] = req.sd_type



    # -------------------- Category Filter --------------------
    category_clause = ""
    if str(req.category) and str(req.category).upper() != "ALL":
        category_clause = "AND client_category = :category"
        params["category"] = req.category


    # 1) Fetch campaigns
    # -------------------- Fetch Campaigns --------------------
    if str(req.company_id).upper() == "ALL":
        rows = db_main.execute(
            text(f"""
                SELECT campaignid
                FROM registration_master
                WHERE status='A'
                  AND is_dd_client='1'
                  {shared_clause}
                  {category_clause}
                ORDER BY campaignid ASC
            """),
            {**params, **params_shared}
        ).fetchall()
        print(rows)

        if not rows:
            raise HTTPException(404, "No campaigns found")

        campaign_list = []
        for r in rows:
            if r[0]:  # skip None
                campaign_list.extend([c.strip().strip("'") for c in r[0].split(",")])

    else:
        # Single Company
        rows = db_main.execute(
            text(f"""
                SELECT campaignid
                FROM registration_master
                WHERE company_id = :cid
                  AND status='A'
                  AND is_dd_client='1'
                  {shared_clause}
                  {category_clause}
                ORDER BY campaignid ASC
            """),
            {**params, **params_shared, "cid": req.company_id}
        ).fetchall()

        if not rows:
            raise HTTPException(404, "Company ID not found")

        campaign_list = []
        for r in rows:
            if r[0]:  # skip None
                campaign_list.extend([c.strip().strip("'") for c in r[0].split(",")])

    campaign_list = list(set(campaign_list))  # remove duplicates

    params["cids"] = tuple(campaign_list)
    camp_clause = "AND t2.campaign_id IN :cids"

    # 2) SQL
    sql = f"""
        SELECT
            COUNT(*)                         AS Total,
            SUM(IF(t2.user <> 'VDCL',1,0))   AS Answered,
            SUM(IF(t2.user = 'VDCL',1,0))    AS Abandon,
            HOUR(t2.call_date)               AS ghour,
            t2.campaign_id                   AS campaign,
            SUM(IF(t2.`user` !='VDCL' AND t2.queue_seconds<=20,1,0)) `SLA`
        FROM asterisk.vicidial_closer_log t2
        LEFT JOIN asterisk.vicidial_agent_log t3
               ON t2.uniqueid = t3.uniqueid
              AND t2.user     = t3.user
              AND t2.lead_id = t3.lead_id
        WHERE DATE(t2.call_date) BETWEEN :from_date AND :to_date
          {camp_clause}
          AND t2.term_reason <> 'AFTERHOURS'
          AND t2.lead_id IS NOT NULL
        GROUP BY
            t2.campaign_id,
            HOUR(t2.call_date)
        ORDER BY
            t2.campaign_id,
            ghour
            
    """

    rows = db.execute(text(sql), params).mappings().fetchall()

    result = [
        HourlyCampaignRow(
            Total=r["Total"] or 0,
            Answered=r["Answered"] or 0,
            Abandon=r["Abandon"] or 0,
            gdate=None,
            ghour=r["ghour"],
            campaign=r["campaign"],
            SLA=r["SLA"]
        )
        for r in rows
    ]

    # Calculate AL (total) per campaign per date
    al_summary = defaultdict(lambda: {"Total": 0, "Answered": 0, "Abandon": 0})

    for r in result:
        key = r.campaign # group by campaign and date
        al_summary[key]["Total"] += r.Total
        al_summary[key]["Answered"] += r.Answered
        al_summary[key]["Abandon"] += r.Abandon

    # Convert summary dict to list of AL rows
    al_rows = []
    for k, v in al_summary.items():
        AL = (v["Answered"] / v["Total"] * 100) if v["Total"] > 0 else 0
        al_rows.append(
            HourlyCampaignALRow(
                Total=v["Total"],
                Answered=v["Answered"],
                Abandon=v["Abandon"],
                AL=round(AL, 0),  # round to 2 decimals
                gdate=None,
                campaign=k,
            )
        )


    return HourlyCampaignResp(al_rows=al_rows, rows=result)







# Return Customer daily data Date wise
@router.post("/hourly_date_wise_report_old", response_model=HourlyCampaignResp)
def get_hourly_date_wise_report_old(
    req: DashboardReq,
    db: Session = Depends(get_db3),
    db_main: Session = Depends(get_db4),
) -> Any:

    params = {
        "from_date": req.from_date,
        "to_date": req.to_date,
    }

    # -------------------- Shared / Dedicated --------------------
    shared_clause = ""
    params_shared = {}

    if str(req.sd_type) in ("0", "1"):
        shared_clause = "AND is_shared = :is_shared"
        params_shared["is_shared"] = req.sd_type

    # -------------------- Category --------------------
    category_clause = ""
    if str(req.category).upper() != "ALL":
        category_clause = "AND client_category = :category"
        params["category"] = req.category

    # -------------------- Fetch Campaigns --------------------
    if str(req.company_id).upper() == "ALL":
        rows = db_main.execute(
            text(f"""
                SELECT campaignid
                FROM registration_master
                WHERE status='A'
                  AND is_dd_client='1'
                  {shared_clause}
                  {category_clause}
            """),
            {**params, **params_shared}
        ).fetchall()
    else:
        rows = db_main.execute(
            text(f"""
                SELECT campaignid
                FROM registration_master
                WHERE company_id = :cid
                  AND status='A'
                  AND is_dd_client='1'
                  {shared_clause}
                  {category_clause}
            """),
            {**params, **params_shared, "cid": req.company_id}
        ).fetchall()

    if not rows:
        raise HTTPException(404, "No campaigns found")

    campaign_list = []
    for r in rows:
        if r[0]:
            campaign_list.extend([c.strip().strip("'") for c in r[0].split(",")])

    campaign_list = list(set(campaign_list))

    if not campaign_list:
        raise HTTPException(404, "No valid campaign IDs")

    params["cids"] = tuple(campaign_list)

    # -------------------- DATE-WISE SQL --------------------
    sql = """
        SELECT
            COUNT(*)                       AS Total,
            SUM(IF(t2.user <> 'VDCL',1,0)) AS Answered,
            SUM(IF(t2.user = 'VDCL',1,0))  AS Abandon,
            DATE(t2.call_date)             AS gdate,
            HOUR(t2.call_date)             AS ghour
        FROM asterisk.vicidial_closer_log t2
        LEFT JOIN asterisk.vicidial_agent_log t3
            ON t2.uniqueid = t3.uniqueid
            AND t2.user     = t3.user
            AND t2.lead_id = t3.lead_id
        WHERE DATE(t2.call_date) BETWEEN :from_date AND :to_date
          AND t2.campaign_id IN :cids
          AND t2.term_reason <> 'AFTERHOURS'
          AND t2.lead_id IS NOT NULL
        GROUP BY
            DATE(t2.call_date),
            HOUR(t2.call_date)
        ORDER BY
            gdate,
            ghour
    """

    rows = db.execute(text(sql), params).mappings().fetchall()

    # -------------------- Hourly Rows --------------------
    result = [
        HourlyCampaignRow(
            Total=r["Total"] or 0,
            Answered=r["Answered"] or 0,
            Abandon=r["Abandon"] or 0,
            gdate=r["gdate"].isoformat(),
            ghour=r["ghour"],
            campaign=None,   # NOT campaign-wise
        )
        for r in rows
    ]

    # -------------------- DATE-WISE AL --------------------
    al_summary = defaultdict(lambda: {"Total": 0, "Answered": 0, "Abandon": 0})

    for r in result:
        al_summary[r.gdate]["Total"] += r.Total
        al_summary[r.gdate]["Answered"] += r.Answered
        al_summary[r.gdate]["Abandon"] += r.Abandon

    al_rows = []
    for gdate, v in al_summary.items():
        AL = (v["Answered"] / v["Total"] * 100) if v["Total"] > 0 else 0
        al_rows.append(
            HourlyCampaignALRow(
                Total=v["Total"],
                Answered=v["Answered"],
                Abandon=v["Abandon"],
                AL=round(AL, 0),
                gdate=gdate,
                campaign=None,   # DATE-WISE
            )
        )

    return HourlyCampaignResp(
        rows=result,
        al_rows=al_rows
    )











def auto_fit_columns(ws):
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[column_letter].width = max_length + 2


@router.get("/did-logs/export-excel_old")
def export_did_logs_excel_old(db: Session = Depends(get_db4)):

    query = text("""
        SELECT id, client_name, did_number, vendor_name, unique_id, call_time
        FROM did_logs
        ORDER BY id ASC
    """)

    rows = db.execute(query).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "DID Logs Report"

    # ==== Report Title ====
    ws["A1"] = "Report: DID Logs"
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append([])

    # ==== Table Headers ====
    headers = [
        "ID",
        "Client Name",
        "DID Number",
        "Vendor Name",
        "Unique ID",
        "Call Time",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    # ==== Insert Data Rows ====
    for row in rows:
        ws.append([
            row.id,
            row.client_name,
            row.did_number,
            row.vendor_name,
            row.unique_id,
            row.call_time.strftime("%Y-%m-%d %H:%M:%S") if row.call_time else "",
        ])

    # Auto-size columns
    auto_fit_columns(ws)

    # Save file to memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="DID_Logs_Report.xlsx"'
        }
    )





















def auto_fit_columns(ws):
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells) if column_cells else 10
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2


# ---------------- LIST API ----------------
@router.get("/channel-utilizations/list_old")
def get_channel_list_old(fromDate: str, toDate: str, db: Session = Depends(get_db3)):
    query = text("""
        SELECT id, update_date, channel_count, vendor
        FROM channel_info
        WHERE update_date BETWEEN :from AND :to
        ORDER BY update_date ASC
    """)
    rows = db.execute(query, {
        "from": fromDate + " 00:00:00",
        "to": toDate + " 23:59:59"
    }).mappings().all()

    return {"status": True, "data": rows}


# ---------------- MAX COUNT API ----------------
@router.get("/channel-utilizations/max-count_old")
def get_max_count_old(fromDate: str, toDate: str, db: Session = Depends(get_db3)):
    query = text("""
        SELECT ci.vendor, ci.channel_count AS max_count, ci.update_date
        FROM channel_info ci
        INNER JOIN (
            SELECT vendor, MAX(channel_count) AS max_count
            FROM channel_info
            WHERE update_date BETWEEN :from AND :to
            GROUP BY vendor
        ) mx ON ci.vendor = mx.vendor AND ci.channel_count = mx.max_count
        WHERE ci.update_date = (
            SELECT MAX(update_date)
            FROM channel_info ci2
            WHERE ci2.vendor = ci.vendor
              AND ci2.channel_count = ci.channel_count
              AND ci2.update_date BETWEEN :from AND :to
        )
        ORDER BY ci.vendor
    """)
    rows = db.execute(query, {
        "from": fromDate + " 00:00:00",
        "to": toDate + " 23:59:59"
    }).mappings().all()

    return {"status": True, "data": rows}


# ---------------- EXCEL DOWNLOAD ----------------
@router.get("/channel-utilizations/download_old")
def download_excel_old(fromDate: str, toDate: str, db: Session = Depends(get_db3)):

    # Sheet 1: All data
    sheet1_query = text("""
        SELECT id, update_date, channel_count, vendor
        FROM channel_info
        WHERE update_date BETWEEN :from AND :to
        ORDER BY update_date ASC
    """)
    sheet1 = db.execute(sheet1_query, {
        "from": fromDate + " 00:00:00",
        "to": toDate + " 23:59:59"
    }).mappings().all()

    # Sheet 2: Max count per vendor
    sheet2_query = text("""
        SELECT ci.vendor, ci.channel_count AS max_count, ci.update_date
        FROM channel_info ci
        INNER JOIN (
            SELECT vendor, MAX(channel_count) AS max_count
            FROM channel_info
            WHERE update_date BETWEEN :from AND :to
            GROUP BY vendor
        ) mx ON ci.vendor = mx.vendor AND ci.channel_count = mx.max_count
        WHERE ci.update_date = (
            SELECT MAX(update_date)
            FROM channel_info ci2
            WHERE ci2.vendor = ci.vendor
              AND ci2.channel_count = ci.channel_count
              AND ci2.update_date BETWEEN :from AND :to
        )
        ORDER BY ci.vendor
    """)
    sheet2 = db.execute(sheet2_query, {
        "from": fromDate + " 00:00:00",
        "to": toDate + " 23:59:59"
    }).mappings().all()

    # ---------------- Create Excel ----------------
    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    alignment_center = Alignment(horizontal="center")

    # Sheet 1
    ws1 = wb.active
    ws1.title = "Channel Utilizations"
    ws1.append([f"Report: Channel Utilizations"])
    ws1.append([f"Date Range: {fromDate} to {toDate}"])
    ws1.append([])

    headers1 = ["ID", "Date", "Channel Count", "Vendor"]
    ws1.append(headers1)
    for col, header in enumerate(headers1, start=1):
        cell = ws1.cell(row=4, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment_center

    for row in sheet1:
        ws1.append([
            row["id"],
            row["update_date"].strftime("%Y-%m-%d %H:%M:%S"),
            row["channel_count"],
            row["vendor"]
        ])
    auto_fit_columns(ws1)

    # Sheet 2
    ws2 = wb.create_sheet("Vendor Max Count")
    ws2.append([f"Report: Vendor Max Count"])
    ws2.append([f"Date Range: {fromDate} to {toDate}"])
    ws2.append([])

    headers2 = ["Vendor", "Max Count Channel User", "Date"]
    ws2.append(headers2)
    for col, header in enumerate(headers2, start=1):
        cell = ws2.cell(row=4, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment_center

    for row in sheet2:
        ws2.append([
            row["vendor"],
            row["max_count"],
            row["update_date"].strftime("%Y-%m-%d %H:%M:%S")
        ])
    auto_fit_columns(ws2)

    # Return Excel
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return Response(
        content=stream.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
            f'attachment; filename="Channel_Utilization_{fromDate}_to_{toDate}.xlsx"'
        }
    )






@router.get("/call/download_excel_raw_old")
def download_excel_raw_old(
        client_id: int,
        from_date: date = Query(...),
        to_date: date = Query(...),
        db=Depends(get_db4),
        db2=Depends(get_db3),
        db3=Depends(get_db4),
):
    # Step 1: Client Info
    client_result = db.execute(text("""
        SELECT company_name, reg_office_address1, phone_no, email, auth_person
        FROM registration_master
        WHERE company_id = :client_id
    """), {"client_id": client_id}).fetchone()

    campaign_q = text("SELECT campaignid FROM registration_master WHERE company_id = :company_id LIMIT 1")
    campaign_row = db.execute(campaign_q, {"company_id": client_id}).mappings().fetchone()
    if not campaign_row or not campaign_row.get("campaignid"):
        raise HTTPException(status_code=404, detail="Company / campaign not found")

    # --- CLEAN CAMPAIGN LIST ---
    raw_campaign = campaign_row["campaignid"]
    campaign_list = [c.strip().strip("'") for c in raw_campaign.split(",") if c.strip()]

    balance_result = db.execute(text("""
        SELECT * FROM balance_master
        WHERE clientId = :client_id
        LIMIT 1
    """), {"client_id": client_id}).mappings().fetchone()

    plan_result = None
    if balance_result and balance_result.PlanId:
        plan_result = db.execute(text("""
            SELECT * FROM plan_master
            WHERE Id = :plan_id
            LIMIT 1
        """), {"plan_id": balance_result.PlanId}).mappings().fetchone()
        

    # parse plan values (with sensible defaults)
    try:
        ib_pulse_sec = float(plan_result.get("pulse_day_shift") or 60)
    except:
        ib_pulse_sec = 60.0
    try:
        ibn_pulse_sec = float(plan_result.get("pulse_night_shift") or ib_pulse_sec)
    except:
        ibn_pulse_sec = ib_pulse_sec

    try:
        ob_pulse_sec = float(plan_result.get("pulse_outbound_call_shift") or 60)
    except:
        ob_pulse_sec = 60.0

    try:
        ib_pulse_rate = Decimal(plan_result.get("rate_per_pulse_day_shift") or 0)
    except:
        ib_pulse_rate = Decimal(0)
    try:
        ibn_pulse_rate = Decimal(plan_result.get("rate_per_pulse_night_shift") or 0)
    except:
        ibn_pulse_rate = Decimal(0)
    try:
        ob_pulse_rate = Decimal(plan_result.get("rate_per_pulse_outbound_call_shift") or 0)
    except:
        ob_pulse_rate = Decimal(0)

    # plan-level flat/charge fields (kept same names as PHP)
    ib_charge_plan = Decimal(plan_result.get("InboundCallCharge") or 0)
    ibn_charge_plan = Decimal(plan_result.get("InboundCallChargeNight") or 0)
    ob_charge_plan = Decimal(plan_result.get("OutboundCallCharge") or 0)

    # first minute logic (match PHP 'Enable' exactly)
    first_minute_enabled = (str(plan_result.get("first_minute", "")).lower() == "enable")
    ifmp = ceil(60.0 / ib_pulse_sec) if ib_pulse_sec > 0 else 1
    ifmp_n = ceil(60.0 / ibn_pulse_sec) if ibn_pulse_sec > 0 else 1
    ofmp = ceil(60.0 / ob_pulse_sec) if ob_pulse_sec > 0 else 1
    ob_first_min = first_minute_enabled  # same as PHP's $ob_first_min

    # used_amount = sum([
    #     balance_result[key] if key in balance_result and balance_result[key] else 0
    #     for key in [
    #         "ib_total", "ibn_total", "ob_total", "TMinAmount",
    #         "TMouAmount", "TvfAmount", "sms_total", "email_total",
    #         "TivAmount", "TWhatsAppAmount", "TBoatAmount"
    #     ]
    # ]) if balance_result else 0

    Used_Amount = Decimal(0)

    

    # Step 2: Call log data from vicidial DB
    call_data = db2.execute(text(f"""
        SELECT
            IF(t3.talk_sec IS NULL, t2.length_in_sec, t3.talk_sec) AS length_in_sec,
            t2.phone_number,
            t2.call_date,
            t2.user
        FROM vicidial_closer_log t2
        LEFT JOIN vicidial_agent_log t3 ON t2.uniqueid = t3.uniqueid AND t2.user = t3.user
        WHERE t2.user != 'VDCL'
          AND t2.campaign_id IN :campaigns
          AND DATE(t2.call_date) BETWEEN :from_date AND :to_date
    """), {"campaigns": tuple(campaign_list),"from_date": from_date, "to_date": to_date}).mappings().fetchall()

    html_day_rows = ""
    html_night_rows = ""

    ib_pulse = 0
    ib_secs = 0
    ib_total = Decimal(0)

    ibn_pulse = 0
    ibn_secs = 0
    ibn_total = Decimal(0)

    ab_pulse = 0
    ab_secs = 0
    ab_total = Decimal(0)

    ob_pulse = 0
    ob_secs = 0
    ob_total = Decimal(0)

    

    # multilang_call_data = db2.execute(text(f"""
    #     SELECT 
    #         IF(t3.talk_sec IS NULL, t2.length_in_sec, t3.talk_sec) AS length_in_sec,
    #         t2.phone_number,
    #         t2.call_date,
    #         t2.user
    #     FROM vicidial_closer_log t2
    #     LEFT JOIN vicidial_agent_log t3 ON t2.uniqueid = t3.uniqueid AND t2.user = t3.user
    #     WHERE t2.user != 'VDCL'
    #       AND t2.campaign_id IN ('ML01', 'ML02', 'ML03')  -- replace with your actual multi-language campaign_ids
    #       AND DATE(t2.call_date) BETWEEN :from_date AND :to_date
    # """), {"from_date": from_date, "to_date": to_date}).fetchall()

    # --- OUTBOUND (Vicidial Log) Section ---
    aband_data = db2.execute(text("""
            SELECT
                (va.talk_sec-va.dead_sec) length_in_sec,
                LEFT(v.phone_number, 10) AS phone_number,
                v.call_date,
                v.user
            FROM vicidial_log v
            JOIN vicidial_agent_log va ON v.uniqueid = va.uniqueid
            WHERE length_in_sec != 0
                AND v.user != 'VDAD'
                AND v.campaign_id IN :campaigns
              AND DATE(v.call_date) BETWEEN :from_date AND :to_date
        """), {"campaigns": tuple(campaign_list),"client_id": client_id, "from_date": from_date, "to_date": to_date}).mappings().fetchall()

    sql_get_ob = text("""
        SELECT LEFT(PhoneNo,10) AS PhoneNumber, DATE(Callbackdate) AS CallbackDate
        FROM aband_call_master
        WHERE clientid = :client_id
        AND DATE(Callbackdate) BETWEEN :from_date AND :to_date
        GROUP BY 1,2
    """)
    ob_rows = db.execute(sql_get_ob, {"client_id": client_id, "from_date": from_date, "to_date": to_date}).fetchall()

    # Prepare the direct IN string just like working API
    phone_date_pairs = [(r.PhoneNumber, r.CallbackDate) for r in ob_rows]
    if phone_date_pairs:
        in_values = ", ".join(f"('{pn}','{dt}')" for pn, dt in phone_date_pairs)
    else:
        in_values = "('','0000-00-00')"  # guaranteed no match
    
    query = text(f"""
                SELECT t2.list_id,
                t2.call_date AS CallDate,
                TIME(FROM_UNIXTIME(t2.start_epoch)) AS StartTime,
                LEFT(t2.phone_number,10) AS PhoneNumber,
                t2.`user` AS Agent,
                t3.talk_sec AS TalkSec
            FROM asterisk.vicidial_log t2
            LEFT JOIN vicidial_agent_log t3 ON t2.uniqueid = t3.uniqueid
            WHERE DATE(t2.call_date) BETWEEN :from_date AND :to_date
            AND t2.campaign_id = 'dialdesk'
            AND t2.lead_id IS NOT NULL
            AND (LEFT(t2.phone_number,10), DATE(t2.call_date)) IN ({in_values})
            """)

    ab_data = db2.execute(query, {"from_date": from_date, "to_date": to_date}).mappings().fetchall()

    # --- Initialize SMS variables
    sms_pulse = 0
    sms_secs = 0
    sms_charge = Decimal(0)
    sms_flat = 0
    sms_total = Decimal(0)

    # Get plan SMS charge
    sms_charge = Decimal(plan_result.get("SMSCharge", 0) or 0)

    sms_query = text("""
        SELECT 
            DATE_FORMAT(CallDate,'%d %b %y') AS CallDate1,
            CallDate,
            CallTime,
            CallFrom,
            Unit,
            AlertTo
        FROM billing_master
        WHERE clientId = :client_id
          AND DedType = 'SMS'
          AND DATE(CallDate) BETWEEN :from_date AND :to_date
    """)

    sms_data = db.execute(sms_query, {
        "client_id": client_id,
        "from_date": from_date,
        "to_date": to_date
    }).mappings().fetchall()


    # --- Initialize EMAIL variables (same as PHP) ---
    email_pulse = 0
    email_secs = 0
    email_charge = Decimal(0)
    email_flat = 0
    email_total = Decimal(0)

    # Plan charge for Email
    email_charge = Decimal(plan_result.get("EmailCharge", 0) or 0)
    email_flat = 0

    email_query = text("""
        SELECT 
            DATE_FORMAT(CallDate,'%d %b %y') AS CallDate1,
            CallDate,
            CallTime,
            CallFrom,
            Unit
        FROM billing_master
        WHERE clientId = :client_id
          AND DedType = 'Email'
          AND DATE(CallDate) BETWEEN :from_date AND :to_date
    """)

    email_data = db.execute(email_query, {
        "client_id": client_id,
        "from_date": from_date,
        "to_date": to_date
    }).mappings().fetchall()


    ivr_charge = Decimal(plan_result.get("IVR_Charge"))

    rx_query = text("""
        SELECT 
            DATE_FORMAT(call_time,'%d %b %y') AS CallDate1,
            call_time AS CallDate,
            TIME(call_time) AS CallTime,
            1 AS Unit,
            source_number AS CallFrom
        FROM rx_log
        WHERE clientId = :client_id
          AND DATE(call_time) BETWEEN :from_date AND :to_date
    """)

    rx_data = db3.execute(rx_query, {
        "client_id": client_id,
        "from_date": from_date,
        "to_date": to_date
    }).mappings().fetchall()

    total_talk_time = 0
    total_pulse = 0
    total_rate = 0.0

    total_talk_time2 = 0
    total_pulse2 = 0
    total_rate2 = 0.0

    total_talk_time3 = 0
    total_pulse3 = 0
    total_rate3 = 0.0

    total_talk_time4 = 0
    total_pulse4 = 0
    total_rate4 = 0.0

    total_pulse5 = 0
    total_rate5 = 0.0

    total_pulse6 = 0
    total_rate6 = 0.0

    total_pulse7 = 0
    total_rate7 = Decimal(0)

    # Step 3: Build HTML for Excel
    html = f"""
    <html><head><meta http-equiv="Content-Type" content="application/vnd.ms-excel; charset=utf-8" /></head><body>
    <table border='0' width='600' cellpadding='2' cellspacing='2'>
        <tr><td colspan='6' align='center'>
            <img src='http://dialdesk.co.in/dialdesk/app/webroot/billing_statement/logo.jpg' height='80'><br>
            <strong style='font-size:16pt;'>A UNIT OF ISPARK DATA CONNECT PVT LTD</strong>
        </td></tr>
    </table>

    <table border='1' width='600' cellpadding='2' cellspacing='2' style="font-size:11pt;">
        <tr><td colspan='7' style='font-size:15pt;background-color:#607d8b;color:#fff;'>Client Details</td></tr>
        <tr><th>Company</th><th colspan='3'>Address</th><th>Mobile No</th><th>Email</th><th>Authorised Person</th></tr>
        <tr>
            <td>{client_result.company_name if client_result else ''}</td>
            <td colspan='3'>{client_result.reg_office_address1 if client_result else ''}</td>
            <td>{client_result.phone_no if client_result else ''}</td>
            <td>{client_result.email if client_result else ''}</td>
            <td>{client_result.auth_person if client_result else ''}</td>
        </tr>
    </table>

    <table><tr><td>&nbsp;</td></tr></table>
  

    <!-- SUMMARY_PLACEHOLDER -->
    {{SUMMARY_TABLE}}


    <table><tr><td>&nbsp;</td></tr></table>

    """

    # for row in call_data:
    #     dt = row.call_date
    #     talk_time = int(row.length_in_sec or 0)
    #     pulse = (talk_time // 60) + (1 if talk_time % 60 else 0)
    #     rate = pulse * 0.5  # Replace with actual rate

    #     total_talk_time += talk_time
    #     total_pulse += pulse
    #     total_rate += rate  # Example rate, replace with actual



    for r in call_data:
        length = r.get("length_in_sec")
        call_date = r.get("call_date")
        # skip empty durations
        if not length:
            continue
        try:
            duration = float(length)
        except:
            continue

        # normalize call_date to datetime
        if isinstance(call_date, str):
            try:
                call_dt = datetime.fromisoformat(call_date)
            except:
                try:
                    call_dt = datetime.strptime(call_date, "%Y-%m-%d %H:%M:%S")
                except:
                    call_dt = None
        else:
            call_dt = call_date

        call_time_str = call_dt.strftime("%H:%M:%S") if call_dt else "00:00:00"
        # match PHP night logic: night if >=20:00 or <=08:00 (note: PHP used <= '08:00:00')
        is_night = (call_time_str >= "20:00:00") or (call_time_str <= "08:00:00")

        # PHP keeps a call_pulsesec variable equal to 0 and then adds it to *_secs.
        call_pulsesec = 0

        if is_night:
            ibn_secs += duration
            convrt_pulse = duration / ibn_pulse_sec if ibn_pulse_sec > 0 else duration
            if first_minute_enabled:
                if convrt_pulse > ifmp_n:
                    subsequent = convrt_pulse - ifmp_n
                    call_pulse = int(ifmp_n + ceil(subsequent))
                else:
                    call_pulse = int(ifmp_n)
            else:
                call_pulse = int(ceil(duration / ibn_pulse_sec)) if ibn_pulse_sec > 0 else int(ceil(duration))
            call_rate = (Decimal(call_pulse) * ibn_pulse_rate).quantize(Decimal("0.0001"))
            ibn_pulse += call_pulse
            ibn_secs += call_pulsesec    # PHP increments by call_pulsesec which is 0
            ibn_total += call_rate

            # Add row to night table
            html_night_rows += f"<tr><td>{call_dt.date()}</td><td>{call_dt.time()}</td><td>{r.phone_number}</td><td>{r.user}</td><td>{length}</td><td>{call_pulse}</td><td>{call_rate:.2f}</td></tr>"

        else:
            ib_secs += duration
            convrt_pulse = duration / ib_pulse_sec if ib_pulse_sec > 0 else duration
            if first_minute_enabled:
                if convrt_pulse > ifmp:
                    subsequent = convrt_pulse - ifmp
                    call_pulse = int(ifmp + ceil(subsequent))
                else:
                    call_pulse = int(ifmp)
            else:
                call_pulse = int(ceil(duration / ib_pulse_sec)) if ib_pulse_sec > 0 else int(ceil(duration))
            call_rate = (Decimal(call_pulse) * ib_pulse_rate).quantize(Decimal("0.0001"))
            ib_pulse += call_pulse
            ib_secs += call_pulsesec    # PHP increments by call_pulsesec which is 0
            ib_total += call_rate

            # Add row to day table
            html_day_rows += f"<tr><td>{call_dt.date()}</td><td>{call_dt.time()}</td><td>{r.phone_number}</td><td>{r.user}</td><td>{length}</td><td>{call_pulse}</td><td>{call_rate:.2f}</td></tr>"

    # html += f"""
    #     <tr style='font-weight:bold; background-color:#e0e0e0;'>
    #         <td colspan='4' align='right'>Total</td>
    #         <td>{ib_secs}</td>
    #         <td>{ib_pulse}</td>
    #         <td>{ib_total:.2f}</td>
    #     </tr>
    # """


    # --- Create DAY table ---
    html += f"""
    <table border='1' width='600' cellpadding='2' cellspacing='2' style="font-size:11pt;">
    <tr><td colspan='7' style='font-size:15pt;background-color:#607d8b;color:#fff;'>{client_result.company_name if client_result else ''} (INBOUND DAY)</td></tr>
    <tr><th>Date</th><th>Time</th><th>Call From</th><th>Agent</th><th>Talk Time</th><th>Pulse</th><th>Rate</th></tr>
    {html_day_rows}
    <tr style='font-weight:bold; background-color:#e0e0e0;'>
        <td colspan='4' align='right'>Total</td>
        <td>{ib_secs}</td>
        <td>{ib_pulse}</td>
        <td>{ib_total:.2f}</td>
    </tr>
    </table>
    """

    # --- Create NIGHT table ---
    html += f"""
    <table><tr><td>&nbsp;</td></tr></table>
    <table border='1' width='600' cellpadding='2' cellspacing='2' style="font-size:11pt;">
    <tr><td colspan='7' style="font-size:15pt;background-color:#607d8b;color:#fff;">{client_result.company_name if client_result else ''} (INBOUND NIGHT)</td></tr>
    <tr><th>Date</th><th>Time</th><th>Call From</th><th>Agent</th><th>Talk Time</th><th>Pulse</th><th>Rate</th></tr>
    {html_night_rows}
    <tr style='font-weight:bold; background-color:#e0e0e0;'>
        <td colspan='4' align='right'>Total</td>
        <td>{ibn_secs}</td>
        <td>{ibn_pulse}</td>
        <td>{ibn_total:.2f}</td>
    </tr>
    </table>
    """



    # html += """
    #     <table><tr><td>&nbsp;</td></tr></table>

    #     <table border='1' width='600' cellpadding='2' cellspacing='2'>
    #         <tr><td colspan='7' style='font-size:15pt;background-color:#607d8b;color:#fff;'>""" + \
    #         f"""{client_result.company_name if client_result else ''} (Multi Language INBOUND)</td></tr>
    #         <tr><th>Date</th><th>Time</th><th>Call From</th><th>Agent</th><th>Talk Time</th><th>Pulse</th><th>Rate</th></tr>
    #     """

    # for row in multilang_call_data:
    #     dt = row.call_date
    #     talk_time = int(row.length_in_sec or 0)
    #     pulse = (talk_time // 60) + (1 if talk_time % 60 else 0)
    #     rate = pulse * 0.5  # Adjust if multi-lang rate is different

    #     total_talk_time2 += talk_time
    #     total_pulse2 += pulse
    #     total_rate2 += rate

    #     html += f"<tr><td>{dt.date()}</td><td>{dt.time()}</td><td>{row.phone_number}</td><td>{row.user}</td><td>{talk_time}</td><td>{pulse}</td><td>{rate:.2f}</td></tr>"

    # html += f"""
    #         <tr style='font-weight:bold; background-color:#e0e0e0;'>
    #             <td colspan='4' align='right'>Total</td>
    #             <td>{total_talk_time2}</td>
    #             <td>{total_pulse2}</td>
    #             <td>{total_rate2:.2f}</td>
    #         </tr>
    #     """

    html += f"""
        <table><tr><td>&nbsp;</td></tr></table>
        <table border="1" width="600" cellpadding="2" cellspacing="2" style="font-size:11pt;">
            <tr><td colspan="7" style="font-size:15pt;background-color:#607d8b;color:#fff;">{client_result.company_name if client_result else ''} (OUTBOUND)</td></tr>
            <tr>
                <th>Date</th>
                <th>Time</th>
                <th>Call From</th>
                <th>Agent</th>
                <th>Talk Time</th>
                <th>Pulse</th>
                <th>Rate</th>
            </tr>
        """

    
    for ob in aband_data:
        raw_len = ob.get("length_in_sec")
        dt = ob.get("call_date")
        if not raw_len:
            continue

        callLength = round(float(raw_len))  # PHP round()
        if callLength <= 0:
            continue
        ab_secs += callLength

        # pulse conversion
        convrt_pulse = callLength / ob_pulse_sec if ob_pulse_sec > 0 else callLength

        # ---- First-minute logic (PHP exact replica) ----
        if ob_first_min:
            if convrt_pulse > ofmp:
                subsequent = math.ceil(convrt_pulse - ofmp)
                total_pulse = ofmp + subsequent
            elif callLength == 0:
                total_pulse = 0
            else:
                total_pulse = ofmp
        else:
            total_pulse = math.ceil(callLength / ob_pulse_sec)

        # bill amount
        amount = Decimal(total_pulse) * Decimal(ob_pulse_rate)

        # aggregate
        ab_pulse += total_pulse
        ab_total += amount
        ab_secs += 0


    # for row in outbound_data:
    #     dt = row.call_date
    #     talk_time = int(row.length_in_sec or 0)
    #     pulse = (talk_time // 60) + (1 if talk_time % 60 else 0)
    #     rate = pulse * 0.5  # You can replace with actual per-minute rate

    #     total_talk_time3 += talk_time
    #     total_pulse3 += pulse
    #     total_rate3 += rate

        html += f"<tr><td>{dt.date()}</td><td>{dt.time()}</td><td>{ob.phone_number}</td><td>{ob.user}</td><td>{callLength}</td><td>{total_pulse}</td><td>{amount:.2f}</td></tr>"

    html += f"""
            <tr style='font-weight:bold; background-color:#e0e0e0;'>
                <td colspan='4' align='right'>Total</td>
                <td>{ab_secs}</td>
                <td>{ab_pulse}</td>
                <td>{ab_total:.2f}</td>
            </tr>
        """

    html += f"""
        <table><tr><td>&nbsp;</td></tr></table>
        <table border="1" width="600" cellpadding="2" cellspacing="2" style="font-size:11pt;">
            <tr><td colspan="7" style="font-size:15pt;background-color:#607d8b;color:#fff;">{client_result.company_name if client_result else ''} (ABAND CALLBACK)</td></tr>
            <tr>
                <th>Date</th>
                <th>Time</th>
                <th>Call From</th>
                <th>Agent</th>
                <th>Talk Time</th>
                <th>Pulse</th>
                <th>Rate</th>
            </tr>
    """

    for call in ab_data:
        talk_sec = float(call.get("TalkSec") or 0)
        call_date = call.get("CallDate")
        if talk_sec <= 0:
            continue

        ob_secs += talk_sec

        # Calculate pulses using plan logic
        if ob_first_min:
            convrt_pulse = talk_sec / ob_pulse_sec
            if convrt_pulse > ofmp:
                subsequent = math.ceil(convrt_pulse - ofmp)
                total_pulse = ofmp + subsequent
            elif talk_sec == 0:
                total_pulse = 0
            else:
                total_pulse = ofmp
        else:
            total_pulse = math.ceil(talk_sec / ob_pulse_sec)

        ob_pulse += total_pulse
        ob_total += Decimal(total_pulse) * Decimal(ob_pulse_rate)



        html += f"""
            <tr>
                <td>{call_date.date()}</td>
                <td>{call_date.time()}</td>
                <td>{call.PhoneNumber}</td>
                <td>{call.Agent}</td>
                <td>{talk_sec}</td>
                <td>{total_pulse}</td>
                <td>{Decimal(total_pulse) * Decimal(ob_pulse_rate):.2f}</td>
            </tr>
        """

    html += f"""
            <tr style='font-weight:bold; background-color:#e0e0e0;'>
                <td colspan='4' align='right'>Total</td>
                <td>{ob_secs}</td>
                <td>{ob_pulse}</td>
                <td>{ob_total:.2f}</td>
            </tr>
        """

    html += f"""
        <table><tr><td>&nbsp;</td></tr></table>
        <table border="1" width="600" cellpadding="2" cellspacing="2" style="font-size:11pt;">
            <tr><td colspan="6" style="font-size:15pt;background-color:#607d8b;color:#fff;">{client_result.company_name if client_result else ''} (SMS)</td></tr>
            <tr>
                <th>Date</th>
                <th>Time</th>
                <th>Call From</th>
                <th>Alert To</th>
                <th>Pulse</th>
                <th>Rate</th>
            </tr>
    """


    for sms in sms_data:
        smsChar = int(sms.get("Duration") or 0)
        sms_unit = int(sms.get("Unit") or 0)

        sms_pulse += sms_unit
        sms_secs += smsChar
        sms_total += Decimal(sms_charge) * Decimal(sms_unit)


    # for row in sms_data:
    #     pulse = int(row.Unit) if row.Unit else 0
    #     rate = pulse * 0.2  # Set your actual rate here

    #     total_pulse5 += pulse
    #     total_rate5 += rate

        html += f"""
            <tr>
                <td>{sms.CallDate1}</td>
                <td>{sms.CallTime}</td>
                <td>{sms.CallFrom}</td>
                <td>{sms.AlertTo}</td>
                <td>{sms_unit}</td>
                <td>{sms_charge:.2f}</td>
            </tr>
        """

    html += f"""
            <tr style='font-weight:bold; background-color:#e0e0e0;'>
                <td colspan='4' align='right'>Total</td>
                <td>{sms_pulse}</td>
                <td>{sms_total:.2f}</td>
            </tr>
        """

    html += f"""
        <table><tr><td>&nbsp;</td></tr></table>
        <table border="1" width="600" cellpadding="2" cellspacing="2" style="font-size:11pt;">
            <tr><td colspan="5" style="font-size:15pt;background-color:#607d8b;color:#fff;">{client_result.company_name if client_result else ''} (EMAIL)</td></tr>
            <tr>
                <th>Date</th>
                <th>Time</th>
                <th>Call From</th>
                <th>Pulse</th>
                <th>Rate</th>
            </tr>
    """

    for email_row in email_data:
        EmailUnit = int(email_row.get("Unit") or 0)
        email_rate = Decimal(EmailUnit) * email_charge

        email_pulse += EmailUnit
        email_total += email_rate

    # for row in email_data:
    #     pulse = int(row.Unit) if row.Unit else 0
    #     rate = pulse * 0.25  # Replace with actual per-email rate

    #     total_pulse6 += pulse
    #     total_rate6 += rate

        html += f"""
            <tr>
                <td>{email_row.CallDate1}</td>
                <td>{email_row.CallTime}</td>
                <td>{email_row.CallFrom}</td>
                <td>{EmailUnit}</td>
                <td>{email_rate:.2f}</td>
            </tr>
        """

    html += f"""
            <tr style='font-weight:bold; background-color:#e0e0e0;'>
                <td colspan='3' align='right'>Total</td>
                <td>{email_pulse}</td>
                <td>{email_total:.2f}</td>
            </tr>
        """

    html += f"""
        <table><tr><td>&nbsp;</td></tr></table>
        <table border="1" width="600" cellpadding="2" cellspacing="2" style="font-size:11pt;">
            <tr><td colspan="5" style="font-size:15pt;background-color:#607d8b;color:#fff;">{client_result.company_name if client_result else ''} (IVR)</td></tr>
            <tr>
                <th>Date</th>
                <th>Time</th>
                <th>Call From</th>
                <th>Pulse</th>
                <th>Rate</th>
            </tr>
    """

    for row in rx_data:
        pulse = 1
        rate = pulse * ivr_charge

        total_pulse7 += pulse
        total_rate7 += rate

        html += f"""
            <tr>
                <td>{row.CallDate1}</td>
                <td>{row.CallTime}</td>
                <td>{row.CallFrom}</td>
                <td>{pulse}</td>
                <td>{rate:.2f}</td>
            </tr>
        """

    html += f"""
            <tr style='font-weight:bold; background-color:#e0e0e0;'>
                <td colspan='3' align='right'>Total</td>
                <td>{total_pulse7}</td>
                <td>{total_rate7:.2f}</td>
            </tr>
        """

    # === 1️⃣ Get dynamic rates with fallback ===
    rate_icb = plan_result.InboundRate if plan_result and hasattr(plan_result, "InboundRate") else 0.5
    rate_multilang = plan_result.MultiLangInboundRate if plan_result and hasattr(plan_result,
                                                                                 "MultiLangInboundRate") else 0.5
    rate_ocb = plan_result.OutboundRate if plan_result and hasattr(plan_result, "OutboundRate") else 0.5
    rate_abcb = plan_result.AbandCallRate if plan_result and hasattr(plan_result, "AbandCallRate") else 0.5
    rate_sms = plan_result.SMSRate if plan_result and hasattr(plan_result, "SMSRate") else 0.2
    rate_email = plan_result.EmailRate if plan_result and hasattr(plan_result, "EmailRate") else 0.25
    rate_rx = plan_result.IVR_Charge if plan_result and hasattr(plan_result, "IVR_Charge") else 0.2

    # === 2️⃣ Recalculate final amounts ===
    amount_icb = total_pulse * rate_icb
    amount_multilang = total_pulse2 * rate_multilang
    amount_ocb = total_pulse3 * rate_ocb
    amount_abcb = total_pulse4 * rate_abcb
    amount_sms = total_pulse5 * rate_sms
    amount_email = total_pulse6 * rate_email
    amount_rx = total_pulse7 * ivr_charge

    # grand_total = ib_total + ibn_total + ob_total + ab_total + amount_sms + amount_email + amount_rx
    grand_total = (
        Decimal(ib_total) +
        Decimal(ibn_total) +
        Decimal(ob_total) +
        Decimal(ab_total) +
        Decimal(sms_total) +
        Decimal(email_total) +
        Decimal(amount_rx)
    )

    # used_amount = (
    #     Decimal(ib_total) +
    #     Decimal(ibn_total) +
    #     Decimal(ob_total) +
    #     Decimal(ab_total) +
    #     Decimal(sms_total) +
    #     Decimal(email_total) +
    #     Decimal(amount_rx)
    # )
    # print("#######",used_amount)
    # html = html.replace("{Used_Amount}", f"{used_amount:.2f}")


    # === 3️⃣ Append Summary Table ===
    summary_html  = f"""
    <table><tr><td>&nbsp;</td></tr></table>
    <table border='1' width='600' cellpadding='2' cellspacing='2' style='font-size:11pt;'>
        <tr>
            <td colspan='4' style='font-size:15pt;background-color:#607d8b;color:#fff;'>Summary</td>
        </tr>
        <tr>
            <th>Description</th>
            <th>Pulse/Unit</th>
            <th>Rate</th>
            <th>Amount</th>
        </tr>
        <tr><td>ICB</td><td>{ib_pulse}</td><td>{ib_pulse_rate} Rs./{ib_pulse_sec} Sec</td><td>{ib_total:.2f}</td></tr>
        <tr><td>ICB Night</td><td>{ibn_pulse}</td><td>{ibn_pulse_rate} Rs./{ibn_pulse_sec} Sec</td><td>{ibn_total:.2f}</td></tr>
        <tr><td>ABCB</td><td>{ob_pulse}</td><td>{ob_pulse_rate}Rs./{ob_pulse_sec} Sec</td><td>{ob_total:.2f}</td></tr>
        <tr><td>OCB</td><td>{ab_pulse}</td><td>{ob_pulse_rate}Rs./{ob_pulse_sec} Sec</td><td>{ab_total:.2f}</td></tr>
        <tr><td>SMS</td><td>{sms_pulse}</td><td>{sms_charge} Rs./Min</td><td>{sms_total:.2f}</td></tr>
        <tr><td>Email</td><td>{email_pulse}</td><td>{email_charge} Rs./Min</td><td>{email_total:.2f}</td></tr>
        <tr><td>IVR</td><td>{total_pulse7}</td><td>{ivr_charge} Rs./Min</td><td>{amount_rx:.2f}</td></tr>
        <tr style='font-weight:bold; background-color:#e0e0e0;'>
            <td colspan='3' align='right'>Grand Total ({from_date}/{to_date})</td>
            <td>{grand_total:.2f}</td>
        </tr>
    </table>
    """

    html = html.replace("{SUMMARY_TABLE}", summary_html)



    html += "</table></body></html>"

    # Step 4: Return as Excel
    buffer = BytesIO(html.encode('utf-8'))
    filename = f"statement_{datetime.now().strftime('%d_%m_%y_%H_%M_%S')}.xls"
    headers = {
        "Content-Disposition": f"attachment; filename={filename}",
        "Pragma": "no-cache",
        "Expires": "0"
    }

    return StreamingResponse(buffer, media_type="application/vnd.ms-excel", headers=headers)


