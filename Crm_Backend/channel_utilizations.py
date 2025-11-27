from fastapi import APIRouter, Depends, Response
from sqlalchemy.orm import Session
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from database import get_db2

router = APIRouter(prefix="/channel-utilizations", tags=["Channel Utilizations"])


def auto_fit_columns(ws):
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells) if column_cells else 10
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2


# ---------------- LIST API ----------------
@router.get("/list")
def get_channel_list(fromDate: str, toDate: str, db: Session = Depends(get_db2)):
    query = text("""
        SELECT id, update_date, channel_count, vendor
        FROM channel_info
        WHERE update_date BETWEEN :from AND :to
        ORDER BY update_date ASC
    """)
    rows = db.execute(query, {
        "from": fromDate + " 00:00:00",
        "to": toDate + " 23:59:59"
    }).mappings().all()

    return {"status": True, "data": rows}


# ---------------- MAX COUNT API ----------------
@router.get("/max-count")
def get_max_count(fromDate: str, toDate: str, db: Session = Depends(get_db2)):
    query = text("""
        SELECT ci.vendor, ci.channel_count AS max_count, ci.update_date
        FROM channel_info ci
        INNER JOIN (
            SELECT vendor, MAX(channel_count) AS max_count
            FROM channel_info
            WHERE update_date BETWEEN :from AND :to
            GROUP BY vendor
        ) mx ON ci.vendor = mx.vendor AND ci.channel_count = mx.max_count
        WHERE ci.update_date = (
            SELECT MAX(update_date)
            FROM channel_info ci2
            WHERE ci2.vendor = ci.vendor
              AND ci2.channel_count = ci.channel_count
              AND ci2.update_date BETWEEN :from AND :to
        )
        ORDER BY ci.vendor
    """)
    rows = db.execute(query, {
        "from": fromDate + " 00:00:00",
        "to": toDate + " 23:59:59"
    }).mappings().all()

    return {"status": True, "data": rows}


# ---------------- EXCEL DOWNLOAD ----------------
@router.get("/download")
def download_excel(fromDate: str, toDate: str, db: Session = Depends(get_db2)):

    # Sheet 1: All data
    sheet1_query = text("""
        SELECT id, update_date, channel_count, vendor
        FROM channel_info
        WHERE update_date BETWEEN :from AND :to
        ORDER BY update_date ASC
    """)
    sheet1 = db.execute(sheet1_query, {
        "from": fromDate + " 00:00:00",
        "to": toDate + " 23:59:59"
    }).mappings().all()

    # Sheet 2: Max count per vendor
    sheet2_query = text("""
        SELECT ci.vendor, ci.channel_count AS max_count, ci.update_date
        FROM channel_info ci
        INNER JOIN (
            SELECT vendor, MAX(channel_count) AS max_count
            FROM channel_info
            WHERE update_date BETWEEN :from AND :to
            GROUP BY vendor
        ) mx ON ci.vendor = mx.vendor AND ci.channel_count = mx.max_count
        WHERE ci.update_date = (
            SELECT MAX(update_date)
            FROM channel_info ci2
            WHERE ci2.vendor = ci.vendor
              AND ci2.channel_count = ci.channel_count
              AND ci2.update_date BETWEEN :from AND :to
        )
        ORDER BY ci.vendor
    """)
    sheet2 = db.execute(sheet2_query, {
        "from": fromDate + " 00:00:00",
        "to": toDate + " 23:59:59"
    }).mappings().all()

    # ---------------- Create Excel ----------------
    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    alignment_center = Alignment(horizontal="center")

    # Sheet 1
    ws1 = wb.active
    ws1.title = "Channel Utilizations"
    ws1.append([f"Report: Channel Utilizations"])
    ws1.append([f"Date Range: {fromDate} to {toDate}"])
    ws1.append([])

    headers1 = ["ID", "Date", "Channel Count", "Vendor"]
    ws1.append(headers1)
    for col, header in enumerate(headers1, start=1):
        cell = ws1.cell(row=4, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment_center

    for row in sheet1:
        ws1.append([
            row["id"],
            row["update_date"].strftime("%Y-%m-%d %H:%M:%S"),
            row["channel_count"],
            row["vendor"]
        ])
    auto_fit_columns(ws1)

    # Sheet 2
    ws2 = wb.create_sheet("Vendor Max Count")
    ws2.append([f"Report: Vendor Max Count"])
    ws2.append([f"Date Range: {fromDate} to {toDate}"])
    ws2.append([])

    headers2 = ["Vendor", "Max Count Channel User", "Date"]
    ws2.append(headers2)
    for col, header in enumerate(headers2, start=1):
        cell = ws2.cell(row=4, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment_center

    for row in sheet2:
        ws2.append([
            row["vendor"],
            row["max_count"],
            row["update_date"].strftime("%Y-%m-%d %H:%M:%S")
        ])
    auto_fit_columns(ws2)

    # Return Excel
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return Response(
        content=stream.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
            f'attachment; filename="Channel_Utilization_{fromDate}_to_{toDate}.xlsx"'
        }
    )
