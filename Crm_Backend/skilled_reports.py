from pydantic import BaseModel
from typing import List, Optional
from fastapi import APIRouter, Depends
from sqlalchemy.orm import Session
from sqlalchemy import text
from typing import Dict, Any
from database import get_db2,get_db4
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import tempfile



router = APIRouter()


class ClientLiveAgentRequest(BaseModel):
    client_ids: Optional[List[int]] = None




@router.post("/client_live_agent")
def get_client_live_agent(
    request: ClientLiveAgentRequest,
    db: Session = Depends(get_db4),   # registration DB
    db2: Session = Depends(get_db2),   # vicidial DB
) -> Dict[str, Dict[str, int]]:
    """
    Returns live agent count per campaign grouped by client
    (Python version of client_live_agent PHP function)
    """

    # -----------------------------
    # Build client filter condition
    # -----------------------------
    client_filter = ""
    params: Dict[str, Any] = {}

    if request.client_ids:
        client_filter = "AND r.company_id IN :client_ids"
        params["client_ids"] = tuple(request.client_ids)

    # -----------------------------
    # Fetch client + campaign list
    # -----------------------------
    client_campaign_query = text(f"""
        SELECT
            r.Company_name AS company_name,
            i.campaign_name AS campaign_name
        FROM registration_master r
        JOIN ingroup_campaign_master i
            ON r.company_id = i.client_id
        WHERE r.status = 'A'
          AND r.is_dd_client = '1'
          AND i.camp_type IS NULL
          {client_filter}
        ORDER BY r.Company_name ASC
    """)

    rows = db.execute(client_campaign_query, params).mappings().all()

    # -----------------------------
    # Count live agents per campaign
    # -----------------------------
    result: Dict[str, Dict[str, int]] = {}

    for row in rows:
        company_name = row["company_name"]
        campaign_name = row["campaign_name"]

        agent_count_query = text("""
            SELECT COUNT(*) AS cnt
            FROM vicidial_live_agents
            WHERE campaign_id = 'Dialdesk'
            AND closer_campaigns LIKE :campaign
        """)

        count_row = db2.execute(
            agent_count_query,
            {"campaign": f"%{campaign_name}%"}
        ).mappings().first()

        if company_name not in result:
            result[company_name] = {}

        result[company_name][campaign_name] = count_row["cnt"]

    return result









@router.get("/skill_wise_excel/export")
def export_skill_wise_excel(
    db: Session = Depends(get_db4),
    db2: Session = Depends(get_db2),
):
    # ----------------------------------
    # Fetch active agents
    # ----------------------------------
    agent_query = text("""
        SELECT
            username,
            displayname,
            dateofjoining
        FROM agent_master
        WHERE status = 'A'
        AND processname IN ('Shared IB', 'IB Dedicated')
    """)

    agents = db.execute(agent_query).mappings().all()

    userid = []
    name = {}
    doj = {}

    for row in agents:
        userid.append(row["username"])
        name[row["username"]] = row["displayname"]
        doj[row["username"]] = row["dateofjoining"]

    # ----------------------------------
    # Fetch skill mapping from Vicidial
    # ----------------------------------
    skills_map = {}
    max_skill_count = 0

    if userid:
        vicidial_query = text("""
            SELECT user, closer_campaigns
            FROM vicidial_users
            WHERE user IN :users
        """)

        rows = db2.execute(
            vicidial_query,
            {"users": tuple(userid)}
        ).mappings().all()

        for r in rows:
            skill_list = []
            if r["closer_campaigns"]:
                # Remove "-" and empty strings
                skill_list = [s for s in r["closer_campaigns"].split() if s.strip() and s != "-"]

            skills_map[r["user"]] = skill_list
            max_skill_count = max(max_skill_count, len(skill_list))


    # ----------------------------------
    # Create Excel
    # ----------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Skill Wise Agents"

    # Base headers
    headers = [
        "User ID",
        "Name",
        "Date of Joining",
        "Skilled Count"
    ]

    # Dynamic skill headers
    for i in range(1, max_skill_count + 1):
        headers.append(f"Skill {i}")

    header_fill = PatternFill(start_color="317EAC", end_color="317EAC", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # ----------------------------------
    # Fill rows
    # ----------------------------------
    row_num = 2
    for user in userid:
        skill_list = skills_map.get(user, [])

        ws.cell(row=row_num, column=1, value=user)
        ws.cell(row=row_num, column=2, value=name.get(user, ""))
        ws.cell(row=row_num, column=3, value=str(doj.get(user, "")))
        ws.cell(row=row_num, column=4, value=len(skill_list))

        col_num = 5
        for skill in skill_list:
            ws.cell(row=row_num, column=col_num, value=skill)
            col_num += 1

        row_num += 1

    # ----------------------------------
    # Save & return file
    # ----------------------------------
    tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp_file.name)

    return FileResponse(
        path=tmp_file.name,
        filename="skill_wise_agents.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )






@router.get("/agent_wise_skill_excel/export")
def export_agent_wise_skill_excel(
    db: Session = Depends(get_db4),   # Main DB
    db2: Session = Depends(get_db2),  # Vicidial DB
):
    # ----------------------------------
    # 1️⃣ Fetch all active DD clients + campaigns
    # ----------------------------------
    client_query = text("""
        SELECT 
            r.Company_name AS company_name,
            i.campaign_name AS campaign_name
        FROM registration_master r
        JOIN ingroup_campaign_master i
            ON r.company_id = i.client_id
        WHERE r.status = 'A'
          AND r.is_dd_client = '1'
          AND i.camp_type IS NULL
        ORDER BY r.Company_name ASC
    """)
    clients = db.execute(client_query).mappings().all()

    # ----------------------------------
    # 2️⃣ Campaign → Vicidial users mapping
    # ----------------------------------
    campaign_users = {}  # { campaign_name: [user1, user2...] }

    for c in clients:
        campaign = c["campaign_name"]

        vicidial_query = text("""
            SELECT GROUP_CONCAT(user) AS users
            FROM vicidial_users
            WHERE closer_campaigns LIKE :campaign
        """)
        result = db2.execute(
            vicidial_query,
            {"campaign": f"%{campaign}%"}
        ).mappings().first()

        if result and result["users"]:
            campaign_users[campaign] = result["users"].split(",")
        else:
            campaign_users[campaign] = []

    # ----------------------------------
    # 3️⃣ Fetch agent display names
    # ----------------------------------
    all_users = list(
        set(user for users in campaign_users.values() for user in users)
    )

    agents_map = {}  # { username: displayname }

    if all_users:
        agents_query = text("""
            SELECT username, displayname
            FROM agent_master
            WHERE username IN :users
              AND status = 'A'
              AND processname = 'Shared IB'
        """)
        rows = db.execute(
            agents_query,
            {"users": tuple(all_users)}
        ).mappings().all()

        for r in rows:
            agents_map[r["username"]] = r["displayname"]

    # ----------------------------------
    # 4️⃣ Excel preparation (IMAGE FORMAT)
    # ----------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "agent_skilled_mapped"

    # Dynamic agent columns
    max_agents = max(
        (len(v) for v in campaign_users.values()),
        default=0
    )

    headers = (
        ["Sno", "Client Name", "Skilled", "Agent Count", "Agent Name"]
        + [f"Agent {i}" for i in range(1, max_agents + 1)]
    )

    header_fill = PatternFill(start_color="317EAC", end_color="317EAC", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ----------------------------------
    # 5️⃣ Write data rows
    # ----------------------------------
    row_num = 2
    sno = 1

    for c in clients:
        company = c["company_name"]
        campaign = c["campaign_name"]

        usernames = campaign_users.get(campaign, [])
        display_names = [
            agents_map[u]
            for u in usernames
            if u in agents_map
        ]

        ws.cell(row=row_num, column=1, value=sno)
        ws.cell(row=row_num, column=2, value=company)
        ws.cell(row=row_num, column=3, value=campaign)
        ws.cell(row=row_num, column=4, value=len(display_names))

        # Agent Name → FIRST agent only
        if display_names:
            ws.cell(row=row_num, column=5, value=display_names[0])

        # Agent 1…N → START FROM SECOND AGENT
        for idx, agent in enumerate(display_names[1:]):
            ws.cell(row=row_num, column=6 + idx, value=agent)


        row_num += 1
        sno += 1

    # ----------------------------------
    # 6️⃣ Save & return Excel
    # ----------------------------------
    tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp_file.name)

    return FileResponse(
        path=tmp_file.name,
        filename="Agent_Skilled_Mapped.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
