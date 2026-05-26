from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from sqlalchemy import text
from typing import Any, List, Optional
from pydantic import BaseModel
from database import get_db2, get_db4
from datetime import date
from fastapi.responses import StreamingResponse
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

router = APIRouter()

# ----------------- Request & Response Models -----------------
class SLAClientwiseReq(BaseModel):
    from_date: date
    to_date: date
    company_id: str  # single or 'ALL'
    sd_type: Optional[str] = None  # "0" = Dedicated, "1" = Shared, None/All
    filter_type: Optional[str] = None  # "All", "SLA", "AL"


class SLAClientwiseRow(BaseModel):
    campaign_id: str
    Total: int
    TalkTime: str
    dispo_time: str
    WrapTime: str
    Answered: int
    Abandon: int
    TotalAcht: int
    WIthinSLA: int
    WIthinSLATen: int
    AbndWithinThresold: int
    AbndAfterThresold: int

class SLAClientwiseResp(BaseModel):
    rows: List[SLAClientwiseRow]



class Companies(BaseModel):
    company_id: int
    company_name: str






# def timedelta_to_seconds(td):
#     if td is None:
#         return 0
#     return int(td.total_seconds())



def timedelta_to_seconds(td):
    if isinstance(td, int):  # ✅ already seconds
        return td

    if isinstance(td, str):  # ✅ "HH:MM:SS"
        h, m, s = map(int, td.split(":"))
        return h * 3600 + m * 60 + s

    # ✅ timedelta case
    return int(td.total_seconds())


def seconds_to_hms(seconds):
    seconds = int(seconds)  # convert float to int
    h = seconds // 3600
    m = (seconds % 3600) // 60
    s = seconds % 60
    return f"{h:02d}:{m:02d}:{s:02d}"


def sec_to_time(seconds):
    h = int(seconds // 3600)
    m = int((seconds % 3600) // 60)
    s = int(seconds % 60)
    return f"{h:02}:{m:02}:{s:02}"





# fetches Company Name using filter is_shared
@router.get("/companies", response_model=List[Companies])
def get_companies(
    is_shared: Optional[str] = Query(
        default=None,
        description="Allowed values: 0, 1 or empty"
    ),
    db: Session = Depends(get_db4)
):
    base_query = """
        SELECT company_name, company_id
        FROM registration_master
        WHERE status = 'A'
          AND is_dd_client = '1'
    """

    params = {}

    if is_shared in ("0", "1"):
        base_query += " AND is_shared = :is_shared"
        params["is_shared"] = is_shared

    base_query += " ORDER BY company_name ASC"

    result = db.execute(text(base_query), params).fetchall()

    return [{"company_id": row.company_id, "company_name": row.company_name} for row in result]







@router.post("/sla_clientwise_report_excel")
def sla_clientwise_report_excel(req: SLAClientwiseReq, db2: Session = Depends(get_db2), db: Session = Depends(get_db4)):

    params_shared = {}
    shared_clause = ""
    if str(req.sd_type) in ("0", "1"):
        shared_clause = "AND is_shared = :is_shared"
        params_shared["is_shared"] = req.sd_type

    # Fetch campaigns
    if req.company_id.upper() == "ALL":
        rows = db.execute(
            text(f"""
                SELECT campaignid, company_id, company_name
                FROM registration_master
                WHERE status='A'
                  AND is_dd_client='1'
                  {shared_clause}
                ORDER BY company_id ASC
            """),
            params_shared
        ).fetchall()
    else:
        rows = db.execute(
            text(f"""
                SELECT campaignid, company_id, company_name
                FROM registration_master
                WHERE company_id = :cid
                  AND status='A'
                  AND is_dd_client='1'
                  {shared_clause}
            """),
            {**params_shared, "cid": req.company_id}
        ).fetchall()

    if not rows:
        raise HTTPException(404, "No campaigns found")

    # Flatten campaigns per company
    client_campaigns = {}
    for r in rows:
        company_id = r.company_id
        company_name = r.company_name
        if r.campaignid:
            campaigns = [c.strip().strip("'") for c in r.campaignid.split(",") if c.strip()]
            # client_campaigns[company_id] = campaigns
        else:
            campaigns = []
            # client_campaigns[company_id] = []

        client_campaigns[company_id] = {
            "company_name": company_name,
            "campaigns": campaigns
        }

    # Fetch plan rates per company
    company_rates = {}
    for company_id in client_campaigns.keys():
        balance = db.execute(
            text("SELECT PlanId FROM balance_master WHERE clientId=:cid LIMIT 1"),
            {"cid": company_id}
        ).fetchone()
        if balance:
            plan = db.execute(
                text("""
                    SELECT InboundCallCharge, rate_per_pulse_day_shift, pulse_day_shift
                    FROM plan_master
                    WHERE id = :plan_id LIMIT 1
                """), {"plan_id": balance.PlanId}
            ).fetchone()
            if plan:
                pulse_day_shift = int(plan.pulse_day_shift or 0)
                rate_per_pulse_day_shift = float(plan.rate_per_pulse_day_shift or 0)
                inbound_charge = float(plan.InboundCallCharge or 0)
                rate_per_sec = rate_per_pulse_day_shift / pulse_day_shift if pulse_day_shift > 0 else 0
                company_rates[company_id] = {
                    "InboundCallCharge": inbound_charge,
                    "rate_per_sec": rate_per_sec
                }
            else:
                company_rates[company_id] = {"InboundCallCharge": 0, "rate_per_sec": 0}
        else:
            company_rates[company_id] = {"InboundCallCharge": 0, "rate_per_sec": 0}


    # Prepare data
    all_data = {}
    total_handle = 0
    total_acht = 0
    total_answered = 0
    total_offered = 0
    # rl_list = set()


    grand_totals = {
        "Offered": 0,
        "Handled": 0,
        "Total Talk Time": "00:00:00",
        "Calls Ans (20 Sec)": 0,
        "Calls Ans (10 Sec)": 0,
        "Total Calls Abandoned": 0,
        "Abnd Within (20)": 0,
        "AHT_total_sec": 0,
        "Amount": 0,
        "RL": 0,
        "Call Rate": 0,
        "Total Number Of Tagging": 0
    }   
        

    for company_id, info in client_campaigns.items():
        # ---------------- RL & RL% calculation ----------------

        rl_sql = """
            SELECT COUNT(1) AS cnt
            FROM aband_call_master
            WHERE ClientId = :cid
            AND call_status = 'answer'
            AND DATE(calldate) BETWEEN :from_date AND :to_date
        """

        rl_row = db.execute(
            text(rl_sql),
            {
                "cid": company_id,
                "from_date": req.from_date,
                "to_date": req.to_date
            }
        ).fetchone()

        company_rl = int(rl_row.cnt or 0)


        # ---------------- TAGGING ----------------
        total_tagged = db.execute(
            text("""
                SELECT COUNT(Id)
                FROM call_master
                WHERE ClientId = :cid
                AND DATE(calldate) BETWEEN :from_date AND :to_date
                AND CallType <> 'Upload'
            """),
            {
                "cid": company_id,
                "from_date": req.from_date,
                "to_date": req.to_date
            }
        ).scalar() or 0

        company_name = info["company_name"]
        campaigns = info["campaigns"]

        if company_name not in all_data:
            all_data[company_name] = {
                "Client Name": company_name,
                "Offered": 0,
                "Handled": 0,
                "Calls Ans (20 Sec)": 0,
                "Total Calls Abandoned": 0,
                "Abnd Within (20)": 0,
                "Total Talk Time_sec": 0,
                "AHT_total": 0,
                "RL": company_rl,
                "Total Number Of Tagging": 0
            }

        rate_info = company_rates[company_id]
        for campaign in campaigns:
            sql = f"""
                SELECT 
                    t2.campaign_id,
                    t2.user,
                    t2.queue_seconds,
                    dispo_sec,
                    talk_sec,
                    pause_sec,
                    wait_sec,
                    t1.length_in_sec
                FROM asterisk.vicidial_closer_log t2
                LEFT JOIN asterisk.vicidial_users vu ON t2.user = vu.user
                LEFT JOIN asterisk.call_log t1 ON t1.uniqueid = t2.uniqueid
                LEFT JOIN asterisk.vicidial_agent_log t3 ON t2.uniqueid = t3.uniqueid AND t2.user = t3.user
                WHERE DATE(t2.call_date) BETWEEN :from_date AND :to_date
                  AND t2.term_reason <> 'AFTERHOURS'
                  AND t2.lead_id IS NOT NULL
                  AND t2.campaign_id = :campaign
            """
            row = db2.execute(text(sql), {"from_date": req.from_date, "to_date": req.to_date, "campaign": campaign}).mappings().all()

            # if not row:
            #     row = {k: 0 for k in ["Total","Answered","Abandon","TotalAcht","WIthinSLA","WIthinSLATen","AbndWithinThresold"]}
            #     row["TalkTime"] = "00:00:00"

            if not row:
                row = []   # ✅ must be list


            total = 0
            answered = 0
            abandon = 0
            total_acht_row = 0

            within_sla = 0
            within_sla_ten = 0
            abnd_within = 0
            abnd_after = 0

            talk_time_sec = 0
            dispo_time_sec = 0
            wrap_time_sec = 0

            for r in row:
                total += 1

                user = r["user"]
                queue_sec = r["queue_seconds"] or 0
                dispo_sec = r["dispo_sec"] or 0
                talk_sec = r["talk_sec"] or 0
                pause_sec = r["pause_sec"] or 0
                wait_sec = r["wait_sec"] or 0
                length_sec = r["length_in_sec"] or 0

                # Time calculations
                talk_time_sec += (talk_sec + pause_sec + wait_sec + dispo_sec)
                dispo_time_sec += dispo_sec
                wrap_time_sec += dispo_sec

                if user != "VDCL":
                    answered += 1
                    total_acht_row += length_sec

                    if queue_sec <= 20:
                        within_sla += 1
                    if queue_sec <= 10:
                        within_sla_ten += 1
                else:
                    abandon += 1

                    if queue_sec <= 20:
                        abnd_within += 1
                    else:
                        abnd_after += 1

            # ✅ KEEP SAME VARIABLE NAMES USED BELOW
            offered = total
            handled = answered
            wi_thin_sla = within_sla
            wi_thin_sla_ten = within_sla_ten

            # abandon = float(row["Abandon"] or 0)
            # total_acht_row = float(row["TotalAcht"] or 0)
            # offered = float(row["Total"] or 0)
            # handled = float(row["Answered"] or 0)
            # wi_thin_sla = float(row["WIthinSLA"] or 0)
            # wi_thin_sla_ten = float(row["WIthinSLATen"] or 0)
            # abnd_within = float(row["AbndWithinThresold"] or 0)

            

            # RL %
            denominator = handled + abandon
            # rl_percent = round(((handled + rl) / denominator) * 100) if denominator > 0 else 0

            # -------- FILTER LOGIC --------
            # skips Offered that have 0
            # if req.filter_type == "without_0" and offered <= 0:
            #     continue

            



            data = all_data[company_name]

            data["Offered"] += offered
            data["Handled"] += handled
            data["Calls Ans (20 Sec)"] += wi_thin_sla
            data["Total Calls Abandoned"] += abandon
            data["Abnd Within (20)"] += abnd_within
            data["Total Talk Time_sec"] += talk_time_sec
            data["Total Number Of Tagging"] = total_tagged
            # data["RL"] += rl

            if handled > 0:
                data["AHT_total"] += total_acht_row


            # all_data[campaign] = {
            #     "Client Name": campaign,
            #     "Offered": offered,
            #     "Handled": handled,
            #     "SL% (20 Sec)": f"{round(wi_thin_sla*100/handled) if handled else 0}%",
            #     "AL": f"{round(handled*100/offered) if offered else 0}%",
            #     "Calls Ans (20 Sec)": wi_thin_sla,
            #     # "Calls Ans (10 Sec)": wi_thin_sla_ten,
            #     "Total Calls Abandoned": abandon,
            #     "Abnd Within (20)": abnd_within,
            #     "Total Talk Time": seconds_to_hms(talk_time_sec),
            #     "Average Aband Time": "",
            #     # "SL% (10 Sec)": f"{round(wi_thin_sla_ten*100/handled) if handled else 0}%",               
            #     "AHT (In Sec)": round(total_acht_row / handled) if handled else 0,
            #     # "Call Rate": rate_info["InboundCallCharge"],
            #     # "Amount": round(abandon * float(rate_info["rate_per_sec"]) * round((total_acht_row / handled if handled else 0)), 2),
            #     "RL": rl,
            #     "RL%": f"{rl_percent}%"
            # }

            # Accumulate totals safely as float
            total_handle += total_acht_row
            total_acht += total_acht_row
            total_answered += handled
            total_offered += offered

    if req.filter_type == "without_0":
        all_data = {
            k: v for k, v in all_data.items()
            if v["Offered"] > 0
        }



    for company_name, data in all_data.items():
        handled = data["Handled"]
        offered = data["Offered"]

        data["Total Talk Time"] = seconds_to_hms(data["Total Talk Time_sec"])

        data["SL% (20 Sec)"] = f"{round(data['Calls Ans (20 Sec)'] * 100 / handled) if handled else 0}%"
        data["AL"] = f"{round(handled * 100 / offered) if offered else 0}%"

        data["AHT (In Sec)"] = round(data["AHT_total"] / handled) if handled else 0

        denominator = handled + data["Total Calls Abandoned"]
        data["RL%"] = f"{round((handled + data['RL']) * 100 / denominator) if denominator else 0}%"


    # -------------------- Generate Excel --------------------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SLA Clientwise Report"

    # headers = [
    #     "Client Name", "Offered", "Handled", "Calls Ans (20 Sec)", "Calls Ans (10 Sec)",
    #     "Total Calls Abandoned", "Abnd Within (20)", "Average Aband Time", "Total Talk Time", "SL% (20 Sec)", 
    #     "SL% (10 Sec)", "AL", "AHT (In Sec)", "Call Rate", "Amount", "RL", "RL%"
    # ]

    headers = [
        "Client Name", "Offered", "Handled", "SL% (20 Sec)", "AL", "Calls Ans (20 Sec)",
        "Total Calls Abandoned", "Abnd Within (20)", "Total Talk Time",  
        "AHT (In Sec)", "RL", "RL%", "Total Number Of Tagging"
    ]
    fill = PatternFill(start_color="317EAC", end_color="317EAC", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for col_num, header in enumerate(headers,1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    last_data_row = 1

    sorted_data = dict(sorted(all_data.items(), key=lambda x: x[0]))

    for row_num, (company, data) in enumerate(sorted_data.items(), start=2):
        last_data_row = row_num        

        for col_num, key in enumerate(headers, start=1):
            ws.cell(row=row_num, column=col_num, value=data.get(key, ""))
        
        # ---- Accumulate totals ----
        offered = data["Offered"]
        handled = data["Handled"]

        grand_totals["Offered"] += offered
        grand_totals["Handled"] += handled

        grand_totals["Total Talk Time_sec"] = grand_totals.get("Total Talk Time_sec", 0)
        grand_totals["Total Talk Time_sec"] += timedelta_to_seconds(data["Total Talk Time"])

        grand_totals["Calls Ans (20 Sec)"] += data["Calls Ans (20 Sec)"]
        # grand_totals["Calls Ans (10 Sec)"] += data["Calls Ans (10 Sec)"]
        grand_totals["Total Calls Abandoned"] += data["Total Calls Abandoned"]
        grand_totals["Abnd Within (20)"] += data["Abnd Within (20)"]
        # grand_totals["Average Aband Time"] = 0
        # grand_totals["Call Rate"] += data["Call Rate"]
        # grand_totals["Amount"] += data["Amount"]
        grand_totals["RL"] += data["RL"]
        grand_totals["Total Number Of Tagging"] += data["Total Number Of Tagging"]

        if handled > 0:
            grand_totals["AHT_total_sec"] += handled * data["AHT (In Sec)"]


    # -------- GRAND TOTAL RL% --------
    denominator = (
        grand_totals["Handled"] +
        grand_totals["Total Calls Abandoned"]
    )

    if denominator > 0:
        grand_totals["RL%"] = round(
            ((grand_totals["Handled"] + grand_totals["RL"]) / denominator) * 100)
    else:
        grand_totals["RL%"] = 0


    # -------- GRAND TOTAL SL% (20 Sec) and SL% (10 Sec) --------
    total_handled = grand_totals["Handled"]

    if total_handle > 0:
        grand_totals["SL% (20 Sec)"] = round(
            ((grand_totals["Calls Ans (20 Sec)"]*100)/total_handled)
            )
        grand_totals["SL% (10 Sec)"] = round(
            ((grand_totals["Calls Ans (10 Sec)"]*100)/total_handled)
            )
    else:
        grand_totals["SL% (20 Sec)"] = 0
        grand_totals["SL% (10 Sec)"] = 0

    # -------- GRAND TOTAL AL --------
    if total_offered > 0:
        grand_totals["AL"] = round(
            ((grand_totals["Handled"]*100)/grand_totals["Offered"]))
    else:
        grand_totals["AL"] = 0


    # Auto-width
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2


    grand_row = last_data_row + 1

    ws.cell(row=grand_row, column=1, value="Grand Total").font = Font(bold=True)

    col_map = {h: i + 1 for i, h in enumerate(headers)}

    ws.cell(row=grand_row, column=col_map["Offered"], value=grand_totals["Offered"])
    ws.cell(row=grand_row, column=col_map["Handled"], value=grand_totals["Handled"])
    ws.cell(row=grand_row, column=col_map["SL% (20 Sec)"], value=f'{grand_totals["SL% (20 Sec)"]}%')
    ws.cell(row=grand_row, column=col_map["AL"], value=f'{grand_totals["AL"]}%')
    ws.cell(row=grand_row, column=col_map["Total Talk Time"], value=seconds_to_hms(grand_totals["Total Talk Time_sec"]))
    ws.cell(row=grand_row, column=col_map["Calls Ans (20 Sec)"], value=grand_totals["Calls Ans (20 Sec)"])
    # ws.cell(row=grand_row, column=col_map["Calls Ans (10 Sec)"], value=grand_totals["Calls Ans (10 Sec)"])
    ws.cell(row=grand_row, column=col_map["Total Calls Abandoned"], value=grand_totals["Total Calls Abandoned"])
    ws.cell(row=grand_row, column=col_map["Abnd Within (20)"], value=grand_totals["Abnd Within (20)"])
    # ws.cell(row=grand_row, column=col_map["Average Aband Time"], value=grand_totals["Average Aband Time"])
    # ws.cell(row=grand_row, column=col_map["Call Rate"], value=grand_totals["Call Rate"])
    # ws.cell(row=grand_row, column=col_map["Amount"], value=round(grand_totals["Amount"], 2))
    ws.cell(row=grand_row, column=col_map["RL"], value=grand_totals["RL"])
    ws.cell(row=grand_row, column=col_map["RL%"], value=f'{grand_totals["RL%"]}%')
    ws.cell(row=grand_row,column=col_map["Total Number Of Tagging"],value=grand_totals["Total Number Of Tagging"])
    
    # ws.cell(row=grand_row, column=col_map["SL% (10 Sec)"], value=f'{grand_totals["SL% (10 Sec)"]}%')
   


    # Weighted AHT
    if grand_totals["Handled"] > 0:
        ws.cell(
            row=grand_row,
            column=col_map["AHT (In Sec)"],
            value=round(grand_totals["AHT_total_sec"] / grand_totals["Handled"])
        )


    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"sla_clientwise_report_{req.from_date}_{req.to_date}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )