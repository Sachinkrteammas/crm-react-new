from typing import Optional

# SLA Report For hours..//
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from database import get_db2
from database import get_db4, get_engine4
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from datetime import datetime, timedelta
from utils.email_manager import send_sla_report_email, EMAIL_RECEIVER


router = APIRouter(prefix="/sla", tags=["SLA Reports"])


# ---------------- Reusable Excel generator for mail have sent----------------
def generate_sla_excel_bytes(start_date: str, end_date: str, campaign_ids: list, db: Session):
    campaigns = [str(c).strip() for c in campaign_ids if str(c).strip()]
    if not campaigns:
        raise HTTPException(status_code=400, detail="No valid campaign IDs provided")

    placeholders = ", ".join([f":c{i}" for i in range(len(campaigns))])
    sql = text(f"""
        SELECT
            DATE_FORMAT(t2.call_date, '%Y-%m-%d %H:00:00') AS timeslot,
            COUNT(*) AS total_calls,
            SUM(CASE WHEN t2.user <> 'VDCL' THEN 1 ELSE 0 END) AS answered,
            COUNT(DISTINCT CASE WHEN t2.user <> 'VDCL' THEN t2.user END) AS manpower,
            ROUND((SUM(CASE WHEN t2.user <> 'VDCL' THEN 1 END) / COUNT(*)) * 100, 2) AS al_percent,
            ROUND(
                (SUM(CASE WHEN t2.user <> 'VDCL' AND t2.queue_seconds <= 20 THEN 1 END) /
                NULLIF(SUM(CASE WHEN t2.user <> 'VDCL' THEN 1 END),0)) * 100, 2
            ) AS sl_percent
        FROM asterisk.vicidial_closer_log t2
        WHERE 
            t2.call_date BETWEEN CONCAT(:start_date,' 00:00:00') AND CONCAT(:end_date,' 23:59:59')
            AND t2.campaign_id IN ({placeholders})
        GROUP BY timeslot
        ORDER BY timeslot
    """)

    params = {"start_date": start_date, "end_date": end_date}
    params.update({f"c{i}": c for i, c in enumerate(campaigns)})
    data = db.execute(sql, params).mappings().all()

    # ---------- Generate Excel ----------
    wb = Workbook()
    ws = wb.active
    ws.title = "SLA Report"

    ws["A1"] = "SLA Report : AL/SL/RL"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Date Range: {start_date} to {end_date}"
    ws["A2"].font = Font(size=12)

    headers = ["Date", "Hour", "Total Calls", "Answered", "Manpower", "AL %", "SL %"]
    ws.append([])
    ws.append(headers)
    header_row = ws.max_row

    fill = PatternFill("solid", fgColor="3A84F7")
    font_white = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")
    for col in range(1, len(headers)+1):
        cell = ws.cell(header_row, col)
        cell.fill = fill
        cell.font = font_white
        cell.alignment = center

    total_calls = total_answered = total_manpower = 0
    sl_values = []

    # ---------- Data Rows ----------
    for row in data:
        ts = row["timeslot"]
        hour = int(ts[11:13])
        if hour < 9 or hour > 20:
            continue

        date = ts[:10]
        tc = int(row.get("total_calls", 0))
        ans = int(row.get("answered", 0))
        mp = int(row.get("manpower", 0))
        al = float(row.get("al_percent", 0) or 0) / 100
        sl = float(row.get("sl_percent", 0) or 0) / 100

        ws.append([date, hour, tc, ans, mp, al, sl])
        total_calls += tc
        total_answered += ans
        total_manpower += mp
        sl_values.append(sl)

    # ---------- TOTAL ROW ----------
    tr = ws.max_row + 1
    ws[f"A{tr}"] = "TOTAL"
    ws[f"C{tr}"] = total_calls
    ws[f"D{tr}"] = total_answered
    ws[f"E{tr}"] = total_manpower
    ws[f"F{tr}"] = (total_answered / total_calls if total_calls else 0)
    ws[f"G{tr}"] = (sum(sl_values) / len(sl_values) if sl_values else 0)
    for col in range(1, len(headers)+1):
        ws.cell(tr, col).font = Font(bold=True)

    # ---------- Format percentages ----------
    for row_cells in ws.iter_rows(min_row=header_row+1, min_col=6, max_col=7):
        for cell in row_cells:
            cell.number_format = "0.00%"

    # ---------- Auto column width ----------
    thin = Side(border_style="thin", color="000000")
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    # ---------- Borders ----------
    for r in ws[f"A{header_row}:G{ws.max_row}"]:
        for cell in r:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # ---------- Save Excel to BytesIO ----------
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


# ---------------- JSON API for table ----------------
@router.get("/agents")
def get_sla_agents(start_date: str, end_date: str, campaign_ids: str, db: Session = Depends(get_db2)):
    campaigns = campaign_ids.split(",")

    placeholders = ", ".join([f":c{i}" for i in range(len(campaigns))])

    sql = text(f"""
        SELECT
            DATE_FORMAT(t2.call_date, '%Y-%m-%d %H:00:00') AS TimeSlot,
            COUNT(*) AS TotalCalls,
            SUM(CASE WHEN t2.user <> 'VDCL' THEN 1 ELSE 0 END) AS Answered,
            SUM(CASE WHEN t2.user = 'VDCL' THEN 1 ELSE 0 END) AS Abandon,
            SUM(CASE WHEN t2.user <> 'VDCL' AND t2.queue_seconds <= 20 THEN 1 ELSE 0 END) AS SLA_Calls,
            COUNT(DISTINCT CASE WHEN t2.user <> 'VDCL' THEN t2.user END) AS Manpower,
            ROUND(
                (SUM(CASE WHEN t2.user <> 'VDCL' THEN 1 END) / COUNT(*)) * 100, 2
            ) AS AL_Percentage,
            ROUND(
                (SUM(CASE WHEN t2.user <> 'VDCL' AND t2.queue_seconds <= 20 THEN 1 END) /
                NULLIF(SUM(CASE WHEN t2.user <> 'VDCL' THEN 1 END),0)) * 100, 2
            ) AS SL_Percentage
        FROM asterisk.vicidial_closer_log t2
        WHERE 
            t2.call_date BETWEEN CONCAT(:start_date,' 00:00:00')
            AND CONCAT(:end_date,' 23:59:59')
            AND t2.campaign_id IN ({placeholders})
        GROUP BY TimeSlot
        ORDER BY TimeSlot
    """)

    params = {"start_date": start_date, "end_date": end_date}
    params.update({f"c{i}": c for i, c in enumerate(campaigns)})

    rows = db.execute(sql, params).mappings().all()
    return {"data": rows}



# ---------------- Excel export ----------------
@router.get("/agents/export")
def export_sla_excel(start_date: str, end_date: str, campaign_ids: str, db: Session = Depends(get_db2)):
    # --- Keep campaign IDs as strings ---
    campaigns = [c.strip() for c in campaign_ids.split(",") if c.strip()]
    if not campaigns:
        return {"error": "No valid campaign IDs provided"}

    placeholders = ", ".join([f":c{i}" for i in range(len(campaigns))])

    sql = text(f"""
        SELECT
            DATE_FORMAT(t2.call_date, '%Y-%m-%d %H:00:00') AS timeslot,
            COUNT(*) AS total_calls,
            SUM(CASE WHEN t2.user <> 'VDCL' THEN 1 ELSE 0 END) AS answered,
            COUNT(DISTINCT CASE WHEN t2.user <> 'VDCL' THEN t2.user END) AS manpower,
            ROUND((SUM(CASE WHEN t2.user <> 'VDCL' THEN 1 END) / COUNT(*)) * 100, 2) AS al_percent,
            ROUND(
                (SUM(CASE WHEN t2.user <> 'VDCL' AND t2.queue_seconds <= 20 THEN 1 END) /
                NULLIF(SUM(CASE WHEN t2.user <> 'VDCL' THEN 1 END),0)) * 100, 2
            ) AS sl_percent
        FROM asterisk.vicidial_closer_log t2
        WHERE 
            t2.call_date BETWEEN CONCAT(:start_date,' 00:00:00')
                            AND CONCAT(:end_date,' 23:59:59')
            AND t2.campaign_id IN ({placeholders})
        GROUP BY timeslot
        ORDER BY timeslot
    """)

    params = {"start_date": start_date, "end_date": end_date}
    params.update({f"c{i}": c for i, c in enumerate(campaigns)})

    data = db.execute(sql, params).mappings().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "SLA Report"

    # ---------- Header ----------
    ws["A1"] = "SLA Report : AL/SL/RL"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Date Range: {start_date} to {end_date}"
    ws["A2"].font = Font(size=12)

    headers = ["Date", "Hour", "Total Calls", "Answered", "Manpower", "AL %", "SL %"]
    ws.append([])
    ws.append(headers)
    header_row = ws.max_row

    fill = PatternFill("solid", fgColor="3A84F7")
    font_white = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for col in range(1, len(headers)+1):
        cell = ws.cell(header_row, col)
        cell.fill = fill
        cell.font = font_white
        cell.alignment = center

    total_calls = total_answered = total_manpower = 0
    sl_values = []

    # ---------- Data Rows ----------
    for row in data:
        ts = row["timeslot"]
        hour = int(ts[11:13])
        if hour < 9 or hour > 20:
            continue

        date = ts[:10]
        tc = int(row.get("total_calls", 0))
        ans = int(row.get("answered", 0))
        mp = int(row.get("manpower", 0))
        al = float(row.get("al_percent", 0) or 0) / 100
        sl = float(row.get("sl_percent", 0) or 0) / 100

        ws.append([date, hour, tc, ans, mp, al, sl])

        total_calls += tc
        total_answered += ans
        total_manpower += mp
        sl_values.append(sl)

    # ---------- TOTAL ROW ----------
    tr = ws.max_row + 1
    ws[f"A{tr}"] = "TOTAL"
    ws[f"C{tr}"] = total_calls
    ws[f"D{tr}"] = total_answered
    ws[f"E{tr}"] = total_manpower
    ws[f"F{tr}"] = (total_answered / total_calls if total_calls else 0)
    ws[f"G{tr}"] = (sum(sl_values) / len(sl_values) if sl_values else 0)

    for col in range(1, len(headers)+1):
        ws.cell(tr, col).font = Font(bold=True)

    # ---------- Format percentages ----------
    for row_cells in ws.iter_rows(min_row=header_row+1, min_col=6, max_col=7):
        for cell in row_cells:
            cell.number_format = "0.00%"

    # ---------- Auto column width ----------
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    # ---------- Borders ----------
    thin = Side(border_style="thin", color="000000")
    for r in ws[f"A{header_row}:G{ws.max_row}"]:
        for cell in r:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # ---------- Return Excel ----------
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sla_report.xlsx"}
    )


# ---------------- Send Data in Email ----------------
@router.get("/agents/email", tags=["SLA Reports"])
def send_sla_report(
    start_date: str,
    end_date: str,
    campaign_ids: str,
    db: Session = Depends(get_db2)
):
    campaigns = [c.strip() for c in campaign_ids.split(",") if c.strip()]
    if not campaigns:
        raise HTTPException(status_code=400, detail="No valid campaign IDs provided")

    excel_stream = generate_sla_excel_bytes(start_date, end_date, campaigns, db)

    send_sla_report_email(excel_stream)

    return {"status": "success", "email_to": EMAIL_RECEIVER}



@router.get("/export-sla-day-wise")
def export_sla_day_wise(
    startdate: str = Query(...),
    enddate: str = Query(...),
    clientID: str = Query(...),
    sd_type: Optional[str] = Query("All"),
    db1: Session = Depends(get_db4),
    db2: Session = Depends(get_db2),
):

    data = {}

    # ---------------- Agents master ----------------
    ag_rows = db2.execute(
        text("SELECT user, full_name FROM vicidial_users WHERE active='Y'")
    ).fetchall()

    ag_list = {r.user: r.full_name for r in ag_rows}

    # ---------------- SD filter ----------------
    sd_str = ""
    sd_str2 = ""

    if sd_type == "Shared":
        sd_str = " AND is_shared='1'"
        sd_str2 = " AND vu.user_nickname='1'"
    elif sd_type == "Dedicated":
        sd_str = " AND is_shared='0'"
        sd_str2 = " AND vu.user_nickname='0'"

    # ---------------- Campaign & Client filter ----------------
    if clientID == "All":
        db1.execute(text("SET SESSION group_concat_max_len = 20000"))

        camp = db1.execute(
            text(f"""
            SELECT GROUP_CONCAT(campaignid) campaignid
            FROM registration_master
            WHERE status='A' AND is_dd_client='1' {sd_str}
            """)
        ).scalar()

        cl = db1.execute(
            text(f"""
            SELECT GROUP_CONCAT(company_id) company_id
            FROM registration_master
            WHERE status='A' AND is_dd_client='1' {sd_str}
            """)
        ).scalar()

        campaignId = f"t2.campaign_id IN ({camp})"
        client_list_str = f"ClientId IN ({cl})"

    else:
        camp = db1.execute(
            text(f"""
            SELECT campaignid FROM registration_master
            WHERE company_id='{clientID}' {sd_str}
            """)
        ).scalar()

        campaignId = f"t2.campaign_id IN ({camp})"
        client_list_str = f"ClientId IN ({clientID})"

    # ---------------- Date handling ----------------
    FromDate = datetime.strptime(startdate + " 00:00:00", "%Y-%m-%d %H:%M:%S")
    ToDate = datetime.strptime(enddate + " 23:59:59", "%Y-%m-%d %H:%M:%S")

    # ---------------- Daily Loop ----------------
    while FromDate < ToDate:

        start_time_start = FromDate
        start_time_end = FromDate + timedelta(days=1)
        FromDate = start_time_end

        dateLabel = start_time_start.strftime("%d-%m-%Y")

        qry = f"""
        SELECT 
            t2.user,
            t2.queue_seconds,
            vu.user_nickname,
            t1.talk_sec,
            t1.wait_sec,
            t1.dispo_sec,
            t1.pause_sec,
            IFNULL(t3.p,0) as hold

        FROM vicidial_closer_log t2
        LEFT JOIN vicidial_users vu ON t2.user=vu.user
        LEFT JOIN vicidial_agent_log t1 ON t1.uniqueid=t2.uniqueid AND t2.user=t1.user
        LEFT JOIN (
            SELECT uniqueid,SUM(parked_sec) p
            FROM park_log
            WHERE STATUS='GRABBED'
              AND parked_time>='{start_time_start}'
              AND parked_time<'{start_time_end}'
            GROUP BY uniqueid
        ) t3 ON t1.uniqueid=t3.uniqueid

        WHERE t2.call_date>='{start_time_start}'
          AND t2.call_date<'{start_time_end}'
          AND {campaignId}
        """

        row = db2.execute(text(qry)).fetchall()

        Total = 0
        WithinSLA = 0
        Answered = 0

        agents = set()
        shared_agents = set()
        dedicated_agents = set()
        other_agents = set()

        Talk = wait = dispo = pause = hold = 0

        # ---------------- Agent Name Mapping ----------------
        def map_agents(csv):
            if not csv:
                return ""
            return ",".join(
                [f"{ag_list.get(a,a)}({a})" for a in csv.split(",") if a]
            )
        
        for r in row:
            Total += 1

            user = r.user
            queue = r.queue_seconds or 0
            nickname = r.user_nickname

            talk_sec = r.talk_sec or 0
            wait_sec = r.wait_sec or 0
            dispo_sec = r.dispo_sec or 0
            pause_sec = r.pause_sec or 0
            hold_sec = r.hold or 0

            if user != 'VDCL':
                Answered += 1
                agents.add(user)

                if queue <= 20:
                    WithinSLA += 1

                if nickname == '1':
                    shared_agents.add(user)
                elif nickname == '0':
                    dedicated_agents.add(user)
                else:
                    other_agents.add(user)

            Talk += talk_sec
            wait += wait_sec
            dispo += dispo_sec
            pause += pause_sec
            hold += hold_sec

        
        Manpower = len(agents)
        Shared = len(shared_agents)
        Dedicated = len(dedicated_agents)
        Other = len(other_agents)

        Al = (Answered / Total * 100) if Total else 0

        Total_login = Talk + wait + dispo + pause + hold
        Net_login = Talk + wait + dispo + hold

        Utilization = (
            (Talk + dispo + hold) / Net_login * 100
            if Net_login else 0
        )


        agents_str = ",".join(agents)
        shared_str = ",".join(shared_agents)
        dedicated_str = ",".join(dedicated_agents)
        other_str = ",".join(other_agents)

        # ---------------- RL ----------------
        rl = db1.execute(
            text(f"""
            SELECT COUNT(1) cnt
            FROM aband_call_master
            WHERE {client_list_str}
              AND call_status='answer'
              AND calldate>='{start_time_start}'
              AND calldate<'{start_time_end}'
            """)
        ).scalar()

        # ---------------- Final Data ----------------
        # data[dateLabel] = {
        #     "Total": row.Total,
        #     "Answered": row.Answered,
        #     "Manpower": row.Manpower,
        #     "Shared": row.Shared,
        #     "Dedicated": row.Dedicated,
        #     "Other": row.Other,
        #     "Talk": row.Talk,
        #     "wait": row.wait,
        #     "dispo": row.dispo,
        #     "hold": row.hold,
        #     "Al %": round(row.Al, 2),
        #     "SL %": round((row.WIthinSLA / row.Answered) * 100, 2) if row.Answered else 0,
        #     "RL": rl,
        #     "RL %": round(((rl + row.Answered) / row.Total) * 100, 2) if row.Total else 0,
        #     "Total login": row.Total_login,
        #     "Net login": row.Net_login,
        #     "Utilization %": round(row.Utilization, 2),
        #     "Manpower Agents": map_agents(row.agents),
        #     "Shared Agents": map_agents(row.Shared_ag),
        #     "Dedicated Agents": map_agents(row.Dedicated_ag),
        #     "Other Agents": map_agents(row.Other_ag),
        #     "WIthinSLA": row.WIthinSLA,
        # }


        data[dateLabel] = {
            "Total": Total,
            "Answered": Answered,
            "Manpower": Manpower,
            "Shared": Shared,
            "Dedicated": Dedicated,
            "Other": Other,
            "Talk": Talk,
            "wait": wait,
            "dispo": dispo,
            "hold": hold,
            "Al %": round(Al, 2),
            "SL %": round((WithinSLA / Answered) * 100, 2) if Answered else 0,
            "RL": rl,
            "RL %": round(((rl + Answered) / Total) * 100, 2) if Total else 0,
            "Total login": Total_login,
            "Net login": Net_login,
            "Utilization %": round(Utilization, 2),

            "Manpower Agents": map_agents(agents_str),
            "Shared Agents": map_agents(shared_str),
            "Dedicated Agents": map_agents(dedicated_str),
            "Other Agents": map_agents(other_str),

            "WIthinSLA": WithinSLA,
        }

    return data


def calculate_manpower(login_time, logout_time, slot_start, slot_end):
    """
    Returns manpower contribution for one slot.
    Example:
        60 min = 1.0
        30 min = 0.5
        15 min = 0.25
    """

    actual_start = max(login_time, slot_start)
    actual_end = min(logout_time, slot_end)

    if actual_end <= actual_start:
        return 0

    active_seconds = (actual_end - actual_start).total_seconds()

    return active_seconds / 3600

@router.get("/slot-wise-utilization")
def slot_wise_utilization(
    startdate: str = Query(...),
    enddate: str = Query(...),
    clientID: str = Query(...),
    sd_type: str = Query("All"),
    db1: Session = Depends(get_db4),   # registration_master
    db2: Session = Depends(get_db2),   # vicidial
):
    data = {}
    datetimeArray = {}
    datearray = []
    timearray = []

    # ------------------ AGENT MASTER ------------------
    ag_sql = text("""
        SELECT user, full_name
        FROM vicidial_users
        WHERE active='Y'
    """)
    ag_rows = db2.execute(ag_sql).fetchall()
    ag_list = {r.user: r.full_name for r in ag_rows}

    # ------------------ DATE SETUP ------------------
    from_dt = datetime.strptime(startdate, "%Y-%m-%d")
    to_dt = datetime.strptime(enddate, "%Y-%m-%d") + timedelta(days=1)

    # ------------------ SD TYPE ------------------
    sd_str = ""
    if sd_type == "Shared":
        sd_str = " AND is_shared='1'"
    elif sd_type == "Dedicated":
        sd_str = " AND is_shared='0'"

    # ------------------ CLIENT / CAMPAIGN ------------------
    if clientID == "All":
        db1.execute(text("SET SESSION group_concat_max_len = 20000"))

        camp_sql = text(f"""
            SELECT GROUP_CONCAT(campaignid) AS campaign_id
            FROM registration_master
            WHERE status='A' AND is_dd_client='1' {sd_str}
        """)
        campaignId = db1.execute(camp_sql).fetchone().campaign_id
        campaign_cond = f"t2.campaign_id IN ({campaignId})" if campaignId else "1=0"

        cli_sql = text(f"""
            SELECT GROUP_CONCAT(company_id) AS company_id
            FROM registration_master
            WHERE status='A' AND is_dd_client='1' {sd_str}
        """)
        client_list = db1.execute(cli_sql).fetchone().company_id
        client_list_cond = f"ClientId IN ({client_list})" if client_list else "1=0"

    else:
        camp_sql = text(f"""
            SELECT campaignid
            FROM registration_master
            WHERE company_id=:clientID {sd_str}
            LIMIT 1
        """)
        campaignId = db1.execute(camp_sql, {"clientID": clientID}).fetchone().campaignid
        campaign_cond = f"t2.campaign_id IN ({campaignId})"
        client_list_cond = f"ClientId IN ({clientID})"

    # ------------------ LOOP SLOT WISE (HOURLY) ------------------
    cur = from_dt
    while cur < to_dt:
        date_label = cur.strftime("%Y-%m-%d")
        time_label = int(cur.strftime("%H"))

        start_time = cur.strftime("%Y-%m-%d %H:%M:%S")
        end_time = (cur + timedelta(hours=1)).strftime("%Y-%m-%d %H:%M:%S")

        slot_start = datetime.strptime(start_time, "%Y-%m-%d %H:%M:%S")
        slot_end = datetime.strptime(end_time, "%Y-%m-%d %H:%M:%S")

        datearray.append(date_label)
        timearray.append(time_label)
        datetimeArray.setdefault(date_label, []).append(time_label)

        login_sql = text("""
        SELECT
            user,
            event,
            event_date
        FROM vicidial_user_log
        WHERE event_date < :slot_end
        AND event IN ('LOGIN','LOGOUT')
        ORDER BY user,event_date
        """)

        login_rows = db2.execute(
            login_sql,
            {
                "slot_start": start_time,
                "slot_end": end_time
            }
        ).fetchall()

        qry = text(f"""
            SELECT 
                t2.user,
                t2.queue_seconds,
                vu.user_nickname,
                t1.talk_sec,
                t1.wait_sec,
                t1.dispo_sec,
                t1.pause_sec,
                IFNULL(t3.p,0) AS hold
            FROM asterisk.vicidial_closer_log t2
            LEFT JOIN asterisk.vicidial_users vu ON t2.user=vu.user
            LEFT JOIN asterisk.vicidial_agent_log t1 ON t1.uniqueid=t2.uniqueid AND t2.user=t1.user
            LEFT JOIN (
                SELECT uniqueid,SUM(parked_sec) p
                FROM park_log
                WHERE STATUS='GRABBED'
                AND parked_time>='{start_time}'
                AND parked_time<'{end_time}'
                GROUP BY uniqueid
            ) t3 ON t1.uniqueid=t3.uniqueid
            WHERE t2.call_date>='{start_time}'
            AND t2.call_date<'{end_time}'
            AND {campaign_cond}
        """)

        rows = db2.execute(qry).fetchall()

        total = 0
        within_sla = 0
        answered = 0

        agents_set = set()
        shared_set = set()
        dedicated_set = set()
        other_set = set()

        talk = wait = dispo = pause = hold = 0

        for r in rows:
            user = (r.user or "")
            nickname = str(r.user_nickname) if r.user_nickname else ""

            total += 1

            if user != "VDCL":
                answered += 1
                agents_set.add(user)

                if (r.queue_seconds or 0) <= 20:
                    within_sla += 1

                if nickname == "1":
                    shared_set.add(user)
                elif nickname == "0":
                    dedicated_set.add(user)
                else:
                    other_set.add(user)

            talk += r.talk_sec or 0
            wait += r.wait_sec or 0
            dispo += r.dispo_sec or 0
            pause += r.pause_sec or 0
            hold += r.hold or 0

        # data.setdefault(date_label, {})[time_label] = {
        #     "Total": r.Total,
        #     "Answered": r.Answered,
        #     "Manpower": r.Manpower,
        #     "Shared": r.Shared,
        #     "Dedicated": r.Dedicated,
        #     "Other": r.Other,
        #     "Talk": r.Talk,
        #     "wait": r.wait,
        #     "dispo": r.dispo,
        #     "pause": r.pause,
        #     "hold": r.hold,
        #     "Al %": round(r.Al or 0, 2),
        #     "SL %": round((r.WIthinSLA / r.Answered) * 100, 2) if r.Answered else 0,
        #     "Total login": r.Total_login,
        #     "Net login": r.Net_login,
        #     "Utilization %": round(r.Utilization or 0, 2),
        #     "WIthinSLA": r.WIthinSLA,
        #     "Manpower Agents": ",".join([f"{ag_list.get(a)}({a})" for a in (r.agents or "").split(",") if a]),
        #     "Shared Agents": ",".join([f"{ag_list.get(a)}({a})" for a in (r.Shared_ag or "").split(",") if a]),
        #     "Dedicated Agents": ",".join([f"{ag_list.get(a)}({a})" for a in (r.Dedicated_ag or "").split(",") if a]),
        #     "Other Agents": ",".join([f"{ag_list.get(a)}({a})" for a in (r.Other_ag or "").split(",") if a]),
        # }

        manpower = 0
        agent_slot_time = {}

        from collections import defaultdict

        user_sessions = defaultdict(list)

        for row in login_rows:
            user_sessions[row.user].append(row)



        # for user, events in user_sessions.items():
            # For client-wise report, count only agents who handled calls
        active_agents = agents_set if clientID != "All" else user_sessions.keys()

        for user in active_agents:
            events = user_sessions.get(user, [])

            login_time = None
            worked_seconds = 0

            # Was the agent already logged in before this slot?
            for e in events:
                if e.event_date >= slot_start:
                    break

                if e.event.upper() == "LOGIN":
                    login_time = slot_start

                elif e.event.upper() == "LOGOUT":
                    login_time = None

            for e in events:

                if e.event_date < slot_start:
                    continue

                if e.event.upper() == "LOGIN":
                    login_time = e.event_date

                elif e.event.upper() == "LOGOUT" and login_time:

                    overlap_start = max(login_time, slot_start)
                    overlap_end = min(e.event_date, slot_end)

                    if overlap_end > overlap_start:
                        worked_seconds += (
                                overlap_end - overlap_start
                        ).total_seconds()

                    login_time = None

            if login_time:
                overlap_start = max(login_time, slot_start)
                overlap_end = slot_end

                if overlap_end > overlap_start:
                    worked_seconds += (
                            overlap_end - overlap_start
                    ).total_seconds()

            manpower += worked_seconds / 3600

            agent_slot_time[user] = {
                "worked": round(worked_seconds / 60),
                "slot": 60,
                "display": f"{round(worked_seconds / 60)}/60"
            }


        data.setdefault(date_label, {})[time_label] = {
            "Total": total,
            "Answered": answered,
            "Manpower": round(manpower,2),
            # "Shared": len(shared_set),
            # "Dedicated": len(dedicated_set),
            "Other": len(other_set),

            "Talk": talk,
            "wait": wait,
            "dispo": dispo,
            "pause": pause,
            "hold": hold,
            "ACHT": round(talk / answered, 2) if answered else 0,

            "Al %": round((answered / total * 100) if total else 0, 2),
            "SL %": round((within_sla / answered * 100) if answered else 0, 2),

            "Total login": talk + wait + dispo + pause + hold,
            "Net login": talk + wait + dispo + hold,

            "Utilization %": round(
                (talk / (talk + wait + dispo + hold) * 100)
                if (talk + wait + dispo + hold) else 0,
                2
            ),

            "WIthinSLA": within_sla,

            "Manpower Agents": ",".join(
                [
                    f"{ag_list.get(a)}({a}) {agent_slot_time.get(a, {}).get('display', '0/60')}"
                    for a in agents_set
                ]
            ),
            "Shared Agents": ",".join([f"{ag_list.get(a)}({a})" for a in shared_set]),
            "Dedicated Agents": ",".join([f"{ag_list.get(a)}({a})" for a in dedicated_set]),
            "Other Agents": ",".join([f"{ag_list.get(a)}({a})" for a in other_set]),
        }

        # --- RL and RL % calculation ---
        rl_sql = text(f"""
            SELECT COUNT(1) as cnt
            FROM aband_call_master
            WHERE {client_list_cond}
            AND call_status='answer'
            AND calldate >= '{start_time}'
            AND calldate < '{end_time}'
        """)
        rl_count = db1.execute(rl_sql).fetchone().cnt or 0
        rl_percent = round(((rl_count + answered) / total) * 100, 2) if total else 0

        data[date_label][time_label].update({
            "RL": rl_count,
            "RL %": rl_percent
        })


        cur += timedelta(hours=1)

    return {
        "data": data,
        "datearray": sorted(set(datearray)),
        "timearray": sorted(set(timearray)),
        "datetimeArray": datetimeArray
    }




def generate_rl_sl_excel(startdate, enddate, clientID, sd_type, db1, db2):

    # ---------------------------------
    # Call existing API logic
    # ---------------------------------
    result = slot_wise_utilization(
        startdate=startdate,
        enddate=enddate,
        clientID=clientID,
        sd_type=sd_type,
        db1=db1,
        db2=db2
    )

    jsonData = result.get("data", {})

    rows = []

    # ---------------------------------
    # JSON → ROWS
    # ---------------------------------
    for date, hours in jsonData.items():
        for hour, values in hours.items():

            rows.append([
                date,
                hour,
                values.get("Total", 0),
                values.get("Answered", 0),
                values.get("Manpower", 0),
                values.get("Shared", 0),
                values.get("Dedicated", 0),
                values.get("Other", 0),
                values.get("Talk", 0),
                values.get("wait", 0),
                values.get("dispo", 0),
                values.get("hold", 0),
                values.get("Al %", 0),
                values.get("SL %", 0),
                values.get("RL %", 0),
                values.get("RL", 0),
                values.get("Total login", 0),
                values.get("Net login", 0),
                values.get("Utilization %", 0),
                values.get("WIthinSLA", 0),
                values.get("Manpower Agents", ""),
                values.get("Shared Agents", ""),
                values.get("Dedicated Agents", ""),
                values.get("Other Agents", "")
            ])

    # ---------------------------------
    # Create Workbook
    # ---------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "RL SL Report"

    headers = [
        "Date","Hour","Total","Answered","Manpower","Shared","Dedicated","Other",
        "Talk","Wait","Dispo","Hold",
        "Al %","SL %","RL %","RL",
        "Total Login","Net Login","Utilization %",
        "Within SLA",
        "Manpower Agents","Shared Agents","Dedicated Agents","Other Agents"
    ]

    ws.append(headers)

    # ---------------------------------
    # Add Data Rows
    # ---------------------------------
    for r in rows:
        ws.append(r)

    # ---------------------------------
    # GRAND TOTAL
    # ---------------------------------
    if rows:

        percent_cols = [12, 13, 14, 18]  # Al %, SL %, RL %, Utilization %

        grand = ["Grand Total", ""]

        for i in range(2, 20):

            col_values = [float(r[i]) if r[i] else 0 for r in rows]

            if i in percent_cols:
                grand.append(round(sum(col_values) / len(col_values), 2))
            else:
                grand.append(round(sum(col_values), 2))

        # agent columns empty
        grand += ["", "", "", ""]

        ws.append(grand)

    # ---------------------------------
    # Auto Column Width
    # ---------------------------------
    for col in ws.columns:

        max_length = 0
        column = col[0].column_letter

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column].width = max_length + 2

    # ---------------------------------
    # Save to Memory
    # ---------------------------------
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output