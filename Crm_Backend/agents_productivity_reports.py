# agents_productivity_reports.py
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from database import get_db2
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO

router = APIRouter(prefix="/agents", tags=["Agent Productivity"])

##############################################
# 1️⃣ JSON API
##############################################
@router.get("/productivity")
def get_agents_productivity(start_date: str, end_date: str, db: Session = Depends(get_db2)):
    """
    Agent productivity by merged OB + IB calls
    """
    try:
        sql = f"""
            SELECT agent_name, calls_taken
            FROM (
                SELECT
                    COALESCE(vu.full_name, 'Unknown / No Agent') AS agent_name,
                    COUNT(DISTINCT merged.lead_id) AS calls_taken,
                    0 AS sort_order
                FROM (
                    -- OUTBOUND
                    SELECT val.user, val.lead_id
                    FROM vicidial_agent_log val
                    WHERE val.talk_sec > 0
                      AND DATE(val.event_time) BETWEEN '{start_date}' AND '{end_date}'

                    UNION ALL

                    -- INBOUND
                    SELECT vcl.user, vcl.lead_id
                    FROM vicidial_closer_log vcl
                    WHERE vcl.length_in_sec > 0
                      AND DATE(vcl.call_date) BETWEEN '{start_date}' AND '{end_date}'
                ) merged
                LEFT JOIN vicidial_users vu ON merged.user = vu.user
                GROUP BY vu.full_name

                UNION ALL

                SELECT
                    'TOTAL' AS agent_name,
                    COUNT(DISTINCT lead_id) AS calls_taken,
                    1 AS sort_order
                FROM (
                    SELECT lead_id
                    FROM vicidial_agent_log
                    WHERE talk_sec > 0
                      AND DATE(event_time) BETWEEN '{start_date}' AND '{end_date}'

                    UNION ALL

                    SELECT lead_id
                    FROM vicidial_closer_log
                    WHERE length_in_sec > 0
                      AND DATE(call_date) BETWEEN '{start_date}' AND '{end_date}'
                ) total
            ) final
            ORDER BY sort_order, agent_name;
        """

        rows = db.execute(text(sql)).mappings().all()
        return {"data": rows}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"DB Error: {str(e)}")





##############################################
# 2️⃣ Export Excel
##############################################
@router.get("/productivity/export")
def export_agents_productivity(start_date: str, end_date: str, db: Session = Depends(get_db2)):
    """
    Export merged OB/IB productivity to Excel
    """

    sql = f"""
        SELECT agent_name, calls_taken
        FROM (
            SELECT
                COALESCE(vu.full_name, 'Unknown / No Agent') AS agent_name,
                COUNT(DISTINCT merged.lead_id) AS calls_taken,
                0 AS sort_order
            FROM (
                SELECT user, lead_id
                FROM vicidial_agent_log
                WHERE talk_sec > 0
                  AND DATE(event_time) BETWEEN '{start_date}' AND '{end_date}'
                UNION ALL
                SELECT user, lead_id
                FROM vicidial_closer_log
                WHERE length_in_sec > 0
                  AND DATE(call_date) BETWEEN '{start_date}' AND '{end_date}'
            ) merged
            LEFT JOIN vicidial_users vu ON vu.user = merged.user
            GROUP BY vu.full_name

            UNION ALL
            SELECT 'TOTAL', COUNT(DISTINCT lead_id), 1
            FROM (
                SELECT lead_id
                FROM vicidial_agent_log
                WHERE talk_sec > 0
                  AND DATE(event_time) BETWEEN '{start_date}' AND '{end_date}'
                UNION ALL
                SELECT lead_id
                FROM vicidial_closer_log
                WHERE length_in_sec > 0
                  AND DATE(call_date) BETWEEN '{start_date}' AND '{end_date}'
            ) t
        ) X
        ORDER BY sort_order, agent_name;
    """

    rows = db.execute(text(sql)).mappings().all()

    # ========= EXCEL =============
    wb = Workbook()
    ws = wb.active
    ws.title = "Agent APR"

    ws["A1"] = "Report: Agent APR (OB + IB)"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Date Range: {start_date} to {end_date}"
    ws["A2"].font = Font(size=12)

    ws.append([])
    headers = ["Agent Name", "Calls Taken (OB + IB)"]
    ws.append(headers)

    # Header style
    fill = PatternFill("solid", fgColor="3A84F7")
    font_white = Font(color="FFFFFF", bold=True)
    for col in range(1, len(headers) + 1):
        c = ws.cell(4, col)
        c.fill = fill
        c.font = font_white
        c.alignment = Alignment(horizontal="center")

    # Add data
    for row in rows:
        ws.append([row["agent_name"], row["calls_taken"]])

    # Auto width
    for col in ws.columns:
        length = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = length + 4

    # Download
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Agent APR.xlsx"}
    )
