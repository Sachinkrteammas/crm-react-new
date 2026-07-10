
from fastapi import APIRouter, Depends, Response
from sqlalchemy.orm import Session
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO

from database import get_db4

router = APIRouter(prefix="/did-logs", tags=["DID Logs Reports"])


def auto_fit_columns(ws):
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[column_letter].width = max_length + 2


@router.get("/export-excel")
def export_did_logs_excel(db: Session = Depends(get_db4)):

    # query = text("""
    #     SELECT id, client_name, did_number, vendor_name, unique_id, call_time
    #     FROM did_logs
    #     ORDER BY id ASC
    # """)
    query = text("""
        SELECT d.id,d.client_name, d.did_number, d.vendor_name, d.unique_id, d.call_time

            FROM did_logs d

            INNER JOIN (

                SELECT MAX(id) AS id

                FROM did_logs

                GROUP BY client_name, did_number, vendor_name

            ) x ON d.id = x.id;
    """)

    rows = db.execute(query).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "DID Logs Report"

    # ==== Report Title ====
    ws["A1"] = "Report: DID Logs"
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append([])

    # ==== Table Headers ====
    headers = [
        "ID",
        "Client Name",
        "DID Number",
        "Vendor Name",
        "Unique ID",
        "Call Time",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    # ==== Insert Data Rows ====
    for row in rows:
        ws.append([
            row.id,
            row.client_name,
            row.did_number,
            row.vendor_name,
            row.unique_id,
            row.call_time.strftime("%Y-%m-%d %H:%M:%S") if row.call_time else "",
        ])

    # Auto-size columns
    auto_fit_columns(ws)

    # Save file to memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="DID_Logs_Report.xlsx"'
        }
    )
